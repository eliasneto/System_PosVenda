"""
Percorre a pasta EACE, extrai INEP e Valor Total da Nota de cada PDF
e gera um arquivo Excel com os resultados.

Estrutura esperada:
    EACE/
        {INEP_FOLDER}/
            KIT/
                *.pdf
            NOBREAK/
                *.pdf
"""

import re
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger

from config import EACE_DIR, OUTPUT_DIR


_RE_INEP = re.compile(r"INEP[:\s]+(\d{7,8})", re.IGNORECASE)
_RE_VALOR = re.compile(r"VALOR\s+TOTAL\s+DA\s+NOTA", re.IGNORECASE)
# Linha de dados do produto: [COD.PROD] [DESCRICAO] [NCM 0000.00.00] [resto...]
_RE_PRODUTO = re.compile(r"^\S+\s+(.+?)\s+\d{4}\.\d{2}\.\d{2}", re.MULTILINE)


def extrair_texto_pdf(caminho_pdf: Path) -> str:
    texto = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                conteudo = pagina.extract_text()
                if conteudo:
                    texto.append(conteudo)
    except Exception as e:
        logger.error(f"Erro ao ler PDF {caminho_pdf.name}: {e}")
    return "\n".join(texto)


def extrair_inep(texto: str) -> str:
    match = _RE_INEP.search(texto)
    return match.group(1) if match else ""


def extrair_produto(texto: str) -> str:
    match = _RE_PRODUTO.search(texto)
    return match.group(1).strip() if match else ""


def extrair_valor(texto: str) -> str:
    linhas = texto.splitlines()
    for i, linha in enumerate(linhas):
        if _RE_VALOR.search(linha):
            # O valor fica na proxima linha, como ultimo numero da sequencia
            if i + 1 < len(linhas):
                numeros = re.findall(r"[\d]+(?:[.,][\d]+)*", linhas[i + 1])
                if numeros:
                    return numeros[-1]
    return ""


def localizar_pdfs(pasta_eace: Path) -> list[dict]:
    """Retorna lista de dicts com pasta_nome, pdf_path para cada PDF encontrado."""
    registros = []
    if not pasta_eace.exists():
        logger.warning(f"Pasta EACE nao encontrada: {pasta_eace}")
        return registros

    for pasta_inep in sorted(pasta_eace.iterdir()):
        if not pasta_inep.is_dir():
            continue
        nome_pasta = pasta_inep.name
        for subpasta in sorted(pasta_inep.iterdir()):
            if not subpasta.is_dir():
                continue
            for pdf in sorted(subpasta.glob("*.pdf")):
                registros.append({"pasta": nome_pasta, "produto": subpasta.name.upper(), "pdf": pdf})

    return registros


def criar_excel(dados: list[dict], caminho_saida: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados NF"

    # Colunas: A=OSP B=Pasta C=INEP D=Produto E=Valor NF F=Valor Portal G=Status
    cabecalho = ["OSP", "Pasta", "INEP", "Produto", "Valor Total da Nota", "Valor Portal", "Status"]
    estilo_cab = Font(bold=True, color="FFFFFF")
    fill_cab = PatternFill(fill_type="solid", fgColor="2E4057")

    for col, titulo in enumerate(cabecalho, start=1):
        celula = ws.cell(row=1, column=col, value=titulo)
        celula.font = estilo_cab
        celula.fill = fill_cab
        celula.alignment = Alignment(horizontal="center")

    for linha, item in enumerate(dados, start=2):
        ws.cell(row=linha, column=1, value=item["osp"])
        ws.cell(row=linha, column=2, value=item["pasta"])
        ws.cell(row=linha, column=3, value=item["inep"])
        ws.cell(row=linha, column=4, value=item["produto"])
        ws.cell(row=linha, column=5, value=item["valor"])

    for col in ws.columns:
        largura = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = largura

    try:
        wb.save(caminho_saida)
    except PermissionError:
        raise PermissionError(
            f"Nao foi possivel salvar o Excel - feche o arquivo antes de rodar: {caminho_saida}"
        )
    logger.info(f"Excel salvo em: {caminho_saida}")


def _normalizar_valor(valor: str) -> float:
    try:
        return float(valor.strip().replace(".", "").replace(",", "."))
    except (ValueError, AttributeError):
        return -1.0


def valores_iguais(v1: str, v2: str) -> bool:
    n1, n2 = _normalizar_valor(v1), _normalizar_valor(v2)
    return n1 >= 0 and n2 >= 0 and abs(n1 - n2) < 0.01


def ler_valor_pdf(pasta: str, tipo_upper: str, caminho_excel: Path) -> str:
    """Le o valor da NF (coluna E) do Excel para a pasta e tipo informados."""
    if not caminho_excel.exists():
        return ""
    wb = openpyxl.load_workbook(caminho_excel, read_only=True)
    ws = wb.active
    for linha in range(2, ws.max_row + 1):
        pasta_cel = str(ws.cell(row=linha, column=2).value or "").strip()   # B=Pasta
        produto_cel = str(ws.cell(row=linha, column=4).value or "").upper() # D=Produto
        if pasta_cel == pasta and tipo_upper in produto_cel:
            wb.close()
            return str(ws.cell(row=linha, column=5).value or "")            # E=Valor NF
    wb.close()
    return ""


def atualizar_planilha_pos_portal(atualizacoes: list[dict], caminho_excel: Path) -> None:
    """Preenche 'Valor Portal' (E) e 'Status' (F) no Excel apos automacao no portal.

    Cada item de `atualizacoes`:
        {"inep": str, "tipo": str|None, "valor_portal": str, "status": str}

    Se `tipo` for None, o status e aplicado a TODAS as linhas do INEP.
    """
    if not caminho_excel.exists():
        logger.warning("Excel nao encontrado para atualizar.")
        return

    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active

    for linha in range(2, ws.max_row + 1):
        pasta_excel = str(ws.cell(row=linha, column=2).value or "").strip()   # B=Pasta
        produto_excel = str(ws.cell(row=linha, column=4).value or "").upper() # D=Produto

        for item in atualizacoes:
            if item["inep"] != pasta_excel:
                continue
            tipo = item.get("tipo")
            if tipo and tipo not in produto_excel:
                continue
            ws.cell(row=linha, column=6, value=item.get("valor_portal", ""))  # F=Valor Portal
            ws.cell(row=linha, column=7, value=item.get("status", ""))        # G=Status
            logger.info(
                f"  Planilha: Pasta={pasta_excel} | {produto_excel[:25]} → {item.get('status', '')}"
            )
            break

    try:
        wb.save(caminho_excel)
        logger.success("Planilha atualizada com status e valores do portal.")
    except PermissionError:
        logger.error("Nao foi possivel salvar o Excel - feche o arquivo antes de rodar.")


def ler_registros_excel(caminho_excel: Path) -> list[dict]:
    """Le todos os registros do Excel gerado.

    Retorna lista de dicts com: osp, pasta, inep, produto, valor.
    """
    if not caminho_excel.exists():
        logger.error("Excel nao encontrado: {}", caminho_excel)
        return []

    wb = openpyxl.load_workbook(caminho_excel, read_only=True)
    ws = wb.active
    registros = []
    for linha in range(2, ws.max_row + 1):
        pasta = str(ws.cell(row=linha, column=2).value or "").strip()
        inep  = str(ws.cell(row=linha, column=3).value or "").strip()
        if not pasta and not inep:
            continue
        registros.append({
            "osp":     str(ws.cell(row=linha, column=1).value or "").strip(),
            "pasta":   pasta,
            "inep":    inep,
            "produto": str(ws.cell(row=linha, column=4).value or "").strip(),
            "valor":   str(ws.cell(row=linha, column=5).value or "").strip(),
        })
    wb.close()
    return registros


def validar_pasta_vs_inep(registros: list[dict]) -> bool:
    """Valida que o INEP extraido da NF bate com o nome da pasta (coluna Pasta).

    Retorna True se todos passaram, False se houver qualquer divergencia.
    """
    ok = True
    for r in registros:
        if r["pasta"] != r["inep"]:
            logger.error(
                "INEP divergente: Pasta='{}' ≠ INEP na NF='{}' | Produto: {}",
                r["pasta"], r["inep"], r["produto"],
            )
            ok = False
    if ok:
        logger.success("Validacao Pasta x INEP OK - todos os registros conferem.")
    return ok


def main(osp: str = "", caminho_excel: Path = OUTPUT_DIR / "dados_notas_fiscais.xlsx") -> bool:
    logger.info(f"Varrendo PDFs em: {EACE_DIR}")
    registros_pdf = localizar_pdfs(EACE_DIR)

    if not registros_pdf:
        logger.error("Nenhum PDF encontrado. Verifique a estrutura da pasta EACE.")
        return False

    dados = []
    houve_erro = False
    for item in registros_pdf:
        pdf_path: Path = item["pdf"]
        logger.info(f"Processando: {pdf_path.relative_to(EACE_DIR.parent)}")
        texto = extrair_texto_pdf(pdf_path)

        inep = extrair_inep(texto)
        produto = extrair_produto(texto)
        valor = extrair_valor(texto)

        if not inep:
            logger.warning(f"  INEP nao encontrado em: {pdf_path.name}")
            houve_erro = True
        if not produto:
            logger.warning(f"  Produto nao encontrado em: {pdf_path.name}")
            houve_erro = True
        if not valor:
            logger.warning(f"  Valor total nao encontrado em: {pdf_path.name}")
            houve_erro = True

        dados.append({"osp": osp, "pasta": item["pasta"], "inep": inep, "produto": produto, "valor": valor})
        logger.success(f"  OSP={osp} | Pasta={item['pasta']} | INEP={inep} | Produto={produto} | Valor={valor}")

    try:
        criar_excel(dados, caminho_excel)
    except PermissionError as e:
        logger.error(str(e))
        return False
    logger.success(f"Planilha gerada. {len(dados)} registro(s) exportado(s).")
    return not houve_erro


if __name__ == "__main__":
    main()
