import datetime
import tempfile
from pathlib import Path

import openpyxl
from django.core.management import call_command
from django.test import TestCase

from apps.escolas.models import Escola

CABECALHO = [
    "LOTE", "UF ", "MUNICIPIO", "INEP", "UNIDADE ESCOLAR ",
    "ENDEREÇO UNIDADE ESCOLAR ", "VELOCIDADE", "KIT WIFI ESTIMADO",
]


def _criar_planilha(tmp_path, linhas, linha_cabecalho=13):
    """Monta um .xlsx no mesmo formato do CONSOLIDADO EACE.xlsx real: dados
    fora da aba comecam em branco, cabecalho na linha `linha_cabecalho` e
    colunas com espacos/acentos identicos ao arquivo de origem."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FATURAMENTO MATERIAIS"

    for indice, titulo in enumerate(CABECALHO, start=1):
        ws.cell(row=linha_cabecalho, column=indice, value=titulo)

    for offset, linha in enumerate(linhas, start=1):
        for indice, valor in enumerate(linha, start=1):
            ws.cell(row=linha_cabecalho + offset, column=indice, value=valor)

    caminho = tmp_path / "planilha_teste.xlsx"
    wb.save(caminho)
    return caminho


class ImportarEscolasPlanilhaTests(TestCase):
    """FEAT-002: importacao de Escola a partir da planilha CONSOLIDADO EACE.xlsx."""

    def setUp(self):
        self._tmp_dir = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp_dir.name)
        self.addCleanup(self._tmp_dir.cleanup)

    def test_importa_escolas_da_planilha(self):
        caminho = _criar_planilha(self.tmp_path, [
            [9, "SP", "Nova Aliança", 35006380, "ESCOLA MUNICIPAL VICENTE FERNANDES",
             "LOURENCO PALA, 276 CENTRO. 15210-000 Nova Aliança - SP.", 50, 2],
            [11, "DF", "Brasília", 53012089, "EC 02 DO RIACHO FUNDO",
             "QUADRA QN 5, 07 AREA ESPECIAL.", 300, 10],
        ])

        call_command("importar_escolas_planilha", str(caminho))

        self.assertEqual(Escola.objects.count(), 2)
        escola = Escola.objects.get(inep="35006380")
        self.assertEqual(escola.nome, "ESCOLA MUNICIPAL VICENTE FERNANDES")
        self.assertEqual(escola.lote, 9)
        self.assertEqual(escola.estado, "SP")
        self.assertEqual(escola.municipio, "Nova Aliança")
        self.assertEqual(escola.velocidade_dl_minima, "50")
        self.assertEqual(escola.kit_inicial, "2")
        self.assertEqual(escola.nobreak_inicial, "Nobreak")  # RN-017

    def test_inep_curto_e_preenchido_com_zeros_a_esquerda(self):
        caminho = _criar_planilha(self.tmp_path, [
            [1, "SP", "Teste", 123, "ESCOLA TESTE", "RUA TESTE, 1", 50, 1],
        ])

        call_command("importar_escolas_planilha", str(caminho))

        self.assertTrue(Escola.objects.filter(inep="00000123").exists())

    def test_linha_em_branco_no_rodape_e_ignorada(self):
        caminho = _criar_planilha(self.tmp_path, [
            [9, "SP", "Nova Aliança", 35006380, "ESCOLA MUNICIPAL VICENTE FERNANDES",
             "LOURENCO PALA, 276 CENTRO.", 50, 2],
            [None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None],
        ])

        call_command("importar_escolas_planilha", str(caminho))

        self.assertEqual(Escola.objects.count(), 1)

    def test_reimportar_nao_duplica_nem_sobrescreve_escola_existente(self):
        Escola.objects.create(inep="35006380", nome="NOME JA CADASTRADO MANUALMENTE")
        caminho = _criar_planilha(self.tmp_path, [
            [9, "SP", "Nova Aliança", 35006380, "ESCOLA MUNICIPAL VICENTE FERNANDES",
             "LOURENCO PALA, 276 CENTRO.", 50, 2],
        ])

        call_command("importar_escolas_planilha", str(caminho))
        call_command("importar_escolas_planilha", str(caminho))  # roda de novo, tem que ser idempotente

        self.assertEqual(Escola.objects.count(), 1)
        escola = Escola.objects.get(inep="35006380")
        self.assertEqual(escola.nome, "NOME JA CADASTRADO MANUALMENTE")

    def test_escola_nova_nasce_desconectada(self):
        caminho = _criar_planilha(self.tmp_path, [
            [9, "SP", "Nova Aliança", 35006380, "ESCOLA MUNICIPAL VICENTE FERNANDES",
             "LOURENCO PALA, 276 CENTRO.", 50, 2],
        ])

        call_command("importar_escolas_planilha", str(caminho))

        escola = Escola.objects.get(inep="35006380")
        self.assertEqual(escola.status_conexao, Escola.DESCONECTADO)

    def test_aba_inexistente_gera_erro_claro(self):
        wb = openpyxl.Workbook()
        wb.active.title = "OUTRA ABA"
        caminho = self.tmp_path / "planilha_sem_aba.xlsx"
        wb.save(caminho)

        with self.assertRaises(Exception):
            call_command("importar_escolas_planilha", str(caminho))


CABECALHO_NOVA_BASE = [
    "Fase", "UF", "Cidade", "Codigo Inep", "Nome da Escola", "Endereço",
    "Velocidade DL Mínima (Mbps)", "Kit Wi-Fi\n(estimado)",
]


def _criar_nova_base(tmp_path, linhas):
    """Monta um .xlsx no mesmo formato de "Nova BASE EACE.xlsx": aba
    "base", cabeçalho na linha 1, colunas com nomes/quebra de linha
    idênticos ao arquivo real."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "base"

    for indice, titulo in enumerate(CABECALHO_NOVA_BASE, start=1):
        ws.cell(row=1, column=indice, value=titulo)

    for offset, linha in enumerate(linhas, start=1):
        for indice, valor in enumerate(linha, start=1):
            ws.cell(row=1 + offset, column=indice, value=valor)

    caminho = tmp_path / "nova_base_teste.xlsx"
    wb.save(caminho)
    return caminho


class ImportarNovaBaseEaceTests(TestCase):
    """Correção pontual (2026-09-01): importação de Escola a partir de
    "Nova BASE EACE.xlsx" (aba "base") — formato de coluna diferente do
    CONSOLIDADO EACE.xlsx (`importar_escolas_planilha`), mesma regra de
    segurança (só cria INEP novo, nunca sobrescreve)."""

    def setUp(self):
        self._tmp_dir = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp_dir.name)
        self.addCleanup(self._tmp_dir.cleanup)

    def test_simulacao_nao_grava_nada(self):
        caminho = _criar_nova_base(self.tmp_path, [
            [11, "TO", "ARAGUAINA", 17057590, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE",
             "ERICO VERISSIMO, 271", 50, 2],
        ])

        call_command("importar_nova_base_eace", str(caminho))

        self.assertEqual(Escola.objects.count(), 0)

    def test_aplicar_cria_escolas_novas(self):
        caminho = _criar_nova_base(self.tmp_path, [
            [11, "TO", "ARAGUAINA", 17057590, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE",
             "ERICO VERISSIMO, 271", 50, 2],
            [9, "SP", "LORENA", 35243875, "MARIA ANTONIETA ARANTES FERREIRA PROFA ESCOLA MUNICIPAL",
             "LUIZA CHAGAS PROFESSORA, SN", 200, 7],
        ])

        call_command("importar_nova_base_eace", str(caminho), "--aplicar")

        self.assertEqual(Escola.objects.count(), 2)
        escola = Escola.objects.get(inep="17057590")
        self.assertEqual(escola.nome, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE")
        self.assertEqual(escola.lote, 11)
        self.assertEqual(escola.estado, "TO")
        self.assertEqual(escola.municipio, "ARAGUAINA")
        self.assertEqual(escola.endereco, "ERICO VERISSIMO, 271")
        self.assertEqual(escola.velocidade_dl_minima, "50")
        self.assertEqual(escola.kit_inicial, "2")
        self.assertEqual(escola.nobreak_inicial, "Nobreak")
        self.assertEqual(escola.status_conexao, Escola.DESCONECTADO)

    def test_nao_duplica_nem_sobrescreve_escola_existente(self):
        Escola.objects.create(inep="17057590", nome="NOME JA CADASTRADO MANUALMENTE")
        caminho = _criar_nova_base(self.tmp_path, [
            [11, "TO", "ARAGUAINA", 17057590, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE",
             "ERICO VERISSIMO, 271", 50, 2],
        ])

        call_command("importar_nova_base_eace", str(caminho), "--aplicar")
        call_command("importar_nova_base_eace", str(caminho), "--aplicar")  # idempotente

        self.assertEqual(Escola.objects.count(), 1)
        self.assertEqual(Escola.objects.get(inep="17057590").nome, "NOME JA CADASTRADO MANUALMENTE")

    def test_linha_duplicada_dentro_do_arquivo_conta_so_a_primeira(self):
        caminho = _criar_nova_base(self.tmp_path, [
            [11, "TO", "ARAGUAINA", 17057590, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE",
             "ERICO VERISSIMO, 271", 50, 2],
            [11, "TO", "ARAGUAINA", 17057590, "CENTRO DE EDUCACAO INFANTIL PEQUENO PRINCIPE",
             "ERICO VERISSIMO, 271", 50, 2],
        ])

        call_command("importar_nova_base_eace", str(caminho), "--aplicar")

        self.assertEqual(Escola.objects.count(), 1)

    def test_inep_curto_e_preenchido_com_zeros_a_esquerda(self):
        caminho = _criar_nova_base(self.tmp_path, [
            [1, "SP", "Teste", 123, "ESCOLA TESTE", "RUA TESTE, 1", 50, 1],
        ])

        call_command("importar_nova_base_eace", str(caminho), "--aplicar")

        self.assertTrue(Escola.objects.filter(inep="00000123").exists())

    def test_linha_sem_inep_e_ignorada(self):
        caminho = _criar_nova_base(self.tmp_path, [
            [11, "TO", "ARAGUAINA", None, "ESCOLA SEM INEP", "RUA TESTE, 1", 50, 1],
        ])

        call_command("importar_nova_base_eace", str(caminho), "--aplicar")

        self.assertEqual(Escola.objects.count(), 0)

    def test_aba_inexistente_gera_erro_claro(self):
        wb = openpyxl.Workbook()
        wb.active.title = "OUTRA ABA"
        caminho = self.tmp_path / "planilha_sem_aba.xlsx"
        wb.save(caminho)

        with self.assertRaises(Exception):
            call_command("importar_nova_base_eace", str(caminho))


class EscolaStatusConexaoTests(TestCase):
    """RN-007: status de conexao derivado do preenchimento das datas de
    instalacao RE/RI."""

    def test_nasce_desconectada(self):
        escola = Escola.objects.create(inep="11111111", nome="Escola A")
        self.assertEqual(escola.status_conexao, Escola.DESCONECTADO)

    def test_fica_parcialmente_conectada_com_apenas_um_processo(self):
        escola = Escola.objects.create(
            inep="22222222", nome="Escola B",
            data_instalacao_ri=datetime.date(2026, 8, 1),
        )
        self.assertEqual(escola.status_conexao, Escola.PARCIALMENTE_CONECTADO)

    def test_fica_conectada_com_os_dois_processos(self):
        escola = Escola.objects.create(
            inep="33333333", nome="Escola C",
            data_instalacao_re=datetime.date(2026, 8, 1),
            data_instalacao_ri=datetime.date(2026, 8, 2),
        )
        self.assertEqual(escola.status_conexao, Escola.CONECTADO)

    def test_volta_a_parcialmente_conectada_se_uma_data_for_removida(self):
        escola = Escola.objects.create(
            inep="44444444", nome="Escola D",
            data_instalacao_re=datetime.date(2026, 8, 1),
            data_instalacao_ri=datetime.date(2026, 8, 2),
        )
        escola.data_instalacao_ri = None
        escola.save()
        self.assertEqual(escola.status_conexao, Escola.PARCIALMENTE_CONECTADO)


class EscolaNobreakTests(TestCase):
    """RN-017: Nobreak declarado é item padrão, igual para toda escola —
    sem passo manual, tanto para escola já existente quanto para nova."""

    def test_escola_nova_nasce_com_nobreak_padrao(self):
        escola = Escola.objects.create(inep="55555555", nome="Escola E")
        self.assertEqual(escola.nobreak_inicial, "Nobreak")

    def test_nobreak_padrao_e_o_mesmo_para_qualquer_escola(self):
        escola_a = Escola.objects.create(inep="66666666", nome="Escola F")
        escola_b = Escola.objects.create(inep="77777777", nome="Escola G")
        self.assertEqual(escola_a.nobreak_inicial, escola_b.nobreak_inicial)
