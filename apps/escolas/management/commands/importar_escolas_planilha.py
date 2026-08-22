from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.escolas.models import Escola

try:
    import openpyxl
except ImportError:
    openpyxl = None


COLUNAS_OBRIGATORIAS = (
    "LOTE",
    "UF",
    "MUNICIPIO",
    "INEP",
    "UNIDADE ESCOLAR",
    "ENDEREÇO UNIDADE ESCOLAR",
    "VELOCIDADE",
    "KIT WIFI ESTIMADO",
)


def _normalizar_cabecalho(valor):
    return (str(valor) if valor is not None else "").strip().upper()


def _texto(valor):
    return str(valor).strip() if valor not in (None, "") else ""


class Command(BaseCommand):
    """FEAT-002/RF-01/RN-007: importa o cadastro de Escola a partir da
    planilha CONSOLIDADO EACE.xlsx (aba FATURAMENTO MATERIAIS), mesma fonte
    usada na migracao inicial do modulo-posVenda para o Sistema_posvenda. A
    origem e uma exportacao manual do sistema legado (EACE), nunca um dado
    fabricado por este comando.

    Repetir a importacao e seguro: escola cujo INEP ja existe no banco e
    apenas contada como existente — nao duplica, nao sobrescreve campo ja
    migrado. Escola nova sempre nasce com status_conexao=desconectado
    (RN-007), pois nao ha data de instalacao RE/RI na planilha.
    """

    help = (
        "Importa Escola (INEP, lote, UF, municipio, nome, endereco, velocidade, "
        "kit estimado) a partir do CONSOLIDADO EACE.xlsx, aba FATURAMENTO MATERIAIS."
    )

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho para o CONSOLIDADO EACE.xlsx.")
        parser.add_argument(
            "--aba", default="FATURAMENTO MATERIAIS",
            help="Nome da aba com os dados (padrao: FATURAMENTO MATERIAIS).",
        )
        parser.add_argument(
            "--linha-cabecalho", type=int, default=13,
            help="Numero da linha (1-based) com os titulos das colunas (padrao: 13).",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError(
                "Dependencia 'openpyxl' nao instalada. Adicione 'openpyxl' ao requirements.txt "
                "e reinstale as dependencias."
            )

        caminho = Path(options["arquivo"])
        if not caminho.exists():
            raise CommandError(f"Arquivo nao encontrado: {caminho}")

        try:
            planilha = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        except Exception as erro:
            raise CommandError(f"Nao foi possivel abrir '{caminho}': {erro}")

        aba = options["aba"]
        if aba not in planilha.sheetnames:
            raise CommandError(f"Aba '{aba}' nao encontrada. Abas disponiveis: {planilha.sheetnames}")

        planilha_aba = planilha[aba]
        linha_cabecalho = options["linha_cabecalho"]

        cabecalho = next(
            planilha_aba.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho, values_only=True), None
        )
        if cabecalho is None:
            raise CommandError(f"Linha de cabecalho {linha_cabecalho} vazia ou inexistente na aba '{aba}'.")

        indice_coluna = {_normalizar_cabecalho(valor): posicao for posicao, valor in enumerate(cabecalho)}
        faltando = [coluna for coluna in COLUNAS_OBRIGATORIAS if coluna not in indice_coluna]
        if faltando:
            raise CommandError(f"Colunas obrigatorias ausentes na planilha: {faltando}")

        criadas = 0
        existentes = 0
        ignoradas = 0

        linhas = planilha_aba.iter_rows(min_row=linha_cabecalho + 1, values_only=True)
        for numero_linha, linha in enumerate(linhas, start=linha_cabecalho + 1):
            inep_bruto = linha[indice_coluna["INEP"]]
            if inep_bruto in (None, ""):
                continue  # linha em branco (rodape da planilha)

            try:
                inep = str(int(inep_bruto)).zfill(8)
            except (TypeError, ValueError):
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP invalido ({inep_bruto!r}) - ignorada."
                ))
                ignoradas += 1
                continue

            if len(inep) != 8:
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP com {len(inep)} digito(s) ({inep}) - ignorada."
                ))
                ignoradas += 1
                continue

            if Escola.objects.filter(inep=inep).exists():
                existentes += 1
                continue

            nome = _texto(linha[indice_coluna["UNIDADE ESCOLAR"]])
            if not nome:
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP {inep} sem nome de escola - ignorada."
                ))
                ignoradas += 1
                continue

            lote_bruto = linha[indice_coluna["LOTE"]]
            try:
                lote = int(lote_bruto) if lote_bruto not in (None, "") else None
            except (TypeError, ValueError):
                lote = None

            Escola.objects.create(
                inep=inep,
                nome=nome,
                endereco=_texto(linha[indice_coluna["ENDEREÇO UNIDADE ESCOLAR"]]),
                lote=lote,
                estado=_texto(linha[indice_coluna["UF"]]).upper(),
                municipio=_texto(linha[indice_coluna["MUNICIPIO"]]),
                kit_inicial=_texto(linha[indice_coluna["KIT WIFI ESTIMADO"]]),
                velocidade_dl_minima=_texto(linha[indice_coluna["VELOCIDADE"]]),
            )
            criadas += 1

        self.stdout.write(self.style.SUCCESS(
            f"Escola: {criadas} criada(s), {existentes} ja existente(s) (ignorada, sem sobrescrever), "
            f"{ignoradas} linha(s) invalida(s)."
        ))
