from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.escolas.models import Escola

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Correção pontual de dados (2026-09-01, a pedido do usuário): "Nova BASE
# EACE.xlsx" (raiz do repositório, pasta doc/) é uma planilha de
# acompanhamento do programa inteiro (todos os Lotes, ~3200 linhas, 82
# colunas de controle interno) — não o `CONSOLIDADO EACE.xlsx` já coberto
# por `importar_escolas_planilha` (RF-01/RN-007). Cabeçalho e nomes de
# coluna são completamente diferentes; por isso um comando novo, em vez de
# estender o existente. Mesma regra de segurança dele: só CRIA Escola cujo
# INEP ainda não existe no banco — nunca sobrescreve nem duplica.
COLUNAS_OBRIGATORIAS = (
    "FASE",
    "UF",
    "CIDADE",
    "CODIGO INEP",
    "NOME DA ESCOLA",
    "ENDEREÇO",
    "VELOCIDADE DL MÍNIMA (MBPS)",
    "KIT WI-FI (ESTIMADO)",
)


def _normalizar_cabecalho(valor):
    """Colunas desta planilha vêm com quebra de linha embutida no título
    (ex.: "Kit Wi-Fi\\n(estimado)") — normaliza espaços/quebras de linha
    antes de comparar, além de maiúsculas (mesmo padrão de
    `importar_escolas_planilha`)."""
    texto = (str(valor) if valor is not None else "").strip().upper()
    return " ".join(texto.split())


def _texto(valor):
    return str(valor).strip() if valor not in (None, "") else ""


class Command(BaseCommand):
    """Importa Escola a partir de "Nova BASE EACE.xlsx" (aba "base",
    cabeçalho na linha 1) — planilha de controle com escolas de todos os
    Lotes, usada aqui só para os INEPs que ainda não existem no sistema
    (novas escolas de um Lote recém-iniciado). Idempotente: repetir a
    importação é seguro — INEP já cadastrado é apenas contado, nunca
    sobrescrito nem duplicado (mesma regra de `importar_escolas_planilha`,
    RN-007). Escola nova sempre nasce com status_conexao=desconectado.

    Por padrão roda em modo simulação (não grava nada) — use --aplicar
    para gravar de fato. Mesmo padrão dos demais comandos de correção
    pontual deste projeto (ex.: `copiar_itens_relatorio_eace_para_ixc`).
    """

    help = (
        "Importa Escola (INEP, lote, UF, municipio, nome, endereco, velocidade, "
        "kit estimado) a partir de 'Nova BASE EACE.xlsx' (aba 'base'). "
        "Por padrao so simula; use --aplicar para gravar."
    )

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho para 'Nova BASE EACE.xlsx'.")
        parser.add_argument("--aba", default="base", help="Nome da aba com os dados (padrao: base).")
        parser.add_argument(
            "--linha-cabecalho", type=int, default=1,
            help="Numero da linha (1-based) com os titulos das colunas (padrao: 1).",
        )
        parser.add_argument(
            "--aplicar", action="store_true",
            help="Grava as escolas novas no banco. Sem esta flag, só mostra o que seria feito.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError(
                "Dependencia 'openpyxl' nao instalada. Adicione 'openpyxl' ao requirements.txt "
                "e reinstale as dependencias."
            )

        caminho = Path(options["arquivo"])
        if not caminho.exists():
            raise CommandError(f"Arquivo nao encontrado: {caminho}")

        try:
            planilha = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        except Exception as erro:
            raise CommandError(f"Nao foi possivel abrir '{caminho}': {erro}")

        aba = options["aba"]
        if aba not in planilha.sheetnames:
            raise CommandError(f"Aba '{aba}' nao encontrada. Abas disponiveis: {planilha.sheetnames}")

        planilha_aba = planilha[aba]
        linha_cabecalho = options["linha_cabecalho"]

        cabecalho = next(
            planilha_aba.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho, values_only=True), None
        )
        if cabecalho is None:
            raise CommandError(f"Linha de cabecalho {linha_cabecalho} vazia ou inexistente na aba '{aba}'.")

        indice_coluna = {_normalizar_cabecalho(valor): posicao for posicao, valor in enumerate(cabecalho)}
        faltando = [coluna for coluna in COLUNAS_OBRIGATORIAS if coluna not in indice_coluna]
        if faltando:
            raise CommandError(f"Colunas obrigatorias ausentes na planilha: {faltando}")

        existentes_no_banco = set(Escola.objects.values_list("inep", flat=True))

        novas = []  # lista de dicts prontos para Escola.objects.create(**dados)
        vistas_no_arquivo = set()
        existentes = 0
        duplicadas_no_arquivo = 0
        ignoradas = 0

        linhas = planilha_aba.iter_rows(min_row=linha_cabecalho + 1, values_only=True)
        for numero_linha, linha in enumerate(linhas, start=linha_cabecalho + 1):
            inep_bruto = linha[indice_coluna["CODIGO INEP"]]
            if inep_bruto in (None, ""):
                continue  # linha em branco

            try:
                inep = str(int(inep_bruto)).zfill(8)
            except (TypeError, ValueError):
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP invalido ({inep_bruto!r}) - ignorada."
                ))
                ignoradas += 1
                continue

            if len(inep) != 8:
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP com {len(inep)} digito(s) ({inep}) - ignorada."
                ))
                ignoradas += 1
                continue

            nome = _texto(linha[indice_coluna["NOME DA ESCOLA"]])
            if not nome:
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: INEP {inep} sem nome de escola - ignorada."
                ))
                ignoradas += 1
                continue

            if inep in vistas_no_arquivo:
                duplicadas_no_arquivo += 1
                continue
            vistas_no_arquivo.add(inep)

            if inep in existentes_no_banco:
                existentes += 1
                continue

            lote_bruto = linha[indice_coluna["FASE"]]
            try:
                lote = int(lote_bruto) if lote_bruto not in (None, "") else None
            except (TypeError, ValueError):
                lote = None

            novas.append(dict(
                inep=inep,
                nome=nome,
                endereco=_texto(linha[indice_coluna["ENDEREÇO"]]),
                lote=lote,
                estado=_texto(linha[indice_coluna["UF"]]).upper(),
                municipio=_texto(linha[indice_coluna["CIDADE"]]),
                kit_inicial=_texto(linha[indice_coluna["KIT WI-FI (ESTIMADO)"]]),
                velocidade_dl_minima=_texto(linha[indice_coluna["VELOCIDADE DL MÍNIMA (MBPS)"]]),
            ))

        self.stdout.write(
            f"Escola: {len(novas)} nova(s) a criar, {existentes} ja existente(s) no banco "
            f"(ignorada, sem sobrescrever), {duplicadas_no_arquivo} duplicada(s) dentro do "
            f"arquivo (mesmo INEP 2x, só a 1ª conta), {ignoradas} linha(s) invalida(s)."
        )
        for dados in novas[:20]:
            self.stdout.write(f"  + {dados['inep']} - {dados['nome']} (Lote {dados['lote']})")
        if len(novas) > 20:
            self.stdout.write(f"  ... e mais {len(novas) - 20} escola(s).")

        if not options["aplicar"]:
            self.stdout.write(self.style.WARNING(
                "Simulação (nada foi gravado). Rode novamente com --aplicar para gravar."
            ))
            return

        for dados in novas:
            Escola.objects.create(**dados)

        self.stdout.write(self.style.SUCCESS(f"{len(novas)} escola(s) criada(s)."))
