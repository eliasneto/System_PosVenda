import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.ri.models import KitPadrao

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _texto(valor):
    return str(valor).strip() if valor not in (None, "") else ""


def _numero(valor):
    if valor in (None, ""):
        return None
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    """FEAT-015/RN-010: importa o catálogo `KitPadrao` a partir do
    CONSOLIDADO EACE.xlsx (aba LPU - "TABELA 1 - LISTA DE PREÇOS
    UNITÁRIOS"). A planilha traz, repetido por bloco de colunas, um valor
    de Equipamento e um de Serviço por Lote; a coluna "Unidade" indica se
    o valor e o KIT fechado da escola (Escola/Escola-Mes) ou preco
    unitario de item avulso (Unidade, km, enlace, metro, par).

    Repetir a importacao e segura: a combinacao (descricao, lote) ja
    existente e atualizada com os valores da planilha (fonte oficial),
    nunca duplicada. Linhas de secao ("REDE EXTERNA"/"REDE INTERNA") e o
    rodape de notas sao ignorados, nao geram registro.
    """

    help = (
        "Importa o catalogo KitPadrao (item, unidade, valor de equipamento e "
        "de servico por Lote) a partir do CONSOLIDADO EACE.xlsx, aba LPU."
    )

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho para o CONSOLIDADO EACE.xlsx.")
        parser.add_argument("--aba", default="LPU", help="Nome da aba com os dados (padrao: LPU).")
        parser.add_argument(
            "--linha-lotes", type=int, default=3,
            help="Numero da linha (1-based) com o rotulo de cada bloco de Lote (padrao: 3).",
        )
        parser.add_argument(
            "--linha-cabecalho", type=int, default=4,
            help="Numero da linha (1-based) com os titulos das colunas (padrao: 4).",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError(
                "Dependencia 'openpyxl' nao instalada. Adicione 'openpyxl' ao requirements.txt "
                "e reinstale as dependencias."
            )

        caminho = Path(options["arquivo"])
        if not caminho.exists():
            raise CommandError(f"Arquivo nao encontrado: {caminho}")

        try:
            planilha = openpyxl.load_workbook(caminho, data_only=True)
        except Exception as erro:
            raise CommandError(f"Nao foi possivel abrir '{caminho}': {erro}")

        aba = options["aba"]
        if aba not in planilha.sheetnames:
            raise CommandError(f"Aba '{aba}' nao encontrada. Abas disponiveis: {planilha.sheetnames}")

        planilha_aba = planilha[aba]
        linha_lotes_num = options["linha_lotes"]
        linha_cabecalho_num = options["linha_cabecalho"]

        linha_lotes = next(
            planilha_aba.iter_rows(min_row=linha_lotes_num, max_row=linha_lotes_num, values_only=True), None
        )
        cabecalho = next(
            planilha_aba.iter_rows(min_row=linha_cabecalho_num, max_row=linha_cabecalho_num, values_only=True), None
        )
        if not linha_lotes or not cabecalho:
            raise CommandError(
                f"Linha de Lote ({linha_lotes_num}) ou de cabecalho ({linha_cabecalho_num}) "
                f"vazia ou inexistente na aba '{aba}'."
            )

        blocos_por_lote = self._mapear_blocos_por_lote(linha_lotes, cabecalho)
        if not blocos_por_lote:
            raise CommandError(
                "Nenhum bloco de Lote ('Equipamentos (R$)' + rotulo 'LOTE n') encontrado no cabecalho."
            )

        criados = 0
        atualizados = 0
        ignorados = 0
        secao_atual = ""

        linhas = planilha_aba.iter_rows(min_row=linha_cabecalho_num + 1, values_only=True)
        for numero_linha, linha in enumerate(linhas, start=linha_cabecalho_num + 1):
            item_bruto = linha[0] if linha else None
            descricao_bruta = linha[1] if linha and len(linha) > 1 else None

            if item_bruto in (None, "") and descricao_bruta in (None, ""):
                continue  # linha em branco

            if _texto(item_bruto).upper().rstrip(":") == "NOTAS":
                break  # fim dos dados; o restante da aba e nota de rodape

            try:
                int(item_bruto)
            except (TypeError, ValueError):
                if descricao_bruta in (None, ""):
                    secao_atual = _texto(item_bruto)  # ex.: "REDE EXTERNA"/"REDE INTERNA"
                    continue
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: item invalido ({item_bruto!r}) - ignorada."
                ))
                ignorados += 1
                continue

            descricao = _texto(descricao_bruta)
            if not descricao:
                self.stderr.write(self.style.WARNING(
                    f"Linha {numero_linha}: sem descricao ({secao_atual}) - ignorada."
                ))
                ignorados += 1
                continue

            unidade = _texto(linha[2]) if len(linha) > 2 else ""

            for lote, coluna_equipamento in blocos_por_lote.items():
                coluna_servico = coluna_equipamento + 1
                valor_equipamento = (
                    _numero(linha[coluna_equipamento]) if coluna_equipamento < len(linha) else None
                )
                valor_servico = _numero(linha[coluna_servico]) if coluna_servico < len(linha) else None
                if valor_equipamento is None and valor_servico is None:
                    continue  # sem valor para este lote nesta linha

                _, criado = KitPadrao.objects.update_or_create(
                    descricao=descricao,
                    lote=lote,
                    defaults={
                        "unidade": unidade,
                        "quantidade_padrao": 1,
                        "valor_equipamento": valor_equipamento,
                        "valor_servico": valor_servico,
                    },
                )
                if criado:
                    criados += 1
                else:
                    atualizados += 1

        self.stdout.write(self.style.SUCCESS(
            f"KitPadrao: {criados} criado(s), {atualizados} atualizado(s), {ignorados} linha(s) ignorada(s)."
        ))

    @staticmethod
    def _mapear_blocos_por_lote(linha_lotes, cabecalho):
        """Cada bloco de Lote comeca na coluna cujo cabecalho e
        'Equipamentos (R$)' (Servicos fica na coluna seguinte). O numero do
        Lote fica na linha acima, numa celula mesclada — por isso o valor
        so existe na primeira coluna do bloco; procura-se para tras até
        achar o rotulo (ex.: 'LOTE 9')."""
        blocos = {}
        for coluna, valor in enumerate(cabecalho):
            if not _texto(valor).upper().startswith("EQUIPAMENTO"):
                continue
            rotulo_lote = ""
            for indice in range(coluna, -1, -1):
                candidato = linha_lotes[indice] if indice < len(linha_lotes) else None
                if candidato not in (None, ""):
                    rotulo_lote = _texto(candidato)
                    break
            numeros = re.findall(r"\d+", rotulo_lote)
            if not numeros:
                continue
            blocos[int(numeros[0])] = coluna
        return blocos
