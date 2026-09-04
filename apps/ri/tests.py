import shutil
import tempfile
from datetime import date, timedelta
from decimal import Decimal
from email.message import EmailMessage as MensagemEmailMime
from io import BytesIO, StringIO
from pathlib import Path
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import CommandError, call_command
from django.db import IntegrityError, transaction
from django.test import TestCase, override_settings
from django.urls import NoReverseMatch, reverse
from django.utils import timezone

from apps.auditoria.models import Auditoria
from apps.core.email_tracking import montar_codigo_rastreio
from apps.escolas.models import Escola
from apps.integracoes.eace.rpa import ResultadoConsultaPendencias, ResultadoRpaEace, RpaEaceIndisponivel

from .models import (
    Documento,
    EmailFinanceiroLog,
    EmailFinanceiroSync,
    KitPadrao,
    LogRpaEace,
    PlanilhaEace,
    Ri,
    RiDivergencia,
    RiHistorico,
    RiItemEace,
    RiItemIxc,
    RiItemRelatorioEace,
)
from .services import (
    RI_BLOQUEADO_FATURAMENTO_CONCLUIDO,
    RI_SEM_LINHA_NA_PLANILHA,
    EmailFinanceiroSyncError,
    PlanilhaEaceSincronizacaoError,
    PlanilhaFaturamentoError,
    comparar_status_escola_relatorio,
    consultar_pendencias_portal_eace,
    detectar_delimitador_planilha_eace,
    gerar_planilha_faturamento,
    montar_corpo_email_financeiro,
    montar_dashboard_financeiro,
    montar_faturamento_por_estado,
    montar_faturamento_por_municipio,
    nome_arquivo_planilha_faturamento,
    sincronizar_divergencia_kit_relatorio,
    sincronizar_relatorio_eace_da_planilha,
    sincronizar_relatorio_eace_de_todas_as_ri,
    sincronizar_respostas_financeiro,
)

User = get_user_model()


class GridInepViewTests(TestCase):
    """FEAT-007: grid principal de INEPs (RF-05/RF-06)."""

    def setUp(self):
        self.user = User.objects.create_user(username="analista", password="senha-teste-123")
        self.escola_sem_ri = Escola.objects.create(
            inep="10000001", nome="Escola Sem RI", municipio="Fortaleza", estado="CE"
        )  # status_conexao = desconectado (padrao, sem data de instalacao)
        self.escola_com_ri = Escola.objects.create(
            inep="10000002",
            nome="Escola Com RI",
            municipio="Sobral",
            estado="CE",
            data_instalacao_re="2026-01-01",
            data_instalacao_ri="2026-01-02",
        )  # RN-007: as duas datas preenchidas -> conectado
        self.ri = Ri.objects.create(
            escola=self.escola_com_ri, status=Ri.ANDAMENTO, responsavel=self.user
        )

    def test_exige_login(self):
        resp = self.client.get(reverse("grid_inep"))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_caminho_principal_lista_todas_as_escolas(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.escola_sem_ri.inep)
        self.assertContains(resp, self.escola_com_ri.inep)
        self.assertContains(resp, "Sem RI")
        self.assertContains(resp, "Andamento")
        self.assertContains(resp, "Desconectado")
        self.assertContains(resp, "Conectado")

    def test_drilldown_kit_declarado_mostra_referencia_quando_ainda_nao_lancado(self):
        """Bug reportado pelo usuário (2026-08-26): o drill-down do Grid
        mostrava "Nenhum item lançado." para o Kit declarado mesmo quando a
        tela de detalhe (fonte master, RN-010) já resolvia uma referência a
        partir de Escola.kit_inicial + catálogo — divergência entre as duas
        telas para o mesmo 1º lado."""
        self.escola_com_ri.kit_inicial = "Kit Wi-Fi Indoor"
        self.escola_com_ri.save()
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Kit Wi-Fi Indoor — referência, ainda não lançado")

    def test_drilldown_mostra_nobreak_declarado(self):
        """RN-017: Nobreak declarado aparece junto ao Kit no card "Kit
        declarado (1º)" do drill-down — item padrão, igual para toda
        escola, sempre quantidade 1."""
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Nobreak: Nobreak — 1 un.")

    def test_drilldown_trava_select_de_status_para_analista_em_faturamento_concluido(self):
        """RN-020: com o RI em "Faturamento Concluído", o <select> de
        status do drill-down trava no valor atual para o Analista — o
        backend também recusa (RiStatusUpdateViewTests), aqui é só a UI."""
        self.ri.status = Ri.FATURAMENTO_CONCLUIDO
        self.ri.save(update_fields=["status"])
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Faturamento Concluído (só Administrador)")
        self.assertContains(resp, "<select name=\"status\" disabled")

    def test_filtro_por_status_do_ri(self):
        """RF-05: coluna e filtro "Status do RI" são o status do RI
        (RN-001, seção 5 dos requisitos)."""
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"), {"status_ri": Ri.ANDAMENTO})
        self.assertContains(resp, self.escola_com_ri.inep)
        self.assertNotContains(resp, self.escola_sem_ri.inep)

    def test_filtro_por_status_de_conexao(self):
        """RF-20: coluna e filtro "Conexão" são atributo da própria Escola
        (INEP) - independentes do status do RI."""
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"), {"status_conexao": Escola.CONECTADO})
        self.assertContains(resp, self.escola_com_ri.inep)
        self.assertNotContains(resp, self.escola_sem_ri.inep)

    def test_assunto_sugerido_do_email_inclui_nome_da_escola(self):
        """RN-050 (2026-09-02): assunto sugerido do e-mail ao financeiro
        inclui o nome da escola, além do INEP — facilita identificar a
        escola sem abrir o e-mail."""
        escola = Escola.objects.create(inep="10000005", nome="Escola Envio Email")
        Ri.objects.create(escola=escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, f"Faturamento EACE — INEP {escola.inep} — {escola.nome}")

    def test_busca_por_inep_nome_municipio_ou_uf(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"), {"q": "Sobral"})
        self.assertContains(resp, self.escola_com_ri.inep)
        self.assertNotContains(resp, self.escola_sem_ri.inep)

    def test_busca_sem_resultado_mostra_estado_vazio(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"), {"q": "inep-que-nao-existe"})
        self.assertContains(resp, "Nenhum INEP encontrado")

    def test_divergencia_aberta_conta_no_card_e_marca_a_linha(self):
        RiDivergencia.objects.create(ri=self.ri, tipo=RiDivergencia.TIPO_QUANTIDADE)
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Divergência aberta")
        # card "Com divergência" conta 1
        self.assertContains(resp, ">1<")

    def test_divergencia_resolvida_nao_conta_como_aberta(self):
        RiDivergencia.objects.create(
            ri=self.ri,
            tipo=RiDivergencia.TIPO_QUANTIDADE,
            resolvida_em="2026-08-22T00:00:00Z",
        )
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertNotContains(resp, "Divergência aberta")

    def test_card_resposta_financeiro_conta_e_link_filtra_pelo_status(self):
        """RN-016: 3º card do grid mostra a contagem de INEPs com RI em
        "Resposta Financeiro" (antigo "Aguardando Anexo portal EACE") e o
        link do card filtra o grid por esse status."""
        self.ri.status = Ri.AGUARDANDO_ANEXO_PORTAL_EACE
        self.ri.save()
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertEqual(resp.context["total_resposta_financeiro"], 1)
        self.assertContains(resp, "Resposta Financeiro")
        self.assertContains(resp, f"?status_ri={Ri.AGUARDANDO_ANEXO_PORTAL_EACE}")

        resp_filtrado = self.client.get(
            reverse("grid_inep"), {"status_ri": Ri.AGUARDANDO_ANEXO_PORTAL_EACE}
        )
        self.assertContains(resp_filtrado, self.escola_com_ri.inep)
        self.assertNotContains(resp_filtrado, self.escola_sem_ri.inep)

    def test_card_com_divergencia_vira_filtro(self):
        """Usuário pediu (2026-08-27) que o card "Com divergência" virasse
        filtro, igual ao card "Resposta Financeiro" (RN-016)."""
        RiDivergencia.objects.create(ri=self.ri, tipo=RiDivergencia.TIPO_QUANTIDADE)
        outra_escola = Escola.objects.create(inep="10000003", nome="Escola Sem Divergência")
        Ri.objects.create(escola=outra_escola, status=Ri.ANDAMENTO)

        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "?divergencia=1")

        resp_filtrado = self.client.get(reverse("grid_inep"), {"divergencia": "1"})
        self.assertContains(resp_filtrado, self.escola_com_ri.inep)
        self.assertNotContains(resp_filtrado, outra_escola.inep)
        self.assertNotContains(resp_filtrado, self.escola_sem_ri.inep)
        # O total do card continua contando todos, mesmo com o filtro ativo.
        self.assertEqual(resp_filtrado.context["total_divergencia"], 1)

    def test_filtro_de_divergencia_combina_com_status_ri(self):
        RiDivergencia.objects.create(ri=self.ri, tipo=RiDivergencia.TIPO_QUANTIDADE)
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("grid_inep"), {"divergencia": "1", "status_ri": Ri.FATURAMENTO_CONCLUIDO}
        )
        # RI está em "Andamento" — filtro combinado (divergência + outro
        # status) não deve trazer nenhuma escola.
        self.assertNotContains(resp, self.escola_com_ri.inep)

    def test_drilldown_mostra_itens_dos_3_lados_do_ri(self):
        RiItemEace.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=2, valor_unitario="350.00"
        )
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=2, valor_unitario="350.00"
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=2, valor_unitario="350.00"
        )
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Kit Wi-Fi")
        self.assertContains(resp, "Relatório EACE")

    def test_grid_tem_as_5_colunas_conexao_e_status_ri_sem_responsavel(self):
        """RF-05 + RF-20 (revisto em 2026-08-25, RN-012): grid tem 5 colunas
        - INEP, Nome, Endereço, Conexão (Escola/RF-20) e Status do RI
        (RN-001) - "Responsável" não é mais coluna da tabela principal."""
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertNotContains(resp, "<th class=\"px-4 py-2\">Responsável</th>")
        self.assertContains(
            resp, '<th class="px-4 py-2" title="Status de conexão da Escola (RF-20)">Conexão</th>'
        )

    def test_drilldown_mostra_responsavel_editavel(self):
        """RN-012: "Responsável" aparece dentro do drill-down, como um
        <select> com os usuários do sistema, pré-selecionando o atual."""
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Responsável")
        self.assertContains(resp, reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}))
        self.assertContains(resp, self.user.username)

    def test_administrador_ve_botao_forcar_resposta_financeiro(self):
        """RN-019: só o Administrador vê a opção de forçar a saída de
        "Aguardando financeiro"."""
        self.ri.status = Ri.AGUARDANDO_FINANCEIRO
        self.ri.save()
        admin = User.objects.create_user(
            username="admin-grid", password="senha-teste-123", perfil=User.PERFIL_ADMINISTRADOR
        )
        self.client.force_login(admin)
        resp = self.client.get(reverse("grid_inep"))
        self.assertContains(resp, "Forçar Resposta Financeiro")

    def test_analista_nao_ve_botao_forcar_resposta_financeiro(self):
        """RN-019: Analista não vê a opção, mesmo com o RI em "Aguardando
        financeiro"."""
        self.ri.status = Ri.AGUARDANDO_FINANCEIRO
        self.ri.save()
        self.client.force_login(self.user)
        resp = self.client.get(reverse("grid_inep"))
        self.assertNotContains(resp, "Forçar Resposta Financeiro")

    def test_botao_forcar_nao_aparece_fora_de_aguardando_financeiro(self):
        """RN-019: botão é exclusivo desse status, mesmo para
        Administrador — o RI de fixture já nasce em "Andamento"."""
        admin = User.objects.create_user(
            username="admin-grid-2", password="senha-teste-123", perfil=User.PERFIL_ADMINISTRADOR
        )
        self.client.force_login(admin)
        resp = self.client.get(reverse("grid_inep"))
        self.assertNotContains(resp, "Forçar Resposta Financeiro")


class RiStatusUpdateViewTests(TestCase):
    """Início da FEAT-006 (fora de ordem): troca manual de status do RI a
    partir do grid, com as regras já fechadas de RN-001/RN-003."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="analista", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.admin = User.objects.create_user(
            username="admin-status", password="senha-teste-123", perfil=User.PERFIL_ADMINISTRADOR
        )
        self.escola = Escola.objects.create(inep="30000001", nome="Escola RI Status")

    def test_exige_login(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        resp = self.client.post(reverse("ri_status_update", kwargs={"pk": ri.pk}))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_troca_manual_permitida(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_bloqueia_status_automatico_do_sistema(self):
        """"Aguardando financeiro" só é trocado pelo sistema (RN-001)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_FINANCEIRO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_correcao_mega_so_a_partir_de_andamento(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_VALIDACAO_EACE)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.CORRECAO_MEGA, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_VALIDACAO_EACE)

    def test_troca_de_status_gera_entrada_automatica_no_historico(self):
        """FEAT-014/RN-008: mudança de status grava rótulo + valor
        anterior/novo na linha do tempo do RI, não só uma frase livre."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_STATUS)
        self.assertEqual(entrada.campo, "Status do RI")
        self.assertEqual(entrada.valor_anterior, "Em Andamento")
        self.assertEqual(entrada.valor_novo, "Envio de Email para faturamento")
        self.assertEqual(entrada.autor, self.user)

    def test_transicao_bloqueada_nao_gera_entrada_no_historico(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_FINANCEIRO, "next": reverse("grid_inep")},
        )
        self.assertFalse(RiHistorico.objects.filter(ri=ri).exists())

    def test_correcao_mega_permitida_a_partir_de_andamento(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.CORRECAO_MEGA, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.CORRECAO_MEGA)

    def test_correcao_mega_so_retorna_para_andamento(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.CORRECAO_MEGA)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.FATURAMENTO_CONCLUIDO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.CORRECAO_MEGA)

    def test_bloqueia_envio_email_com_divergencia_aberta_bloqueante(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiDivergencia.objects.create(ri=ri, tipo=RiDivergencia.TIPO_QUANTIDADE, bloqueia=True)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ANDAMENTO)

    def test_permite_envio_email_com_divergencia_ja_resolvida(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiDivergencia.objects.create(
            ri=ri,
            tipo=RiDivergencia.TIPO_QUANTIDADE,
            bloqueia=True,
            resolvida_em="2026-08-22T00:00:00Z",
        )
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_bloqueia_envio_email_com_divergencia_gerada_automaticamente(self):
        """FEAT-006 (integração com FEAT-005): a RiDivergencia não precisa
        ser criada manualmente para bloquear — basta o Lado IXC e o Lado
        Relatório EACE terem KIT/Produtos diferentes (RN-003) para o
        gerador automático (`sincronizar_divergencia_kit_relatorio`)
        alimentar a mesma divergência que bloqueia a transição. RN-003
        (ajustada em 2026-09-02): precisa dos dois lados com algum item —
        um lado vazio não conta como divergência."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="KIT Cobertura Wi-Fi - 6 Access Points",
            quantidade=1, valor_unitario=Decimal("0.00"), eh_kit=True,
        )
        # Lado Relatório EACE com um KIT diferente (não vazio) — diverge
        # de verdade. O recálculo roda depois de qualquer lançamento real
        # (RN-003, `ri_detail_view`); chamado direto aqui para isolar o
        # teste do formulário de lançamento (já coberto pela FEAT-004).
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="KIT Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("400.00"), eh_kit=True,
        )
        sincronizar_divergencia_kit_relatorio(ri)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ANDAMENTO)
        self.assertTrue(
            RiDivergencia.objects.filter(
                ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, resolvida_em__isnull=True
            ).exists()
        )

    def test_permite_envio_email_apos_divergencia_gerada_automaticamente_ser_resolvida(self):
        """Mesmo cenário acima, mas o Lado Relatório EACE recebe o mesmo
        KIT do Lado IXC antes da transição — o gerador automático resolve
        a divergência (RN-003) e a transição deixa de ser bloqueada."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="KIT Cobertura Wi-Fi - 6 Access Points",
            quantidade=1, valor_unitario=Decimal("0.00"), eh_kit=True,
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="KIT Cobertura Wi-Fi - 6 Access Points",
            quantidade=1, valor_unitario=Decimal("500.00"), eh_kit=True,
        )
        sincronizar_divergencia_kit_relatorio(ri)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_htmx_atualiza_sem_redirecionar(self):
        """FEAT-019: com o header do HTMX, a resposta é o fragmento (200)
        que atualiza o formulário e o badge do grid — não um redirect, que
        recarregaria a página inteira."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Status do RI atualizado.")
        self.assertContains(resp, f'id="form-status-{ri.pk}"')
        self.assertContains(resp, f'id="status-badge-{self.escola.inep}"')
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_htmx_com_transicao_bloqueada_retorna_fragmento_com_erro(self):
        """FEAT-019: rejeição também responde com fragmento (200) e a
        mensagem de erro, sem alterar o status real do RI."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_FINANCEIRO, "next": reverse("grid_inep")},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Esse status só é alterado automaticamente pelo sistema.")
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_administrador_forca_saida_de_aguardando_financeiro(self):
        """RN-019: exceção do Administrador — sai de "Aguardando financeiro"
        direto para "Resposta Financeiro", sem esperar o gatilho automático."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_ANEXO_PORTAL_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

    def test_analista_nao_forca_saida_de_aguardando_financeiro(self):
        """RN-019: a exceção é só do Administrador — Analista continua sem
        nenhuma opção manual nesse status, igual antes da regra."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_ANEXO_PORTAL_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_FINANCEIRO)

    def test_administrador_nao_forca_para_outro_destino(self):
        """RN-019: único destino liberado é "Resposta Financeiro" — para
        qualquer outro valor, a regra de status automático continua valendo
        mesmo para o Administrador."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.IMPLANTACAO_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_FINANCEIRO)

    def test_administrador_nao_forca_de_outra_origem(self):
        """RN-019: a exceção só vale a partir de "Aguardando financeiro" —
        de qualquer outro status, "Resposta Financeiro" continua sendo só
        automático mesmo para o Administrador."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_ANEXO_PORTAL_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_forcar_saida_gera_entrada_no_historico_com_administrador_como_autor(self):
        """RN-019/RN-008: a transição forçada grava log igual às demais
        trocas de status, identificando o Administrador como autor."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_ANEXO_PORTAL_EACE, "next": reverse("grid_inep")},
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_STATUS)
        self.assertEqual(entrada.valor_anterior, "Aguardando financeiro")
        self.assertEqual(entrada.valor_novo, "Resposta Financeiro")
        self.assertEqual(entrada.autor, self.admin)

    def test_analista_nao_troca_status_a_partir_de_faturamento_concluido(self):
        """RN-020: com o RI em "Faturamento Concluído", só o Administrador
        troca o status — Analista perde a opção que tem nos demais status
        editáveis (RN-001)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ANDAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.FATURAMENTO_CONCLUIDO)

    def test_administrador_troca_status_a_partir_de_faturamento_concluido(self):
        """RN-020: a exceção do bloqueio é só para o Administrador."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ANDAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ANDAMENTO)

    def test_marca_anexo_eace_a_partir_de_resposta_financeiro(self):
        """FEAT-010/RF-10: Analista e Administrador podem marcar o anexo
        feito no portal EACE a partir de "Resposta Financeiro"."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_VALIDACAO_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_VALIDACAO_EACE)

    def test_marca_anexo_eace_bloqueada_fora_de_resposta_financeiro(self):
        """FEAT-010/RF-10: fora de "Resposta Financeiro", a marcação é
        rejeitada — mesmo para o Administrador (não é uma exceção dele,
        RN-019 é só para a saída de "Aguardando financeiro")."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_VALIDACAO_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ANDAMENTO)

    def test_administrador_marca_anexo_eace_a_partir_de_faturamento_concluido(self):
        """RN-020: correção do Administrador — volta um RI já concluído
        para "Aguardando validação EACE" continua permitida (não é a
        marcação normal do fluxo, é a exceção já aberta pela RN-020)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.AGUARDANDO_VALIDACAO_EACE, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_VALIDACAO_EACE)

    def test_conclui_faturamento_a_partir_de_aguardando_validacao_eace(self):
        """FEAT-010/RF-11: conclusão manual, disponível a partir de
        "Aguardando validação EACE" — grava `concluido_em`."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_VALIDACAO_EACE)
        self.assertIsNone(ri.concluido_em)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.FATURAMENTO_CONCLUIDO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.FATURAMENTO_CONCLUIDO)
        self.assertIsNotNone(ri.concluido_em)

    def test_conclusao_bloqueada_fora_de_aguardando_validacao_eace(self):
        """FEAT-010/RF-11: "Botão de conclusão só habilitado depois da
        marcação de anexo" (checklist.md) — sem passar por "Aguardando
        validação EACE", a conclusão manual é rejeitada."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.FATURAMENTO_CONCLUIDO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.assertIsNone(ri.concluido_em)

    def test_conclusao_gera_entrada_no_historico(self):
        """RN-008: a conclusão manual grava log igual às demais trocas de
        status, identificando o usuário como autor."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_VALIDACAO_EACE)
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.FATURAMENTO_CONCLUIDO, "next": reverse("grid_inep")},
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_STATUS)
        self.assertEqual(entrada.valor_anterior, "Aguardando validação EACE")
        self.assertEqual(entrada.valor_novo, "Faturamento Concluído")
        self.assertEqual(entrada.autor, self.user)


class RiLogRpaEaceDispararViewTests(TestCase):
    """FEAT-033 (Fase 3, RN-056/RN-058): "Disparar RPA" só enfileira - quem
    executa de fato é `processar_proximo_da_fila_rpa_eace` (testado à
    parte em `ProcessarFilaRpaEaceTests`). Aqui o núcleo do RPA nunca
    precisa ser mockado, porque a view não chama mais."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="analista-rpa", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.escola = Escola.objects.create(inep="35083938", nome="Escola RPA EACE")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF,
            arquivo=SimpleUploadedFile("nota.pdf", b"%PDF-fake"),
        )
        self.xml = Documento.objects.create(
            ri=self.ri, tipo=Documento.XML,
            arquivo=SimpleUploadedFile("nota.xml", b"<nfe/>"),
        )
        self.log = LogRpaEace.objects.create(ri=self.ri)

    def _disparar(self, **post_extra):
        post = {"documento_pdf": self.pdf.pk, "documento_xml": self.xml.pk, "next": ""}
        post.update(post_extra)
        return self.client.post(reverse("ri_log_rpa_eace_disparar", kwargs={"pk": self.log.pk}), post)

    def test_exige_login(self):
        resp = self._disparar()
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_sem_documentos_selecionados_nao_enfileira(self):
        self.client.force_login(self.user)
        self._disparar(documento_pdf="", documento_xml="")
        self.log.refresh_from_db()
        self.assertEqual(self.log.resultado, LogRpaEace.PENDENTE)

    def test_sem_num_osp_nao_enfileira(self):
        """Sem RiItemRelatorioEace.num_osp (FEAT-024), não há como saber a
        OSP no portal — bloqueia antes de entrar na fila."""
        self.client.force_login(self.user)
        self._disparar()
        self.log.refresh_from_db()
        self.assertEqual(self.log.resultado, LogRpaEace.PENDENTE)

    def test_disparo_valido_entra_na_fila_sem_executar_nada(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        self.client.force_login(self.user)
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal") as mock_rpa:
            self._disparar()
            mock_rpa.assert_not_called()
        self.log.refresh_from_db()
        self.assertEqual(self.log.resultado, LogRpaEace.NA_FILA)
        self.assertEqual(self.log.tentativas, 0)
        self.assertIsNotNone(self.log.enfileirado_em)
        self.assertEqual(self.log.documento_pdf_id, self.pdf.pk)
        self.assertEqual(self.log.documento_xml_id, self.xml.pk)

    def test_tentar_novamente_reseta_tentativas_e_reenfileira(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        self.log.resultado = LogRpaEace.ERRO
        self.log.motivo_erro = "valor_divergente"
        self.log.tentativas = 1
        self.log.save()
        self.client.force_login(self.user)

        self._disparar()

        self.log.refresh_from_db()
        self.assertEqual(self.log.resultado, LogRpaEace.NA_FILA)
        self.assertEqual(self.log.tentativas, 0)
        self.assertEqual(self.log.motivo_erro, "")

    def test_log_ja_com_sucesso_nao_pode_ser_reenviado(self):
        """Pedido do usuário (2026-09-03): depois de "Sucesso" os inputs
        não podem mais ser editados - o template já esconde o formulário,
        isso aqui garante que o backend também recusa (defesa em
        profundidade, ex.: reenvio direto do formulário via POST)."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        self.log.resultado = LogRpaEace.SUCESSO
        self.log.documento_pdf = self.pdf
        self.log.documento_xml = self.xml
        self.log.inep_pdf = "35083938"
        self.log.save()
        self.client.force_login(self.user)

        outro_pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF,
            arquivo=SimpleUploadedFile("outra-nota.pdf", b"%PDF-fake-2"),
        )
        self._disparar(documento_pdf=outro_pdf.pk)

        self.log.refresh_from_db()
        self.assertEqual(self.log.resultado, LogRpaEace.SUCESSO, "não pode sair de Sucesso")
        self.assertEqual(self.log.documento_pdf_id, self.pdf.pk, "não pode trocar o documento já aceito")


class RotularDocumentosPdfTests(TestCase):
    """Melhoria (2026-09-04, RN-057): usuário reportou não ter como saber,
    antes de escolher o PDF no select de "Disparar RPA", qual Nota Fiscal
    correspondia a qual produto/valor - só descobria depois de um "Erro
    (valor divergente)". O select passa a mostrar Produto/Valor extraídos
    do PDF junto do nome do arquivo."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="analista-rotulo", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.escola = Escola.objects.create(inep="35083938", nome="Escola Rótulo NF")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF,
            arquivo=SimpleUploadedFile("1617_nota.pdf", b"%PDF-fake"),
        )
        Documento.objects.create(
            ri=self.ri, tipo=Documento.XML, arquivo=SimpleUploadedFile("1617_nota.xml", b"<nfe/>"),
        )
        LogRpaEace.objects.create(ri=self.ri)

    def test_select_mostra_produto_e_valor_extraidos_do_pdf(self):
        self.client.force_login(self.user)
        dados = {"inep": "35083938", "produto": "KIT WI-FI - 15 ACCESS POINT (LT 11)", "valor": "25.330,63"}
        with patch("apps.integracoes.eace.extrair_dados_pdf.extrair_dados_nota_fiscal", return_value=dados):
            resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "KIT WI-FI - 15 ACCESS POINT (LT 11)")
        self.assertContains(resp, "R$ 25.330,63")

    def test_falha_ao_ler_pdf_nao_quebra_a_tela(self):
        """PDF ilegível (ou lib ausente) só deixa o rótulo sem o
        complemento - a tela não pode quebrar por causa de um preview."""
        self.client.force_login(self.user)
        with patch(
            "apps.integracoes.eace.extrair_dados_pdf.extrair_dados_nota_fiscal",
            side_effect=RuntimeError("boom"),
        ):
            resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(resp.status_code, 200)
        # Nome exato no disco pode ganhar sufixo do storage se outro teste
        # já tiver salvo um arquivo homônimo antes - confere pelo nome real
        # gravado neste teste, não por um literal fixo.
        self.pdf.refresh_from_db()
        self.assertContains(resp, self.pdf.arquivo.name[-40:])


class ConsultarPendenciasPortalEaceServiceTests(TestCase):
    """RN-063 (melhoria 2026-09-04): serviço que dispara a consulta
    somente-leitura ao portal e grava o resultado no próprio `Ri`."""

    def setUp(self):
        self.escola = Escola.objects.create(inep="53005090", nome="Escola Pendências")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

    def test_sem_osp_levanta_value_error(self):
        with self.assertRaises(ValueError):
            consultar_pendencias_portal_eace(self.ri)

    def test_sucesso_grava_linhas_e_consultado_em(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3905",
        )
        linhas = [{"inep": "53005090", "status": "Pendente", "descricao": "Nobreak", "valor": "1.491,72", "indice": 1}]
        resultado_mock = ResultadoConsultaPendencias(sucesso=True, linhas=linhas)
        with patch("apps.integracoes.eace.rpa.consultar_pendencias_eace", return_value=resultado_mock) as mock_consulta:
            resultado = consultar_pendencias_portal_eace(self.ri)

        mock_consulta.assert_called_once_with(osp="3905", inep="53005090")
        self.assertTrue(resultado.sucesso)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.pendencias_portal_eace, [{**linhas[0], "osp": "3905"}])
        self.assertEqual(self.ri.pendencias_portal_eace_motivo_erro, "")
        self.assertIsNotNone(self.ri.pendencias_portal_eace_consultado_em)

    def test_erro_grava_motivo_e_zera_linhas(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3905",
        )
        # Pendências de uma consulta anterior não podem "vazar" como se
        # fossem desta tentativa, que falhou.
        self.ri.pendencias_portal_eace = [{"status": "Pendente", "descricao": "Antiga", "valor": "1,00"}]
        self.ri.save(update_fields=["pendencias_portal_eace"])
        resultado_mock = ResultadoConsultaPendencias(sucesso=False, motivo="osp_nao_encontrada")
        with patch("apps.integracoes.eace.rpa.consultar_pendencias_eace", return_value=resultado_mock):
            resultado = consultar_pendencias_portal_eace(self.ri)

        self.assertFalse(resultado.sucesso)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.pendencias_portal_eace, [])
        self.assertEqual(self.ri.pendencias_portal_eace_motivo_erro, "3905:osp_nao_encontrada")

    def test_consulta_todas_as_osps_distintas_do_ri_e_junta_as_linhas(self):
        """RN-064 (correção 2026-09-04): RI com itens em OSPs diferentes -
        a consulta precisa cobrir TODAS, não só a 1ª OSP não vazia (senão
        as pendências da 2ª OSP ficam escondidas da tela)."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=1, valor_unitario=1491.72, num_osp="3905",
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=1, valor_unitario=25330.63, num_osp="4867",
        )
        linhas_3905 = [{"inep": "53005090", "status": "Enviado", "descricao": "Nobreak", "valor": "1.491,72", "indice": 1}]
        linhas_4867 = [{"inep": "53005090", "status": "Pendente", "descricao": "Kit Wi-Fi", "valor": "25.330,63", "indice": 1}]

        def _side_effect(*, osp, inep):
            linhas = linhas_3905 if osp == "3905" else linhas_4867
            return ResultadoConsultaPendencias(sucesso=True, linhas=linhas)

        with patch("apps.integracoes.eace.rpa.consultar_pendencias_eace", side_effect=_side_effect) as mock_consulta:
            resultado = consultar_pendencias_portal_eace(self.ri)

        self.assertEqual(mock_consulta.call_count, 2)
        self.assertTrue(resultado.sucesso)
        self.ri.refresh_from_db()
        osps_nas_linhas = {linha["osp"] for linha in self.ri.pendencias_portal_eace}
        self.assertEqual(osps_nas_linhas, {"3905", "4867"})

    def test_falha_parcial_mostra_o_que_deu_certo(self):
        """1 OSP falhando (ex.: ambiente fora do ar naquele momento) não
        pode esconder as pendências das OSPs que responderam certo."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=1, valor_unitario=25330.63, num_osp="4867",
        )
        linhas_4867 = [{"inep": "53005090", "status": "Pendente", "descricao": "Kit Wi-Fi", "valor": "25.330,63", "indice": 1}]

        def _side_effect(*, osp, inep):
            if osp == "3905":
                return ResultadoConsultaPendencias(sucesso=False, motivo="erro_playwright")
            return ResultadoConsultaPendencias(sucesso=True, linhas=linhas_4867)

        with patch("apps.integracoes.eace.rpa.consultar_pendencias_eace", side_effect=_side_effect):
            resultado = consultar_pendencias_portal_eace(self.ri)

        self.assertTrue(resultado.sucesso)
        self.ri.refresh_from_db()
        self.assertEqual(len(self.ri.pendencias_portal_eace), 1)
        self.assertEqual(self.ri.pendencias_portal_eace[0]["osp"], "4867")
        self.assertEqual(self.ri.pendencias_portal_eace_motivo_erro, "")


class RiConsultarPendenciasEaceViewTests(TestCase):
    """RN-063 (melhoria 2026-09-04): view que dispara a consulta a partir
    da tela do RI."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="analista-pendencias", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.escola = Escola.objects.create(inep="53005090", nome="Escola Pendências View")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3905",
        )

    def _consultar(self):
        return self.client.post(
            reverse("ri_consultar_pendencias_eace", kwargs={"pk": self.ri.pk}), {"next": ""},
        )

    def test_exige_login(self):
        resp = self._consultar()
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_nao_roda_se_ja_tem_log_processando(self):
        self.client.force_login(self.user)
        LogRpaEace.objects.create(ri=self.ri, resultado=LogRpaEace.PROCESSANDO)
        with patch("apps.ri.views.consultar_pendencias_portal_eace") as mock_consulta:
            self._consultar()
        mock_consulta.assert_not_called()

    def test_sucesso_chama_o_servico(self):
        self.client.force_login(self.user)
        resultado_mock = ResultadoConsultaPendencias(sucesso=True, linhas=[])
        with patch("apps.ri.views.consultar_pendencias_portal_eace", return_value=resultado_mock) as mock_consulta:
            self._consultar()
        mock_consulta.assert_called_once_with(self.ri)

    def test_ambiente_indisponivel_nao_quebra_a_tela(self):
        self.client.force_login(self.user)
        with patch(
            "apps.ri.views.consultar_pendencias_portal_eace",
            side_effect=RpaEaceIndisponivel("Playwright nao instalado"),
        ):
            resp = self._consultar()
        self.assertIn(resp.status_code, (200, 302))


class ProcessarFilaRpaEaceTests(TestCase):
    """FEAT-033 (Fase 3, RN-058): processo consumidor da fila - 1 execução
    por chamada, FIFO, reprocessamento automático só de erro não mapeado.
    `anexar_nota_fiscal` é sempre mockado — não há como (nem se deve)
    testar contra o portal real aqui."""

    def setUp(self):
        self.escola = Escola.objects.create(inep="35083938", nome="Escola RPA EACE")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        self.pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF,
            arquivo=SimpleUploadedFile("nota.pdf", b"%PDF-fake"),
        )
        self.xml = Documento.objects.create(
            ri=self.ri, tipo=Documento.XML,
            arquivo=SimpleUploadedFile("nota.xml", b"<nfe/>"),
        )

    def _enfileirar(self, ri=None, enfileirado_em=None, **extra):
        return LogRpaEace.objects.create(
            ri=ri or self.ri, documento_pdf=self.pdf, documento_xml=self.xml,
            resultado=LogRpaEace.NA_FILA, enfileirado_em=enfileirado_em or timezone.now(), **extra,
        )

    def test_fila_vazia_retorna_none(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace
        self.assertIsNone(processar_proximo_da_fila_rpa_eace())

    def test_marca_processando_antes_de_chamar_a_rpa(self):
        """O usuário reportou (2026-09-03) que o status ia direto de "Na
        fila" pra "Erro", sem nunca mostrar "Processando" - a troca
        precisa estar gravada (e comitada) já quando a RPA está rodando,
        não só depois que ela termina."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        estados_capturados = []

        def _side_effect(**kwargs):
            estados_capturados.append(LogRpaEace.objects.get(pk=log.pk).resultado)
            return ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", side_effect=_side_effect):
            processar_proximo_da_fila_rpa_eace()

        self.assertEqual(estados_capturados, [LogRpaEace.PROCESSANDO])

    def test_sucesso_marca_log_e_avanca_status_com_1_log_so(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        resultado = ResultadoRpaEace(
            sucesso=True,
            dados_pdf={"inep": "35083938", "produto": "Kit Cobertura Wi-Fi", "valor": "22.644,43"},
            valor_portal="22.644,43",
        )
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=resultado) as mock_rpa:
            saida = processar_proximo_da_fila_rpa_eace()
            mock_rpa.assert_called_once()
            kwargs = mock_rpa.call_args.kwargs
            self.assertEqual(kwargs["osp"], "3929")
            self.assertEqual(kwargs["inep"], "35083938")
            self.assertEqual(kwargs["caminho_pdf"], self.pdf.arquivo.path)
            self.assertEqual(kwargs["caminho_xml"], self.xml.arquivo.path)
            self.assertTrue(callable(kwargs["progresso_callback"]), "barra de progresso (2026-09-03)")
        log.refresh_from_db()
        self.ri.refresh_from_db()
        self.assertEqual(saida["resultado"], LogRpaEace.SUCESSO)
        self.assertEqual(log.resultado, LogRpaEace.SUCESSO)
        self.assertEqual(log.tentativas, 1)
        self.assertEqual(log.produto_pdf, "Kit Cobertura Wi-Fi")
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_VALIDACAO_EACE)

    def test_erro_de_regra_de_negocio_e_definitivo_na_1a_tentativa(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        resultado = ResultadoRpaEace(sucesso=False, motivo="valor_divergente", dados_pdf={"inep": "35083938"})
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=resultado):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.ri.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.ERRO)
        self.assertEqual(log.motivo_erro, "valor_divergente")
        self.assertEqual(log.tentativas, 1)
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

    def test_erro_nao_mapeado_reprocessa_1_vez_antes_de_ser_definitivo(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        erro_tecnico = ResultadoRpaEace(sucesso=False, motivo="login")

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=erro_tecnico):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.NA_FILA, "1ª falha não mapeada volta pra fila")
        self.assertEqual(log.tentativas, 1)
        primeiro_enfileiramento = log.enfileirado_em

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=erro_tecnico):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.ERRO, "2ª falha seguida vira definitiva")
        self.assertEqual(log.tentativas, 2)
        self.assertGreaterEqual(log.enfileirado_em, primeiro_enfileiramento)

    def test_reprocessamento_com_sucesso_na_2a_tentativa(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        erro_tecnico = ResultadoRpaEace(sucesso=False, motivo="erro_playwright")
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=erro_tecnico):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.NA_FILA)

        sucesso = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "22.644,43"})
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.SUCESSO)
        self.assertEqual(log.tentativas, 2)

    def test_ambiente_indisponivel_conta_como_nao_mapeado(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        with patch(
            "apps.integracoes.eace.rpa.anexar_nota_fiscal",
            side_effect=RpaEaceIndisponivel("Playwright não instalado"),
        ):
            processar_proximo_da_fila_rpa_eace()
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.NA_FILA)
        self.assertEqual(log.motivo_erro, "ambiente_indisponivel")

    def test_ordem_fifo_processa_o_mais_antigo_primeiro(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        antigo = self._enfileirar(enfileirado_em=timezone.now() - timedelta(minutes=5), tentativas=0)
        recente = self._enfileirar()
        resultado = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=resultado):
            saida = processar_proximo_da_fila_rpa_eace()
        self.assertEqual(saida["log_id"], antigo.pk)
        recente.refresh_from_db()
        self.assertEqual(recente.resultado, LogRpaEace.NA_FILA, "o mais recente nem foi tocado")

    def test_so_avanca_status_quando_todos_os_logs_do_ri_derem_sucesso(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log1 = self._enfileirar()
        log2 = self._enfileirar()
        sucesso = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso):
            processar_proximo_da_fila_rpa_eace()
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE, "ainda falta o 2º log")

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso):
            processar_proximo_da_fila_rpa_eace()
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_VALIDACAO_EACE)

    def test_usa_a_osp_do_item_que_bate_com_o_valor_da_nf_quando_ri_tem_varias_osps(self):
        """RN-064 (correção 2026-09-04): usuário reportou (INEP 53005090,
        RI 202) um RI com itens em OSPs diferentes (Nobreak numa OSP, Kit
        Wi-Fi/Access Point Adicional em outra) - pegar a 1ª OSP não vazia
        do RI inteiro (ignorando qual NF está sendo processada) mandava a
        NF certa pra OSP errada, mesmo com a OSP certa já cadastrada no
        item certo. `self.ri` já nasce (setUp) com 1 item OSP=3929/
        valor=1 - este teste soma um 2º item, em outra OSP, com o mesmo
        valor da NF, e confere que é essa OSP (não a do 1º item) que é
        usada."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Wi-Fi", quantidade=1, valor_unitario=25330.63, num_osp="4867",
        )
        self._enfileirar()
        sucesso = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "25.330,63"})

        with patch(
            "apps.integracoes.eace.extrair_dados_pdf.extrair_dados_nota_fiscal",
            return_value={"inep": "35083938", "produto": "Kit Wi-Fi", "valor": "25.330,63"},
        ), patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso) as mock_rpa:
            processar_proximo_da_fila_rpa_eace()

        self.assertEqual(mock_rpa.call_args.kwargs["osp"], "4867")

    def test_casa_pelo_valor_total_do_item_nao_so_pelo_unitario(self):
        """Correção (2026-09-04): usuário reportou um item de 3 Access
        Points com Valor Unitário R$ 699,09 - a NF traz o valor TOTAL
        (R$ 2.097,27 = 699,09 × 3), não o unitário. Comparar só o
        unitário (sem multiplicar pela quantidade) nunca bate pra item
        com quantidade > 1 - só "funcionava por acaso" pros itens de
        quantidade 1 do RI de teste."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Access Point Adicional", quantidade=3, valor_unitario=699.09, num_osp="4867",
        )
        self._enfileirar()
        sucesso = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "2.097,27"})

        with patch(
            "apps.integracoes.eace.extrair_dados_pdf.extrair_dados_nota_fiscal",
            return_value={"inep": "35083938", "produto": "Access Point Adicional", "valor": "2.097,27"},
        ), patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso) as mock_rpa:
            processar_proximo_da_fila_rpa_eace()

        self.assertEqual(mock_rpa.call_args.kwargs["osp"], "4867")

    def test_cada_tentativa_grava_1_registro_no_historico_do_ri_e_na_auditoria(self):
        """Pedido do usuário (2026-09-03): as informações da NF e do
        status de cada tentativa (mesmo reprocessando, o que sobrescreve
        os campos do `LogRpaEace`) precisam ficar na mesma linha do tempo
        onde já aparecem as trocas de status/descrições do RI
        (`RiHistorico`) - e também em `Auditoria` (log técnico)."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        erro_tecnico = ResultadoRpaEace(sucesso=False, motivo="login")
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=erro_tecnico):
            processar_proximo_da_fila_rpa_eace()

        entradas = RiHistorico.objects.filter(
            ri=self.ri, tipo=RiHistorico.LOG_CAMPO, campo=f"RPA EACE (Nota Fiscal #{log.pk})",
        )
        self.assertEqual(entradas.count(), 1)
        primeira_entrada = entradas.first()
        self.assertIsNone(primeira_entrada.autor)
        self.assertIn("Na fila", primeira_entrada.valor_novo)
        self.assertIn("tentativa 1", primeira_entrada.valor_novo)
        self.assertIn("Motivo: login", primeira_entrada.valor_novo)

        registros = Auditoria.objects.filter(
            acao=Auditoria.EXECUCAO_RPA_EACE, entidade="LogRpaEace", entidade_id=log.pk,
        )
        self.assertEqual(registros.count(), 1)
        primeiro_registro = registros.first()
        self.assertIsNone(primeiro_registro.usuario)
        self.assertIn("Na fila", primeiro_registro.valor_novo)
        self.assertIn("tentativa 1", primeiro_registro.valor_novo)
        self.assertIn("Motivo: login", primeiro_registro.valor_novo)

        sucesso = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=sucesso):
            processar_proximo_da_fila_rpa_eace()

        entradas = RiHistorico.objects.filter(
            ri=self.ri, tipo=RiHistorico.LOG_CAMPO, campo=f"RPA EACE (Nota Fiscal #{log.pk})",
        )
        self.assertEqual(entradas.count(), 2, "reprocessamento soma, não substitui, a entrada anterior")
        segunda_entrada = entradas.order_by("criado_em").last()
        self.assertIn("Sucesso", segunda_entrada.valor_novo)
        self.assertIn("tentativa 2", segunda_entrada.valor_novo)

        registros = Auditoria.objects.filter(
            acao=Auditoria.EXECUCAO_RPA_EACE, entidade="LogRpaEace", entidade_id=log.pk,
        )
        self.assertEqual(registros.count(), 2, "reprocessamento soma, não substitui, o registro anterior")
        segundo_registro = registros.order_by("criado_em").last()
        self.assertIn("Sucesso", segundo_registro.valor_novo)
        self.assertIn("tentativa 2", segundo_registro.valor_novo)

    def test_estado_inconsistente_tambem_grava_historico_e_auditoria(self):
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = LogRpaEace.objects.create(
            ri=self.ri, documento_pdf=None, documento_xml=self.xml,
            resultado=LogRpaEace.NA_FILA, enfileirado_em=timezone.now(),
        )
        processar_proximo_da_fila_rpa_eace()

        entrada = RiHistorico.objects.get(
            ri=self.ri, tipo=RiHistorico.LOG_CAMPO, campo=f"RPA EACE (Nota Fiscal #{log.pk})",
        )
        self.assertIn("Motivo: fila_sem_osp_ou_documento", entrada.valor_novo)

        registro = Auditoria.objects.get(
            acao=Auditoria.EXECUCAO_RPA_EACE, entidade="LogRpaEace", entidade_id=log.pk,
        )
        self.assertIn("Motivo: fila_sem_osp_ou_documento", registro.valor_novo)

    def test_reporta_progresso_da_rpa_no_log_durante_a_execucao(self):
        """Pedido do usuário (2026-09-03): a barra de progresso da tela
        precisa refletir, em tempo real, a etapa que a RPA está
        executando - aqui simula a RPA reportando 1 etapa no meio da
        execução (via `progresso_callback`) e confere que o log já mostra
        esse valor ANTES do processamento terminar, não só no resultado
        final."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar()
        capturado_no_meio = {}

        def _side_effect(**kwargs):
            callback = kwargs["progresso_callback"]
            callback("Preenchendo usuário", 19)
            atual = LogRpaEace.objects.get(pk=log.pk)
            capturado_no_meio["etapa"] = atual.etapa_atual
            capturado_no_meio["pct"] = atual.progresso_pct
            callback("Enviando as notas", 100)
            return ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", side_effect=_side_effect):
            processar_proximo_da_fila_rpa_eace()

        self.assertEqual(capturado_no_meio["etapa"], "Preenchendo usuário")
        self.assertEqual(capturado_no_meio["pct"], 19)

    def test_zera_progresso_ao_iniciar_um_novo_processamento(self):
        """Reprocessamento reusa o mesmo log (RN-058) - o progresso de uma
        tentativa anterior não pode "vazar" pra tela como se já fosse a
        tentativa atual."""
        from apps.ri.services import processar_proximo_da_fila_rpa_eace

        log = self._enfileirar(etapa_atual="Enviando as notas", progresso_pct=100)
        estados_capturados = []

        def _side_effect(**kwargs):
            atual = LogRpaEace.objects.get(pk=log.pk)
            estados_capturados.append((atual.etapa_atual, atual.progresso_pct))
            return ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})

        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", side_effect=_side_effect):
            processar_proximo_da_fila_rpa_eace()

        self.assertEqual(estados_capturados, [("", 0)])


class ProcessarFilaRpaEaceCommandTests(TestCase):
    """FEAT-033 (Fase 3): comando de terminal que dá 1 passada na fila -
    o loop de repetição fica por conta do agendador externo (DevOps)."""

    def setUp(self):
        self.escola = Escola.objects.create(inep="35083938", nome="Escola RPA EACE Cmd")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

    def test_fila_vazia(self):
        out = StringIO()
        call_command("processar_fila_rpa_eace", stdout=out)
        self.assertIn("vazia", out.getvalue())

    def test_processa_1_item_da_fila(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF, arquivo=SimpleUploadedFile("nota.pdf", b"%PDF"),
        )
        xml = Documento.objects.create(
            ri=self.ri, tipo=Documento.XML, arquivo=SimpleUploadedFile("nota.xml", b"<nfe/>"),
        )
        log = LogRpaEace.objects.create(
            ri=self.ri, documento_pdf=pdf, documento_xml=xml,
            resultado=LogRpaEace.NA_FILA, enfileirado_em=timezone.now(),
        )
        resultado = ResultadoRpaEace(sucesso=True, dados_pdf={"inep": "35083938", "valor": "1"})
        out = StringIO()
        with patch("apps.integracoes.eace.rpa.anexar_nota_fiscal", return_value=resultado):
            call_command("processar_fila_rpa_eace", stdout=out)
        log.refresh_from_db()
        self.assertEqual(log.resultado, LogRpaEace.SUCESSO)
        self.assertIn(str(log.pk), out.getvalue())


class ContextoLogsRpaEaceTests(TestCase):
    """FEAT-033: a seção "Notas Fiscais para anexar no portal EACE" só
    aparece com o RI em "Resposta Financeiro" (pedido do usuário,
    2026-09-03) - os logs continuam existindo depois que o RI avança,
    só deixam de ser exibidos."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="analista-contexto-rpa", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.escola = Escola.objects.create(inep="35083938", nome="Escola Contexto RPA")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        LogRpaEace.objects.create(ri=self.ri)

    def test_com_ri_em_resposta_financeiro_mostra_os_logs(self):
        from apps.ri.views import _contexto_logs_rpa_eace

        contexto = _contexto_logs_rpa_eace(self.ri, "")
        self.assertEqual(len(contexto["logs_rpa_eace"]), 1)

    def test_com_ri_em_outro_status_esconde_os_logs(self):
        from apps.ri.views import _contexto_logs_rpa_eace

        self.ri.status = Ri.AGUARDANDO_VALIDACAO_EACE
        self.ri.save()
        contexto = _contexto_logs_rpa_eace(self.ri, "")
        self.assertEqual(contexto["logs_rpa_eace"], [])
        self.assertFalse(contexto["existe_log_ativo"])

    def test_secao_some_da_tela_quando_ri_avanca_de_status(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Notas Fiscais para anexar no portal EACE")

        self.ri.status = Ri.AGUARDANDO_VALIDACAO_EACE
        self.ri.save()
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, "Notas Fiscais para anexar no portal EACE")

    def test_processando_tambem_conta_como_log_ativo(self):
        """RN-058: o polling precisa continuar enquanto o log estiver
        "Processando", não só "Na fila" - senão a tela para de atualizar
        sozinha bem na hora em que a RPA está rodando de verdade."""
        from apps.ri.views import _contexto_logs_rpa_eace

        log = self.ri.logs_rpa_eace.first()
        log.resultado = LogRpaEace.PROCESSANDO
        log.save()
        contexto = _contexto_logs_rpa_eace(self.ri, "")
        self.assertTrue(contexto["existe_log_ativo"])

    def test_posicao_na_fila_e_calculada_por_ordem_de_chegada(self):
        """RN-058: a fila é única pro sistema todo - a posição de um log
        não conta só os logs deste RI."""
        from apps.ri.views import _contexto_logs_rpa_eace

        escola_2 = Escola.objects.create(inep="35083939", nome="Escola Contexto RPA 2")
        ri_2 = Ri.objects.create(escola=escola_2, status=Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

        log_deste_ri = self.ri.logs_rpa_eace.first()
        log_deste_ri.resultado = LogRpaEace.NA_FILA
        log_deste_ri.enfileirado_em = timezone.now() - timedelta(minutes=5)
        log_deste_ri.save()

        log_de_outro_ri = LogRpaEace.objects.create(
            ri=ri_2, resultado=LogRpaEace.NA_FILA, enfileirado_em=timezone.now()
        )

        contexto = _contexto_logs_rpa_eace(self.ri, "")
        self.assertEqual(contexto["logs_rpa_eace"][0].posicao_na_fila, 1)

        contexto_2 = _contexto_logs_rpa_eace(ri_2, "")
        self.assertEqual(contexto_2["logs_rpa_eace"][0].posicao_na_fila, 2)

    def test_resposta_de_polling_nao_marca_a_secao_de_logs_como_oob(self):
        """Bug real reportado pelo usuário em 2026-09-03: a seção só
        atualizava sozinha com F5. Causa: a resposta do polling também
        marcava a própria seção de logs como `hx-swap-oob`, o que competia
        com o `hx-get`/`hx-trigger` que fez a consulta (o pill de status,
        concatenado na mesma resposta, é outro elemento - continua
        `hx-swap-oob` normalmente, isso nunca foi o problema)."""
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("ri_logs_rpa_eace_status", kwargs={"inep": self.escola.inep})
        )
        self.assertNotContains(resp, f'id="logs-rpa-eace-{self.ri.pk}" hx-swap-oob="true"')

    def test_resposta_do_disparo_continua_marcando_a_secao_de_logs_como_oob(self):
        """A resposta do disparo (`hx-swap="none"` no form) só chega à
        tela via out-of-band - continua precisando do `hx-swap-oob`."""
        pdf = Documento.objects.create(
            ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF, arquivo=SimpleUploadedFile("nota.pdf", b"%PDF"),
        )
        xml = Documento.objects.create(
            ri=self.ri, tipo=Documento.XML, arquivo=SimpleUploadedFile("nota.xml", b"<nfe/>"),
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit", quantidade=1, valor_unitario=1, num_osp="3929",
        )
        log = self.ri.logs_rpa_eace.first()
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_log_rpa_eace_disparar", kwargs={"pk": log.pk}),
            {"documento_pdf": pdf.pk, "documento_xml": xml.pk, "next": ""},
            HTTP_HX_REQUEST="true",
        )
        self.assertContains(resp, f'id="logs-rpa-eace-{self.ri.pk}" hx-swap-oob="true"')


class RiResponsavelUpdateViewTests(TestCase):
    """RN-012: reatribuição manual do responsável do RI, a partir de um
    <select> com os usuários do sistema (drill-down do grid e ri_detail)."""

    def setUp(self):
        self.user = User.objects.create_user(username="analista", password="senha-teste-123")
        self.outro_usuario = User.objects.create_user(
            username="outro-analista", password="senha-teste-123"
        )
        self.escola = Escola.objects.create(inep="40000001", nome="Escola Responsável")
        self.ri = Ri.objects.create(
            escola=self.escola, status=Ri.ANDAMENTO, responsavel=self.user
        )

    def test_exige_login(self):
        resp = self.client.post(reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_reatribui_responsavel(self):
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}),
            {"responsavel": self.outro_usuario.pk, "next": reverse("grid_inep")},
        )
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.responsavel, self.outro_usuario)

    def test_reatribuicao_gera_entrada_no_historico(self):
        self.client.force_login(self.user)
        self.client.post(
            reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}),
            {"responsavel": self.outro_usuario.pk, "next": reverse("grid_inep")},
        )
        entrada = RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada.campo, "Responsável")
        self.assertEqual(entrada.valor_anterior, self.user.username)
        self.assertEqual(entrada.valor_novo, self.outro_usuario.username)
        self.assertEqual(entrada.autor, self.user)

    def test_sem_selecionar_usuario_nao_altera_e_mostra_erro(self):
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}),
            {"responsavel": "", "next": reverse("grid_inep")},
            follow=True,
        )
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.responsavel, self.user)
        self.assertContains(resp, "Selecione um responsável.")

    def test_ri_detail_mostra_select_de_responsavel(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}))
        self.assertContains(resp, self.outro_usuario.username)

    def test_htmx_atualiza_sem_redirecionar_origem_grid(self):
        """FEAT-019: origem=grid devolve o formulário do drill-down do
        grid, sem redirecionar (sem recarregar a página)."""
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}),
            {
                "responsavel": self.outro_usuario.pk,
                "next": reverse("grid_inep"),
                "origem": "grid",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Responsável do RI atualizado.")
        self.assertContains(resp, f'id="form-responsavel-grid-{self.ri.pk}"')
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.responsavel, self.outro_usuario)

    def test_htmx_atualiza_sem_redirecionar_origem_detail(self):
        """FEAT-019: origem=detail devolve o card do cabeçalho da tela do
        RI, não o formulário do grid."""
        self.client.force_login(self.user)
        resp = self.client.post(
            reverse("ri_responsavel_update", kwargs={"pk": self.ri.pk}),
            {
                "responsavel": self.outro_usuario.pk,
                "next": reverse("ri_detail", kwargs={"inep": self.escola.inep}),
                "origem": "detail",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, f'id="card-responsavel-{self.ri.pk}"')
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.responsavel, self.outro_usuario)


_MEDIA_ROOT_TESTE_FINANCEIRO = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_FINANCEIRO)
class RiEnvioFinanceiroTests(TestCase):
    """FEAT-008/RF-16-18, RN-009: tela única de composição de e-mail
    (De automático/Para/Cc/Assunto/Anexo/Mensagem, reaberta em 2026-08-24 a
    pedido do usuário) — confirma os dados e envia o e-mail ao financeiro
    num só passo, com código de rastreio e transição automática de status.
    MEDIA_ROOT isolado (o PDF enviado também fica anexado na linha do
    tempo, FEAT-014) para não gravar no `media/` real."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_FINANCEIRO, ignore_errors=True)

    def setUp(self):
        self.user = User.objects.create_user(username="analista-fin", password="senha-teste-123")
        self.escola = Escola.objects.create(
            inep="50000001", nome="Escola Financeiro", municipio="Fortaleza", estado="CE"
        )
        self.ri = Ri.objects.create(
            escola=self.escola,
            status=Ri.ENVIO_EMAIL_FATURAMENTO,
            municipio_ixc="Fortaleza",
            estado_ixc="CE",
            data_ativacao=date(2026, 8, 1),
            cnpj="00.000.000/0001-00",
            cnpj_ficticio="11.111.111/0001-11",
        )
        # RN-013: KIT sempre usa a aba fixa "NF KIT" da planilha-modelo —
        # não depende de estar cadastrado no catálogo (usado aqui só para
        # testar o fluxo de envio; o mapeamento aba × produto avulso tem
        # sua própria suíte em GerarPlanilhaFaturamentoTests).
        RiItemIxc.objects.create(
            ri=self.ri,
            descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1,
            valor_unitario="0",
            eh_kit=True,
        )
        self.para_padrao = "hilber.lustosa@speedcsc.com.br, financeiro@speedcsc.com.br"
        self.cc_padrao = (
            "logistica-l@speedcsc.com.br, posvendas@megainfraestrutura.com.br, "
            "david.alves@speedcsc.com.br"
        )

    def _enviar_email(self, **campos):
        dados = {
            "para": self.para_padrao,
            "cc": self.cc_padrao,
            "assunto": f"Faturamento EACE — INEP {self.escola.inep}",
            "mensagem": "",
            "next": reverse("grid_inep"),
        }
        dados.update(campos)
        self.client.force_login(self.user)
        return self.client.post(
            reverse("ri_enviar_email_financeiro", kwargs={"pk": self.ri.pk}), dados
        )

    def test_enviar_email_bloqueado_fora_do_status(self):
        self.ri.status = Ri.ANDAMENTO
        self.ri.save(update_fields=["status"])
        self._enviar_email()
        self.assertEqual(len(mail.outbox), 0)

    def test_enviar_email_bloqueado_com_destinatario_invalido(self):
        self._enviar_email(para="nao-e-um-email")
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)
        self.assertEqual(len(mail.outbox), 0)

    def test_enviar_email_bloqueado_sem_destinatario(self):
        self._enviar_email(para="")
        self.assertEqual(len(mail.outbox), 0)

    def test_envio_completo(self):
        """RF-16→18: um só envio confirma a Mensagem e dispara o e-mail;
        confere Para/Cc editados pelo usuário na tela, código de rastreio
        preservado no Assunto (RN-009), planilha de faturamento anexada
        (RN-013, substitui o PDF), log em EmailFinanceiroLog, entrada na
        linha do tempo (FEAT-014) e transição automática de status."""
        self._enviar_email(mensagem="Kit substituído em campo.")

        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_FINANCEIRO)
        self.assertEqual(self.ri.observacoes_envio_financeiro, "Kit substituído em campo.")
        self.assertIsNotNone(self.ri.dados_financeiro_confirmados_em)

        self.assertEqual(len(mail.outbox), 1)
        enviado = mail.outbox[0]
        self.assertEqual(enviado.to, ["hilber.lustosa@speedcsc.com.br", "financeiro@speedcsc.com.br"])
        self.assertEqual(
            enviado.cc,
            [
                "logistica-l@speedcsc.com.br",
                "posvendas@megainfraestrutura.com.br",
                "david.alves@speedcsc.com.br",
            ],
        )
        self.assertIn("#RI-", enviado.subject)
        self.assertIn(self.escola.inep, enviado.subject)
        self.assertEqual(len(enviado.attachments), 1)
        nome_anexo, conteudo_anexo, tipo_anexo = enviado.attachments[0]
        # Pedido do usuário (2026-08-31): nome do anexo identifica a escola.
        self.assertEqual(nome_anexo, "FATURAMENTO MATERIAS EACE - 50000001 - Escola Financeiro.xlsx")
        self.assertEqual(
            tipo_anexo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertTrue(conteudo_anexo)

        log = EmailFinanceiroLog.objects.get(ri=self.ri)
        self.assertEqual(log.direcao, EmailFinanceiroLog.ENVIADO)
        self.assertIn("hilber.lustosa@speedcsc.com.br", log.destinatarios)

        entrada_email = RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.EMAIL)
        self.assertTrue(entrada_email.anexo.name)
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_STATUS, valor_novo="Aguardando financeiro"
            ).exists()
        )

    def test_envio_completo_com_produto_sem_aba_cadastrada_cria_aba_automaticamente(self):
        """RN-013 (ajuste 2026-08-26): produto lançado no Lado IXC sem
        aba cadastrada no catálogo NÃO bloqueia mais — ganha uma aba nova,
        criada na hora, e o e-mail sai normalmente."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Produto fora do catálogo", quantidade=1, valor_unitario="0"
        )
        resp = self._enviar_email()
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_FINANCEIRO)

        _, conteudo_anexo, _ = mail.outbox[0].attachments[0]
        workbook = openpyxl.load_workbook(BytesIO(conteudo_anexo))
        self.assertEqual(
            {titulo.strip() for titulo in workbook.sheetnames},
            {"NF KIT", "Produto fora do catálogo"},
        )

    def test_envio_bloqueado_sem_data_ativacao(self):
        """RN-013/RN-014 (2026-08-26): KIT/Data de Ativação/Município/
        Estado do Lado IXC são exigidos só na hora de enviar — RI com KIT
        lançado (setUp) mas sem Data de Ativação ainda bloqueia."""
        self.ri.data_ativacao = None
        self.ri.save(update_fields=["data_ativacao"])
        resp = self._enviar_email()
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(len(mail.outbox), 0)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_usuario_pode_editar_destinatarios_e_assunto(self):
        """Decisão do usuário (2026-08-24): Para/Cc/Assunto vêm
        pré-preenchidos, mas deixam de ser estritamente fixos."""
        self._enviar_email(
            para="outro-financeiro@speedcsc.com.br",
            cc="",
            assunto="Assunto customizado pelo analista",
        )
        self.assertEqual(len(mail.outbox), 1)
        enviado = mail.outbox[0]
        self.assertEqual(enviado.to, ["outro-financeiro@speedcsc.com.br"])
        self.assertEqual(enviado.cc, [])
        # RN-009: o código de rastreio é garantido mesmo quando o usuário
        # substitui o assunto sugerido por um texto próprio.
        self.assertIn("#RI-", enviado.subject)
        self.assertIn("Assunto customizado pelo analista", enviado.subject)

    def test_anexo_extra_e_adicional_a_planilha_automatica(self):
        """Decisão do usuário (2026-08-24, mantida na troca do anexo para
        planilha em 2026-08-26): o campo Anexo soma um arquivo opcional à
        planilha gerada automaticamente (RN-013), não a substitui."""
        arquivo_extra = SimpleUploadedFile(
            "comprovante.txt", b"conteudo do comprovante", content_type="text/plain"
        )
        self._enviar_email(anexo_extra=arquivo_extra)
        enviado = mail.outbox[0]
        self.assertEqual(len(enviado.attachments), 2)
        nomes_anexos = [nome for nome, _, _ in enviado.attachments]
        self.assertIn("comprovante.txt", nomes_anexos)
        self.assertTrue(any(nome.endswith(".xlsx") for nome in nomes_anexos))

    def test_um_email_por_inep_sem_lote(self):
        """RF-16-18: a rota é sempre por `pk` de um RI — não existe envio
        em lote nesta feature."""
        with self.assertRaises(NoReverseMatch):
            reverse("ri_enviar_email_financeiro_lote")

    def test_baixar_planilha_gera_arquivo_sem_enviar_email(self):
        """FEAT-017: botão "Baixar planilha" gera a mesma planilha que
        seria anexada, sem enviar e-mail nem mudar o status do RI."""
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("ri_baixar_planilha_financeiro", kwargs={"pk": self.ri.pk})
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", resp["Content-Disposition"])
        # Pedido do usuário (2026-08-31): nome do arquivo identifica a escola.
        self.assertIn(
            'filename="FATURAMENTO MATERIAS EACE - 50000001 - Escola Financeiro.xlsx"',
            resp["Content-Disposition"],
        )
        self.assertTrue(resp.content)
        self.assertEqual(len(mail.outbox), 0)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_baixar_planilha_com_produto_sem_aba_cadastrada_cria_aba_automaticamente(self):
        """FEAT-017 (ajuste 2026-08-26): baixa normalmente mesmo com
        produto sem aba cadastrada — a aba nasce na hora, com o nome do
        produto."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Produto fora do catálogo", quantidade=1, valor_unitario="0"
        )
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("ri_baixar_planilha_financeiro", kwargs={"pk": self.ri.pk})
        )
        self.assertEqual(resp.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(resp.content))
        self.assertIn("Produto fora do catálogo", {t.strip() for t in workbook.sheetnames})

    def test_baixar_planilha_bloqueada_sem_nenhum_item_lancado(self):
        ri_vazio = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("ri_baixar_planilha_financeiro", kwargs={"pk": ri_vazio.pk}), follow=True
        )
        self.assertContains(resp, "nenhum item lançado")

    def test_baixar_planilha_vencimento_e_30_dias_apos_geracao(self):
        """Correção 2026-08-31: o VENCIMENTO (E10) não é a data em que a
        planilha foi gerada — é a data de geração + 30 dias corridos."""
        self.client.force_login(self.user)
        resp = self.client.get(
            reverse("ri_baixar_planilha_financeiro", kwargs={"pk": self.ri.pk})
        )
        workbook = openpyxl.load_workbook(BytesIO(resp.content))
        aba = workbook[next(n for n in workbook.sheetnames if n.strip() == "NF KIT")]
        self.assertEqual(aba["E10"].value.date(), timezone.localdate() + timedelta(days=30))

    def test_enviar_email_vencimento_e_30_dias_apos_geracao(self):
        """Mesma correção do teste acima, mas no fluxo de envio de e-mail
        (a planilha anexada usa a mesma função de geração)."""
        self._enviar_email()
        _, conteudo_anexo, _ = mail.outbox[0].attachments[0]
        workbook = openpyxl.load_workbook(BytesIO(conteudo_anexo))
        aba = workbook[next(n for n in workbook.sheetnames if n.strip() == "NF KIT")]
        self.assertEqual(aba["E10"].value.date(), timezone.localdate() + timedelta(days=30))


class GerarPlanilhaFaturamentoTests(TestCase):
    """FEAT-017/RN-013: geração da planilha de faturamento a partir da
    planilha-modelo real (`doc/FATURAMENTO MATERIAS EACE.xlsx`) — mapeamento
    de célula, texto fixo preservado e bloqueio por produto sem aba."""

    def setUp(self):
        self.escola = Escola.objects.create(
            inep="60000001",
            nome="Escola Planilha Teste",
            endereco="Rua Teste, 123",
            municipio="Recife",
            estado="PE",
            lote=9,
        )
        self.ri = Ri.objects.create(
            escola=self.escola,
            status=Ri.ENVIO_EMAIL_FATURAMENTO,
            municipio_ixc="Recife",
            estado_ixc="PE",
            data_ativacao=date(2026, 8, 1),
            cnpj="00.000.000/0001-00",
            cnpj_ficticio="11.111.111/0001-11",
        )
        KitPadrao.objects.create(
            descricao="Rack 5U", lote=9, unidade="Unidade",
            valor_equipamento="500.00", valor_servico="100.00",
            aba_planilha_financeiro="RACK",
        )
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
            valor_equipamento="3000.00", valor_servico="500.00",
        )
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Rack 5U", quantidade=3, valor_unitario="0", eh_kit=False,
        )

    def test_gera_uma_aba_por_produto_com_celulas_mapeadas(self):
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))

        # RN-013: só as abas dos produtos lançados — as demais abas do
        # modelo (NOBREAK, SWITCH, CONVERSOR, ...) não entram na cópia.
        self.assertEqual({titulo.strip() for titulo in workbook.sheetnames}, {"NF KIT", "RACK"})

        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        # openpyxl devolve datetime (não date) ao reabrir uma célula com
        # formato de data — mesmo valor, tipo diferente no round-trip.
        self.assertEqual(aba_kit["E10"].value.date(), date(2026, 9, 21))
        self.assertEqual(aba_kit["H10"].value, 3000.0)  # só equipamento (3000) x 1
        # RN-048: CNPJ/CNPJ Fictício do Lado IXC, na mesma linha 16.
        self.assertEqual(aba_kit["A16"].value, "00.000.000/0001-00")
        self.assertEqual(aba_kit["B16"].value, "11.111.111/0001-11")
        self.assertEqual(aba_kit["C16"].value, "Escola Planilha Teste")
        self.assertEqual(aba_kit["F16"].value, "Rua Teste, 123")
        self.assertEqual(aba_kit["G16"].value, "Recife")
        self.assertEqual(aba_kit["H16"].value, "PE")
        self.assertEqual(aba_kit["I16"].value, "KIT 4")
        texto = aba_kit["F10"].value
        self.assertIn("INEP: 60000001", texto)
        self.assertIn("ITEM LPU: KIT 4", texto)
        self.assertIn("MUNICIPIO/UF: Recife/PE", texto)
        self.assertIn("VENCIMENTO: 21/09/2026", texto)
        # RN-013: resto do texto fixo do modelo (nº de contrato, texto
        # legal) preservado, não é gerado pelo sistema.
        self.assertIn("Mercadoria remetida para a Escola indicada", texto)
        self.assertIn("Edital de Licitação", texto)
        # Logo do financeiro presente na aba (já vem assim na planilha-
        # modelo original — essa aba não passa por clonagem).
        self.assertEqual(len(aba_kit._images), 1)

        aba_rack = next(ws for ws in workbook.worksheets if ws.title.strip() == "RACK")
        self.assertEqual(aba_rack["H10"].value, 1500.0)  # só equipamento (500) x 3
        # RN-013 (ajuste 2026-08-31): produto avulso com mais de 1 unidade
        # mostra a quantidade entre parênteses na I16 — RACK tem 3 aqui.
        self.assertEqual(aba_rack["I16"].value, "RACK (3)")
        # F10 (texto copiado pra Nota Fiscal) NÃO leva o sufixo — só a I16.
        self.assertIn("ITEM LPU: RACK", aba_rack["F10"].value)
        self.assertNotIn("ITEM LPU: RACK (3)", aba_rack["F10"].value)
        self.assertEqual(len(aba_rack._images), 1)

    def test_a20_usa_mes_selecionado_do_lado_ixc_e_ano_atual(self):
        """RN-053: célula A20 ("OPERAÇÃO COMPRA E VENDA - <MÊS>/<ANO>")
        usa o mês salvo em `Ri.mes_operacao_ixc` — ano nunca vem do RI, é
        sempre o corrente na hora de gerar."""
        self.ri.mes_operacao_ixc = 8  # Agosto
        self.ri.save(update_fields=["mes_operacao_ixc"])
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        ano_atual = timezone.now().year
        self.assertEqual(
            aba_kit["A20"].value, f"OPERAÇÃO COMPRA E VENDA  - AGOSTO/{ano_atual}"
        )

    def test_a20_sem_mes_salvo_usa_mes_corrente(self):
        """RN-053: RI antigo, nunca reaberto pelo Lado IXC depois desta
        feature (`mes_operacao_ixc` nulo) — cai no mês corrente, mesmo
        fallback usado para pré-preencher o select."""
        self.assertIsNone(self.ri.mes_operacao_ixc)
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        agora = timezone.now()
        mes_nome = dict(Ri.MESES_OPERACAO_CHOICES)[agora.month]
        self.assertEqual(
            aba_kit["A20"].value, f"OPERAÇÃO COMPRA E VENDA  - {mes_nome.upper()}/{agora.year}"
        )

    def test_aba_criada_automaticamente_herda_linhas_de_grade_ocultas(self):
        """RN-054: aba clonada na hora (produto sem aba cadastrada) copia
        a configuração de grade da aba-modelo (oculta em toda aba do
        modelo) — sem a correção, `Workbook.copy_worksheet` (openpyxl)
        nascia com a grade visível, diferente da "NF KIT"."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Enlace de Rádio",
            quantidade=1, valor_unitario="0", eh_kit=False,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        aba_nova = workbook["Enlace de Rádio"]
        self.assertFalse(aba_kit.sheet_view.showGridLines)
        self.assertEqual(aba_nova.sheet_view.showGridLines, aba_kit.sheet_view.showGridLines)

    def test_soma_subtotal_de_produtos_diferentes_na_mesma_aba(self):
        """RN-013: 2 produtos diferentes do catálogo que apontam para a
        MESMA aba (ex.: 2 tamanhos de Rack) somam o subtotal de cada um —
        não só a quantidade de um dos dois."""
        KitPadrao.objects.create(
            descricao="Rack 9U", lote=9, unidade="Unidade",
            valor_equipamento="900.00", valor_servico="0",
            aba_planilha_financeiro="RACK",
        )
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Rack 9U", quantidade=1, valor_unitario="0", eh_kit=False,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_rack = next(ws for ws in workbook.worksheets if ws.title.strip() == "RACK")
        # Só equipamento: Rack 5U: 500 x 3 = 1500; Rack 9U: 900 x 1 = 900; soma = 2400.
        self.assertEqual(aba_rack["H10"].value, 2400.0)

    def test_kit_fora_do_catalogo_nao_inventa_valor(self):
        """RN-013/CLAUDE.md §9: KIT sem correspondência no catálogo (ex.:
        opção "Outro", RN-011) ainda gera a aba, com valor 0 — não inventa
        preço para um kit que não está cadastrado."""
        RiItemIxc.objects.all().delete()
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 99 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        self.assertEqual(aba_kit["H10"].value, 0.0)
        self.assertIn("ITEM LPU: KIT 99", aba_kit["F10"].value)

    def test_produto_sem_aba_cadastrada_cria_aba_com_nome_do_proprio_produto(self):
        """RN-013 (ajuste 2026-08-26): produto sem `aba_planilha_
        financeiro` no catálogo (ou sem entrada nenhuma no catálogo) não
        bloqueia mais — ganha uma aba nova, com o nome dele, clonada do
        layout de uma aba já existente na planilha-modelo."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Enlace de Rádio",
            quantidade=2, valor_unitario="0", eh_kit=False,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        self.assertIn("Enlace de Rádio", workbook.sheetnames)
        aba_nova = workbook["Enlace de Rádio"]
        self.assertIn("ITEM LPU: Enlace de Rádio", aba_nova["F10"].value)
        # Sem entrada no catálogo para "Enlace de Rádio" → valor 0, não
        # inventado (CLAUDE.md §9) — mas a aba é gerada do mesmo jeito.
        self.assertEqual(aba_nova["H10"].value, 0.0)
        # Correção (2026-08-26): openpyxl.copy_worksheet não copia imagem
        # — a logo do financeiro precisa ir também na aba criada na hora,
        # não só na aba original usada como modelo (usuário reportou que
        # sumia da 2ª aba em diante).
        self.assertEqual(len(aba_nova._images), 1)
        # Layout clonado de uma aba real do modelo — mesma estrutura
        # (fórmula do H12, texto legal fixo) preservada na aba nova.
        self.assertEqual(aba_nova["H12"].value, "=SUM(H10:H11)")
        self.assertIn("Edital de Licitação", aba_nova["F10"].value)

    def test_nome_de_aba_muito_longo_e_truncado_em_31_caracteres(self):
        """Excel limita nome de aba a 31 caracteres — catálogo real tem
        produto com nome bem maior (ex.: "Implantação de postes para rede
        de acesso em fibra óptica - concreto")."""
        nome_longo = "Implantação de postes para rede de acesso em fibra óptica - concreto"
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item=nome_longo, quantidade=1, valor_unitario="0", eh_kit=False,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        abas_novas = [t for t in workbook.sheetnames if t.strip() not in ("NF KIT", "RACK")]
        self.assertEqual(len(abas_novas), 1)
        self.assertLessEqual(len(abas_novas[0]), 31)
        # O texto completo (sem limite de 31 caracteres) continua correto
        # no ITEM LPU dentro da célula — só o NOME da aba é truncado.
        self.assertIn(f"ITEM LPU: {nome_longo}", workbook[abas_novas[0]]["F10"].value)

    def test_ri_sem_nenhum_item_lancado_bloqueia(self):
        ri_vazio = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        with self.assertRaises(PlanilhaFaturamentoError):
            gerar_planilha_faturamento(ri_vazio, data_vencimento=date(2026, 9, 21))

    def test_ri_sem_kit_lancado_bloqueia_mesmo_com_produto(self):
        """RN-013/RN-014 (2026-08-26): KIT é exigido só na hora de gerar a
        planilha (não a cada "Salvar" do Lado IXC) — RI com só Produtos
        lançados, sem nenhum KIT, ainda bloqueia aqui."""
        self.ri.itens_ixc.filter(eh_kit=True).delete()
        with self.assertRaises(PlanilhaFaturamentoError) as contexto:
            gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        self.assertIn("o KIT Instalado", str(contexto.exception))

    def test_ri_sem_data_ativacao_bloqueia(self):
        self.ri.data_ativacao = None
        self.ri.save(update_fields=["data_ativacao"])
        with self.assertRaises(PlanilhaFaturamentoError) as contexto:
            gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        self.assertIn("a Data de Ativação", str(contexto.exception))

    def test_ri_sem_municipio_e_estado_ixc_bloqueia_com_mensagem_unica(self):
        self.ri.municipio_ixc = ""
        self.ri.estado_ixc = ""
        self.ri.save(update_fields=["municipio_ixc", "estado_ixc"])
        with self.assertRaises(PlanilhaFaturamentoError) as contexto:
            gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        mensagem = str(contexto.exception)
        self.assertIn("o Município (Lado IXC)", mensagem)
        self.assertIn("o Estado (Lado IXC)", mensagem)

    def test_ri_sem_cnpj_e_cnpj_ficticio_bloqueia_com_mensagem_unica(self):
        """RN-048 (2026-09-01): mesmo padrão de bloqueio de Município/Estado
        acima — exigidos só na hora de gerar a planilha, não a cada
        "Salvar" do Lado IXC."""
        self.ri.cnpj = ""
        self.ri.cnpj_ficticio = ""
        self.ri.save(update_fields=["cnpj", "cnpj_ficticio"])
        with self.assertRaises(PlanilhaFaturamentoError) as contexto:
            gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        mensagem = str(contexto.exception)
        self.assertIn("o CNPJ (Lado IXC)", mensagem)
        self.assertIn("o CNPJ Fictício (Lado IXC)", mensagem)

    def test_i16_sem_sufixo_de_quantidade_com_1_unica_unidade(self):
        """Pedido do usuário (2026-08-31): o sufixo "(N)" só aparece com
        mais de 1 equipamento — 1 unidade continua só com o nome."""
        RiItemIxc.objects.filter(descricao_item="Rack 5U").update(quantidade=1)
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_rack = next(ws for ws in workbook.worksheets if ws.title.strip() == "RACK")
        self.assertEqual(aba_rack["I16"].value, "RACK")

    def test_i16_kit_nunca_leva_sufixo_de_quantidade(self):
        """Pedido do usuário (2026-08-31): o sufixo "(N)" é só para
        equipamento avulso — KIT nunca leva, mesmo somando mais de 1
        unidade na mesma aba (RI corrigido trocando o tamanho do KIT,
        cenário já previsto no agrupamento por aba)."""
        RiItemIxc.objects.create(
            ri=self.ri,
            descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=2,
            valor_unitario="0",
            eh_kit=True,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_kit = next(ws for ws in workbook.worksheets if ws.title.strip() == "NF KIT")
        self.assertEqual(aba_kit["I16"].value, "KIT 4")

    def test_i16_produto_sem_aba_cadastrada_tambem_leva_sufixo_de_quantidade(self):
        """A regra da quantidade vale tanto pra aba de catálogo (RACK)
        quanto pra aba criada na hora pra produto sem cadastro."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Enlace de Rádio",
            quantidade=2, valor_unitario="0", eh_kit=False,
        )
        conteudo = gerar_planilha_faturamento(self.ri, data_vencimento=date(2026, 9, 21))
        workbook = openpyxl.load_workbook(BytesIO(conteudo))
        aba_nova = workbook["Enlace de Rádio"]
        self.assertEqual(aba_nova["I16"].value, "Enlace de Rádio (2)")


class NomeArquivoPlanilhaFaturamentoTests(TestCase):
    """Pedido do usuário (2026-08-31): nome do .xlsx anexado no e-mail (e
    baixado pelo botão "Baixar planilha") passa a identificar a escola."""

    def test_monta_nome_com_inep_e_nome_da_escola(self):
        escola = Escola.objects.create(inep="70000001", nome="Escola Modelo", municipio="Recife", estado="PE")
        self.assertEqual(
            nome_arquivo_planilha_faturamento(escola),
            "FATURAMENTO MATERIAS EACE - 70000001 - Escola Modelo.xlsx",
        )

    def test_remove_caractere_invalido_de_nome_de_arquivo(self):
        escola = Escola.objects.create(
            inep="70000002", nome='Escola "Teste"/Especial', municipio="Recife", estado="PE"
        )
        nome = nome_arquivo_planilha_faturamento(escola)
        self.assertNotIn('"', nome)
        self.assertNotIn("/", nome)

    def test_nome_de_escola_muito_longo_e_truncado_sem_quebrar(self):
        """Escola.nome permite até 255 caracteres (há escola real
        cadastrada com 100) — o nome do arquivo não pode estourar o
        `RiHistorico.anexo` (FileField max_length=255, com o prefixo do
        `upload_to` e a folga de deduplicação do Django)."""
        nome_bem_longo = "Escola " + "Muito Comprida " * 15  # 232 chars, dentro do limite da Escola.nome (255)
        escola = Escola.objects.create(
            inep="70000003", nome=nome_bem_longo, municipio="Recife", estado="PE"
        )
        nome = nome_arquivo_planilha_faturamento(escola)
        self.assertLessEqual(len(nome), 200)
        self.assertTrue(nome.startswith("FATURAMENTO MATERIAS EACE - 70000003 - Escola"))
        self.assertTrue(nome.endswith(".xlsx"))


class MontarCorpoEmailFinanceiroTests(TestCase):
    """RN-013: o corpo do e-mail usa a mesma origem de preço da planilha de
    faturamento — o catálogo `KitPadrao`, nunca `RiItemIxc.valor_unitario`
    (nasce 0,00, RN-011). Bug reportado: item e KIT saíam com valor R$ 0,00
    no corpo do e-mail mesmo já cadastrados no catálogo."""

    def setUp(self):
        self.escola = Escola.objects.create(
            inep="60000002",
            nome="Escola Corpo E-mail Teste",
            endereco="Rua Teste, 456",
            municipio="Bragança Paulista",
            estado="SP",
            lote=9,
        )
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        KitPadrao.objects.create(
            descricao="Rack 5U", lote=9, unidade="Unidade",
            valor_equipamento="500.00", valor_servico="100.00",
        )
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
            valor_equipamento="3000.00", valor_servico="500.00",
        )

    def test_usa_valor_do_catalogo_em_vez_do_valor_zerado_do_item(self):
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Rack 5U", quantidade=3, valor_unitario="0", eh_kit=False,
        )
        corpo = montar_corpo_email_financeiro(self.ri)
        self.assertIn("Kit Cobertura Wi-Fi - 4 Access Points — 1 un. — R$ 3000.00 — R$ 3000.00", corpo)
        self.assertIn("Rack 5U — 3 un. — R$ 500.00 — R$ 1500.00", corpo)
        self.assertIn("Valor total: R$ 4500.00", corpo)

    def test_item_fora_do_catalogo_nao_inventa_valor(self):
        """RN-013/CLAUDE.md §9: item sem correspondência no catálogo (ex.:
        opção "Outro", RN-011) mostra valor 0 — não inventa preço."""
        RiItemIxc.objects.create(
            ri=self.ri, descricao_item="Serviço fora do catálogo",
            quantidade=2, valor_unitario="0", eh_kit=False,
        )
        corpo = montar_corpo_email_financeiro(self.ri)
        self.assertIn("Serviço fora do catálogo — 2 un. — R$ 0.00 — R$ 0.00", corpo)
        self.assertIn("Valor total: R$ 0.00", corpo)


def _montar_email_bytes(assunto, remetente="financeiro@speedcsc.com.br", anexos=None):
    """Monta os bytes RFC822 de um e-mail de teste (RN-005/RN-009) — mesmo
    formato que o Microsoft Graph devolve num `GET .../$value`."""
    mensagem = MensagemEmailMime()
    mensagem["Subject"] = assunto
    mensagem["From"] = remetente
    mensagem["To"] = "posvendas@megainfraestrutura.com.br"
    mensagem.set_content("Segue em anexo.")
    for nome, tipo_principal, subtipo, payload in anexos or []:
        mensagem.add_attachment(payload, maintype=tipo_principal, subtype=subtipo, filename=nome)
    return mensagem.as_bytes()


class _RespostaGraphFake:
    """Dublê do `requests.Response` devolvido por `_graph_get` — só o
    `.json()` importa para `sincronizar_respostas_financeiro`."""

    def __init__(self, payload):
        self._payload = payload

    def json(self):
        return self._payload


_CONFIG_GRAPH_FINANCEIRO_TESTE = {
    "GRAPH_FINANCEIRO_ENABLED": True,
    "GRAPH_FINANCEIRO_CLIENT_ID": "client-id-de-teste",
    "GRAPH_FINANCEIRO_CLIENT_SECRET": "client-secret-de-teste",
    "GRAPH_FINANCEIRO_TENANT_ID": "tenant-id-de-teste",
    "GRAPH_FINANCEIRO_MAILBOX": "posvendas@megainfraestrutura.com.br",
}

_MEDIA_ROOT_TESTE_SYNC_FINANCEIRO = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SYNC_FINANCEIRO, **_CONFIG_GRAPH_FINANCEIRO_TESTE)
class SincronizarEmailFinanceiroTests(TestCase):
    """FEAT-009/FEAT-020 (RF-08/RF-09/RF-19, RN-005/RN-009/RN-016): uma
    passada de polling (delta query do Microsoft Graph — IMAP com
    usuário/senha não funciona mais nessa caixa, testado em 2026-08-25)
    identifica o RI pelo código de rastreio do assunto, confirma que o
    remetente é do domínio do financeiro (RN-016, correção 2026-09-02),
    valida a estrutura da resposta (1 PDF + 1 XML) e avança o status para
    "Resposta Financeiro" — só quando está no padrão é que anexa os
    documentos. Fora do padrão só gera alerta (mas avança o status do mesmo
    jeito, RN-016); sem código, sem RI aguardando ou remetente fora do
    domínio do financeiro não altera nada. Credenciais e chamadas de rede
    são todas dublês — não há como (nem se deve) testar contra o Graph de
    verdade."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_SYNC_FINANCEIRO, ignore_errors=True)

    def setUp(self):
        self.escola = Escola.objects.create(inep="50000002", nome="Escola Financeiro 2")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.codigo = montar_codigo_rastreio(self.escola.inep, timezone.localdate())
        self.assunto_padrao = f"#{self.codigo} - Faturamento EACE — INEP {self.escola.inep}"

    def _rodar_sync(self, mensagens):
        """`mensagens`: lista de `(internet_message_id, bytes_rfc822)`. Cada
        chamada devolve uma única página com `@odata.deltaLink` (sem
        paginação — já coberto pelo teste de persistência do cursor)."""
        itens = [
            {"id": f"graph-id-{indice}", "internetMessageId": message_id}
            for indice, (message_id, _) in enumerate(mensagens)
        ]
        bytes_por_id = {f"graph-id-{indice}": bruto for indice, (_, bruto) in enumerate(mensagens)}
        resposta = _RespostaGraphFake(
            {"value": itens, "@odata.deltaLink": "https://graph.microsoft.com/v1.0/.../delta?token=abc"}
        )

        with patch("apps.ri.services._obter_token", return_value="token-de-teste"), patch(
            "apps.ri.services._graph_get", return_value=resposta
        ), patch("apps.ri.services._buscar_mime", side_effect=lambda caixa, id_msg, token: bytes_por_id[id_msg]):
            resultado = sincronizar_respostas_financeiro()
        return resultado

    def test_resposta_padrao_anexa_documentos_e_avanca_status(self):
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("nota_fiscal.pdf", "application", "pdf", b"%PDF-1.4 conteudo"),
                ("nota_fiscal.xml", "text", "xml", b"<nfe></nfe>"),
            ],
        )
        resultado = self._rodar_sync([("<msg-1@financeiro>", bruto)])

        self.assertEqual(resultado, {
            "processados": 1, "identificados": 1, "fora_do_padrao": 0,
            "sem_codigo": 0, "sem_ri_aguardando": 0,
            "remetente_nao_reconhecido": 0, "duplicados": 0,
        })
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.assertEqual(
            Documento.objects.filter(ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF, ativo=True).count(), 1
        )
        self.assertEqual(Documento.objects.filter(ri=self.ri, tipo=Documento.XML, ativo=True).count(), 1)

        log = EmailFinanceiroLog.objects.get(ri=self.ri)
        self.assertEqual(log.direcao, EmailFinanceiroLog.RECEBIDO)
        self.assertEqual(log.status_leitura, EmailFinanceiroLog.OK)
        self.assertEqual(log.mensagem_id_externo, "<msg-1@financeiro>")

        self.assertTrue(RiHistorico.objects.filter(ri=self.ri, tipo=RiHistorico.EMAIL).exists())
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_STATUS, valor_novo="Resposta Financeiro"
            ).exists()
        )

        # RN-008 (pendência 2026-08-26/27): PDF e XML da resposta ficam
        # disponíveis para download na própria entrada do e-mail na linha
        # do tempo (referência ao `Documento` já salvo, sem duplicar
        # arquivo nem espalhar em cards separados).
        entrada_email = RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.EMAIL)
        documentos_vinculados = list(entrada_email.documentos.order_by("tipo"))
        self.assertEqual([d.tipo for d in documentos_vinculados], [Documento.NOTA_FISCAL_PDF, Documento.XML])
        self.assertEqual(
            {d.tipo: d.arquivo.read() for d in documentos_vinculados},
            {Documento.NOTA_FISCAL_PDF: b"%PDF-1.4 conteudo", Documento.XML: b"<nfe></nfe>"},
        )
        self.assertFalse(RiHistorico.objects.filter(ri=self.ri, tipo=RiHistorico.ANEXO).exists())

    def test_resposta_padrao_cria_1_log_rpa_eace_por_nota_fiscal(self):
        """RN-056 (FEAT-033, Fase 2): resposta com 1 PDF + 1 XML cria 1 log,
        sem nenhum Documento pré-selecionado (o usuário escolhe na tela)."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("nota_fiscal.pdf", "application", "pdf", b"%PDF-1.4 conteudo"),
                ("nota_fiscal.xml", "text", "xml", b"<nfe></nfe>"),
            ],
        )
        self._rodar_sync([("<msg-log-1@financeiro>", bruto)])

        self.assertEqual(LogRpaEace.objects.filter(ri=self.ri).count(), 1)
        log = LogRpaEace.objects.get(ri=self.ri)
        self.assertIsNone(log.documento_pdf)
        self.assertIsNone(log.documento_xml)
        self.assertEqual(log.resultado, LogRpaEace.PENDENTE)

    def test_resposta_com_2_notas_fiscais_cria_2_logs(self):
        """RN-005/RN-056: e-mail com N PDF/N XML (N Notas Fiscais no mesmo
        e-mail) cria N logs, um por Nota Fiscal esperada."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("nota1.pdf", "application", "pdf", b"%PDF nota 1"),
                ("nota2.pdf", "application", "pdf", b"%PDF nota 2"),
                ("nota1.xml", "text", "xml", b"<nfe>1</nfe>"),
                ("nota2.xml", "text", "xml", b"<nfe>2</nfe>"),
            ],
        )
        self._rodar_sync([("<msg-log-2@financeiro>", bruto)])

        self.assertEqual(LogRpaEace.objects.filter(ri=self.ri).count(), 2)

    def test_resposta_fora_do_padrao_nao_cria_log(self):
        """RN-016: resposta fora do padrão (quantidade de PDF diferente da
        de XML) avança o status, mas não anexa Documento nem cria log —
        não há como saber quantas Notas Fiscais esperar."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[("nota_fiscal.pdf", "application", "pdf", b"%PDF-1.4 conteudo")],
        )
        self._rodar_sync([("<msg-log-3@financeiro>", bruto)])

        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.assertEqual(LogRpaEace.objects.filter(ri=self.ri).count(), 0)

    def test_email_sem_codigo_de_rastreio_nao_altera_nada(self):
        bruto = _montar_email_bytes("Assunto qualquer, sem código de rastreio")
        resultado = self._rodar_sync([("<msg-2@financeiro>", bruto)])

        self.assertEqual(resultado["sem_codigo"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_FINANCEIRO)
        self.assertFalse(EmailFinanceiroLog.objects.exists())

    def test_email_com_inep_sem_ri_aguardando_financeiro_nao_altera_nada(self):
        outro_codigo = montar_codigo_rastreio("99999999", timezone.localdate())
        bruto = _montar_email_bytes(f"#{outro_codigo} - Assunto qualquer")
        resultado = self._rodar_sync([("<msg-3@financeiro>", bruto)])

        self.assertEqual(resultado["sem_ri_aguardando"], 1)
        self.assertFalse(EmailFinanceiroLog.objects.exists())

    def test_resposta_de_remetente_fora_do_dominio_do_financeiro_nao_altera_nada(self):
        """RN-016 (correção 2026-09-02): código de rastreio no assunto (RN-009)
        não basta — usuário reportou falso positivo (INEP 35271561) em que um
        e-mail de remetente qualquer (não do financeiro) avançava o status do
        RI. Remetente fora do domínio do financeiro não muda o status, não
        grava `EmailFinanceiroLog` nem entrada na linha do tempo — só alerta
        no log do servidor."""
        bruto = _montar_email_bytes(self.assunto_padrao, remetente="Elias Neto <eliasnetoce@gmail.com>")
        resultado = self._rodar_sync([("<msg-externo@gmail.com>", bruto)])

        self.assertEqual(resultado["remetente_nao_reconhecido"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_FINANCEIRO)
        self.assertFalse(EmailFinanceiroLog.objects.exists())
        self.assertFalse(RiHistorico.objects.filter(ri=self.ri).exists())

    def test_resposta_de_subdominio_diferente_nao_conta_como_financeiro(self):
        """RN-016: comparação de domínio é exata (`speedcsc.com.br`), não por
        substring — evita um domínio parecido (mas diferente) ser aceito por
        engano."""
        bruto = _montar_email_bytes(self.assunto_padrao, remetente="alguem@outrospeedcsc.com.br")
        resultado = self._rodar_sync([("<msg-externo-2@dominio-parecido>", bruto)])

        self.assertEqual(resultado["remetente_nao_reconhecido"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_FINANCEIRO)

    def test_resposta_fora_do_padrao_gera_alerta_e_avanca_status(self):
        """RN-005/RN-016: resposta com quantidade de PDF diferente da de XML
        não bloqueia o fluxo, só gera alerta — aqui, só o PDF veio anexado —
        mas, desde a RN-016, o status avança para "Resposta Financeiro" do
        mesmo jeito que a resposta no padrão, sem anexar os documentos."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[("nota_fiscal.pdf", "application", "pdf", b"%PDF-1.4 conteudo")],
        )
        resultado = self._rodar_sync([("<msg-4@financeiro>", bruto)])

        self.assertEqual(resultado["fora_do_padrao"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)
        self.assertFalse(Documento.objects.filter(ri=self.ri).exists())

        log = EmailFinanceiroLog.objects.get(ri=self.ri)
        self.assertEqual(log.status_leitura, EmailFinanceiroLog.FORA_DO_PADRAO)
        self.assertTrue(RiHistorico.objects.filter(ri=self.ri, tipo=RiHistorico.EMAIL).exists())
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_STATUS, valor_novo="Resposta Financeiro"
            ).exists()
        )
        # Fora do padrão não anexa nada (nem em `Documento`, nem na linha
        # do tempo) — só o alerta.
        self.assertFalse(RiHistorico.objects.filter(ri=self.ri, tipo=RiHistorico.ANEXO).exists())
        self.assertFalse(RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.EMAIL).documentos.exists())

    def test_resposta_com_mais_de_uma_nota_fiscal_anexa_todos_os_documentos(self):
        """RN-005 (correção 2026-09-02): financeiro pode responder com mais
        de 1 Nota Fiscal no mesmo e-mail (usuário reportou, INEP 35095874) —
        2 PDF + 2 XML é "no padrão" (quantidade igual, não precisa ser
        exatamente 1+1) e os 4 arquivos são salvos como `Documento`, sem
        substituir um pelo outro."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("1573.pdf", "application", "pdf", b"%PDF-1.4 nf 1573"),
                ("1573.xml", "text", "xml", b"<nfe>1573</nfe>"),
                ("1575.pdf", "application", "pdf", b"%PDF-1.4 nf 1575"),
                ("1575.xml", "text", "xml", b"<nfe>1575</nfe>"),
            ],
        )
        resultado = self._rodar_sync([("<msg-multi-nf@financeiro>", bruto)])

        self.assertEqual(resultado["identificados"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

        pdfs_ativos = Documento.objects.filter(ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF, ativo=True)
        xmls_ativos = Documento.objects.filter(ri=self.ri, tipo=Documento.XML, ativo=True)
        self.assertEqual(pdfs_ativos.count(), 2)
        self.assertEqual(xmls_ativos.count(), 2)

        log = EmailFinanceiroLog.objects.get(ri=self.ri)
        self.assertEqual(log.status_leitura, EmailFinanceiroLog.OK)

        entrada_email = RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.EMAIL)
        self.assertEqual(entrada_email.documentos.count(), 4)

    def test_resposta_com_cinco_notas_fiscais_anexa_todos_os_dez_documentos(self):
        """RN-005: não é um limite de 2 — qualquer quantidade N de Notas
        Fiscais no mesmo e-mail (N PDF + N XML) salva todos os pares,
        usuário pediu confirmação explícita de que N funciona, não só 2."""
        anexos = []
        for numero in range(1, 6):
            anexos.append((f"{numero}.pdf", "application", "pdf", f"%PDF-1.4 nf {numero}".encode()))
            anexos.append((f"{numero}.xml", "text", "xml", f"<nfe>{numero}</nfe>".encode()))
        bruto = _montar_email_bytes(self.assunto_padrao, anexos=anexos)
        resultado = self._rodar_sync([("<msg-cinco-nf@financeiro>", bruto)])

        self.assertEqual(resultado["identificados"], 1)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.AGUARDANDO_ANEXO_PORTAL_EACE)

        self.assertEqual(
            Documento.objects.filter(ri=self.ri, tipo=Documento.NOTA_FISCAL_PDF, ativo=True).count(), 5
        )
        self.assertEqual(Documento.objects.filter(ri=self.ri, tipo=Documento.XML, ativo=True).count(), 5)
        entrada_email = RiHistorico.objects.get(ri=self.ri, tipo=RiHistorico.EMAIL)
        self.assertEqual(entrada_email.documentos.count(), 10)

    def test_resposta_com_quantidade_de_pdf_diferente_de_xml_continua_fora_do_padrao(self):
        """RN-005: quantidade de PDF diferente da de XML continua fora do
        padrão (algo de fato incompleto), mesmo com mais de 1 arquivo — não
        é só "diferente de 1+1"."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("1573.pdf", "application", "pdf", b"%PDF-1.4 nf 1573"),
                ("1575.pdf", "application", "pdf", b"%PDF-1.4 nf 1575"),
                ("1573.xml", "text", "xml", b"<nfe>1573</nfe>"),
            ],
        )
        resultado = self._rodar_sync([("<msg-desbalanceado@financeiro>", bruto)])

        self.assertEqual(resultado["fora_do_padrao"], 1)
        self.assertFalse(Documento.objects.filter(ri=self.ri).exists())

    def test_mensagem_repetida_nao_e_processada_duas_vezes(self):
        """O delta query do Graph garante "ao menos uma vez", não
        "exatamente uma vez" — a mesma mensagem pode reaparecer numa
        passada seguinte; o dedup é pelo `internetMessageId`."""
        bruto = _montar_email_bytes(
            self.assunto_padrao,
            anexos=[
                ("nota_fiscal.pdf", "application", "pdf", b"%PDF-1.4 conteudo"),
                ("nota_fiscal.xml", "text", "xml", b"<nfe></nfe>"),
            ],
        )
        self._rodar_sync([("<msg-5@financeiro>", bruto)])
        resultado = self._rodar_sync([("<msg-5@financeiro>", bruto)])

        self.assertEqual(resultado["duplicados"], 1)
        self.assertEqual(EmailFinanceiroLog.objects.filter(ri=self.ri).count(), 1)

    def test_cursor_delta_link_e_persistido_entre_passadas(self):
        bruto = _montar_email_bytes(self.assunto_padrao)
        self._rodar_sync([("<msg-6@financeiro>", bruto)])

        estado = EmailFinanceiroSync.objects.get(mailbox="posvendas@megainfraestrutura.com.br")
        self.assertEqual(estado.delta_link, "https://graph.microsoft.com/v1.0/.../delta?token=abc")
        self.assertIsNotNone(estado.ultima_sincronizacao_em)
        self.assertEqual(estado.ultimo_erro, "")

    def test_falha_de_autenticacao_no_graph_nao_e_tratada_como_fora_do_padrao(self):
        """Erro de autenticação é uma falha de verdade (avisa o operador),
        diferente de "e-mail fora do padrão" (RN-005, que nunca bloqueia)."""
        with patch("apps.ri.services._obter_token", side_effect=EmailFinanceiroSyncError("falha")):
            with self.assertRaises(EmailFinanceiroSyncError):
                sincronizar_respostas_financeiro()
        estado = EmailFinanceiroSync.objects.get(mailbox="posvendas@megainfraestrutura.com.br")
        self.assertEqual(estado.ultimo_erro, "falha")

    def test_desabilitado_por_padrao_sem_credenciais(self):
        """Fora do `override_settings` da classe (sem credenciais
        configuradas) — o mesmo estado do `.env` local até existir um app do
        Azure AD dedicado a este sistema."""
        with override_settings(GRAPH_FINANCEIRO_ENABLED=False):
            with self.assertRaises(EmailFinanceiroSyncError):
                sincronizar_respostas_financeiro()


class SincronizarEmailFinanceiroCommandTests(TestCase):
    """FEAT-009: o comando (`manage.py sincronizar_email_financeiro`) é o
    que um agendador externo (cron/Task Scheduler, a cargo do DevOps) roda
    a cada ~5 min — aqui só confere que ele repassa o resultado."""

    @override_settings(**_CONFIG_GRAPH_FINANCEIRO_TESTE)
    def test_comando_reporta_resumo_da_passada(self):
        saida = StringIO()
        resposta_vazia = _RespostaGraphFake({"value": []})
        with patch("apps.ri.services._obter_token", return_value="token-de-teste"), patch(
            "apps.ri.services._graph_get", return_value=resposta_vazia
        ):
            call_command("sincronizar_email_financeiro", stdout=saida)
        self.assertIn("E-mails avaliados: 0", saida.getvalue())

    @override_settings(GRAPH_FINANCEIRO_ENABLED=False)
    def test_comando_reporta_erro_sem_quebrar_silenciosamente(self):
        with self.assertRaises(CommandError):
            call_command("sincronizar_email_financeiro")

    @override_settings(**_CONFIG_GRAPH_FINANCEIRO_TESTE)
    def test_comando_conta_fora_do_padrao_como_status_alterado(self):
        """RN-016: resposta fora do padrão também muda o status do RI —
        o resumo do comando não pode subestimar isso contando só a
        resposta no padrão (`identificados`)."""
        escola = Escola.objects.create(inep="50000003", nome="Escola Financeiro 3")
        Ri.objects.create(escola=escola, status=Ri.AGUARDANDO_FINANCEIRO)
        codigo = montar_codigo_rastreio(escola.inep, timezone.localdate())
        bruto = _montar_email_bytes(f"#{codigo} - Faturamento EACE — INEP {escola.inep}")
        resposta = _RespostaGraphFake({"value": [{"id": "graph-id-0", "internetMessageId": "<msg-x@financeiro>"}]})

        saida = StringIO()
        with patch("apps.ri.services._obter_token", return_value="token-de-teste"), patch(
            "apps.ri.services._graph_get", return_value=resposta
        ), patch("apps.ri.services._buscar_mime", return_value=bruto):
            call_command("sincronizar_email_financeiro", stdout=saida)

        self.assertIn("RIs com status alterado: 1", saida.getvalue())
        self.assertIn("documentos anexados: 0", saida.getvalue())
        self.assertIn("fora do padrão: 1", saida.getvalue())


class RiDetailViewTests(TestCase):
    """FEAT-004: cadastro manual de RI e itens EACE/IXC (RN-003, RN-004)."""

    # RN-011 (formulário único, 2026-08-24): "salvar_ixc" valida KIT,
    # Produtos e Data Ativação juntos — mesmo só testando o KIT, o POST
    # precisa do management form do formset de Produtos (senão Django
    # acusa "ManagementForm data is missing").
    FORMSET_PRODUTO_VAZIO = {
        "form-TOTAL_FORMS": "0",
        "form-INITIAL_FORMS": "0",
        "form-MIN_NUM_FORMS": "0",
        "form-MAX_NUM_FORMS": "1000",
    }

    # RN-018: mesma exigência acima, formset do Lado Relatório EACE
    # (prefixo próprio, "eace_produto", para não colidir com o do IXC).
    FORMSET_PRODUTO_EACE_VAZIO = {
        "eace_produto-TOTAL_FORMS": "0",
        "eace_produto-INITIAL_FORMS": "0",
        "eace_produto-MIN_NUM_FORMS": "0",
        "eace_produto-MAX_NUM_FORMS": "1000",
    }

    def setUp(self):
        self.analista = User.objects.create_user(
            username="analista", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.admin = User.objects.create_user(
            username="admin-ri", password="senha-teste-123", perfil=User.PERFIL_ADMINISTRADOR
        )
        self.escola = Escola.objects.create(
            inep="20000001",
            nome="Escola Teste RI",
            municipio="Fortaleza",
            estado="CE",
            kit_inicial="Kit Wi-Fi Indoor",
        )

    def test_exige_login(self):
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("login"), resp.url)

    def test_sem_ri_mostra_acao_de_iniciar(self):
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Nenhum RI iniciado")
        self.assertContains(resp, "Iniciar RI")

    def test_iniciar_ri_cria_com_status_implantacao_eace(self):
        self.client.force_login(self.analista)
        self.client.post(reverse("ri_iniciar", kwargs={"inep": self.escola.inep}))
        ri = Ri.objects.get(escola=self.escola)
        self.assertEqual(ri.status, Ri.IMPLANTACAO_EACE)
        self.assertEqual(ri.responsavel, self.analista)

    def test_iniciar_ri_nao_duplica_quando_ja_existe(self):
        self.client.force_login(self.analista)
        self.client.post(reverse("ri_iniciar", kwargs={"inep": self.escola.inep}))
        self.client.post(reverse("ri_iniciar", kwargs={"inep": self.escola.inep}))
        self.assertEqual(Ri.objects.filter(escola=self.escola).count(), 1)

    def test_painel_kit_declarado_mostra_kit_inicial_sem_botao_de_lancar(self):
        """RN-010 (2026-08-24): não há mais lançamento manual nesta tela —
        o painel só exibe o dado já informado pela EACE (Escola.kit_inicial).
        Criação do item é automática (Lote 1, fora desta tela) ou feita
        pelo administrador via Django admin (Lote 2/3, decisão do usuário)."""
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Kit Wi-Fi Indoor")

    def test_painel_kit_declarado_mostra_nobreak_declarado(self):
        """RN-017: Nobreak declarado aparece no mesmo card do Kit (1º
        lado), item padrão igual para toda escola, sempre quantidade 1."""
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Nobreak declarado")
        self.assertContains(resp, "(1 un.)")
        self.assertContains(resp, 'value="Nobreak"')

    def test_painel_kit_declarado_mostra_item_ja_lancado_no_admin(self):
        """Bug reportado pelo usuário (2026-08-26): o card só mostrava a
        pré-visualização de Escola.kit_inicial, nunca o item de RiItemEace
        de fato lançado pelo administrador — divergindo do que o Grid de
        INEPs (FEAT-007) já exibe para o mesmo 1º lado. Painel agora lista
        também o item real, junto com a pré-visualização de referência."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        RiItemEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=2,
            valor_unitario=Decimal("350.00"),
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "2 un. — R$ 350,00")

    def test_painel_kit_declarado_com_item_lancado_nao_diverge_da_pre_visualizacao(self):
        """Bug reportado pelo usuário (2026-08-26): com o item do 1º lado
        já lançado, o campo "Descrição do item" mostrava a pré-visualização
        calculada de Escola.kit_inicial (podendo ter ficado desatualizada),
        não o item de fato registrado — as duas informações divergiam na
        mesma tela. Com item lançado, o campo passa a refletir esse item."""
        escola = Escola.objects.create(
            inep="20000004", nome="Escola Teste RI 4", municipio="Fortaleza",
            estado="CE", kit_inicial="Kit Wi-Fi Indoor",
        )
        # Catálogo com uma entrada diferente do que foi de fato lançado —
        # simula Escola.kit_inicial desatualizado/divergente do lançamento.
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 15 Access Points", unidade="Escola",
        )
        ri = Ri.objects.create(escola=escola, status=Ri.IMPLANTACAO_EACE)
        RiItemEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=2,
            valor_unitario=Decimal("350.00"),
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": escola.inep}))
        self.assertContains(resp, 'value="Kit Wi-Fi Indoor"')
        # a descrição do catálogo ainda pode aparecer como opção do select
        # "KIT Instalado" (Lado IXC) — o que não pode é ser o valor deste
        # campo, por isso o assert é sobre o atributo `value`, não o texto solto.
        self.assertNotContains(resp, 'value="Kit Cobertura Wi-Fi - 15 Access Points"')

    def test_painel_kit_declarado_resolve_numero_pela_quantidade_de_access_points(self):
        """RN-010 ampliada (FEAT-016): quando Escola.kit_inicial é só o
        número do KIT (formato usado por parte das escolas), o painel
        mostra a Descrição curta do catálogo (RN-011, mesma nomenclatura
        do Lado IXC) — não o número bruto, nem a Descrição completa (que
        tem o qualificador entre parênteses e não cabe no campo)."""
        escola = Escola.objects.create(
            inep="20000002", nome="Escola Teste RI 2", municipio="Fortaleza",
            estado="CE", kit_inicial="4", lote=9,
        )
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
        )
        Ri.objects.create(escola=escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": escola.inep}))
        self.assertContains(resp, 'value="Kit Cobertura Wi-Fi - 4 Access Points"')

    def test_painel_kit_declarado_sem_correspondencia_mostra_o_numero_bruto(self):
        """Sem entrada no catálogo para o número informado, mostra o dado
        bruto mesmo — nenhum valor inventado (CLAUDE.md §9)."""
        escola = Escola.objects.create(
            inep="20000003", nome="Escola Teste RI 3", municipio="Fortaleza",
            estado="CE", kit_inicial="4", lote=9,
        )
        Ri.objects.create(escola=escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": escola.inep}))
        self.assertContains(resp, 'value="4"')
        self.assertNotContains(resp, "Lançar item do Kit declarado")

    def test_lado_eace_nao_tem_rota_de_edicao_ou_exclusao(self):
        with self.assertRaises(NoReverseMatch):
            reverse("ri_item_eace_update", kwargs={"item_pk": 1})
        with self.assertRaises(NoReverseMatch):
            reverse("ri_item_eace_delete", kwargs={"item_pk": 1})

    def test_lancar_kit_instalado_relatorio_eace(self):
        """RN-018 (2026-08-26): "KIT Instalado" do Lado Relatório EACE usa
        o mesmo catálogo `KitPadrao` do Lado IXC (RN-011), quantidade
        sempre 1. Diferente do Lado IXC, o Valor Unitário não nasce 0 —
        vem do preço já cadastrado no catálogo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        kit = KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", unidade="Escola",
            valor_equipamento=Decimal("300.00"), valor_servico=Decimal("50.00"),
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace-kit": kit.pk,
                **self.FORMSET_PRODUTO_EACE_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        item = RiItemRelatorioEace.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Kit Wi-Fi Indoor")
        self.assertEqual(item.quantidade, 1)
        self.assertEqual(item.valor_unitario, Decimal("300.00"))
        self.assertTrue(item.eh_kit)

    def test_lancar_kit_outro_relatorio_eace_sem_correspondencia_fica_com_valor_zero(self):
        """RN-018: opção "Outro" (kit fora do catálogo) sem nenhuma
        entrada correspondente por número de Access Points — nenhum valor
        inventado (CLAUDE.md §9), o item nasce com Valor Unitário 0,00."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace-kit": "outro",
                "eace-kit_outro_numero": "20",
                **self.FORMSET_PRODUTO_EACE_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        item = RiItemRelatorioEace.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Kit Cobertura Wi-Fi - 20 Access Points")
        self.assertEqual(item.valor_unitario, Decimal("0"))

    def test_lancar_multiplos_produtos_relatorio_eace(self):
        """RN-018: "+" lança 0 ou mais produtos numa única submissão, cada
        um com o Valor Unitário resolvido do catálogo (diferente do Lado
        IXC, que nasce 0)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        roteador = KitPadrao.objects.create(
            descricao="Roteador extra", unidade="Unidade", valor_equipamento=Decimal("199.90")
        )
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento=Decimal("5.50")
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace_produto-TOTAL_FORMS": "2",
                "eace_produto-INITIAL_FORMS": "0",
                "eace_produto-MIN_NUM_FORMS": "0",
                "eace_produto-MAX_NUM_FORMS": "1000",
                "eace_produto-0-produto": roteador.pk,
                "eace_produto-0-quantidade": 2,
                "eace_produto-1-produto": cabo.pk,
                "eace_produto-1-quantidade": 10,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=ri).count(), 2)
        item_roteador = RiItemRelatorioEace.objects.get(ri=ri, descricao_item="Roteador extra")
        self.assertEqual(item_roteador.valor_unitario, Decimal("199.90"))
        self.assertFalse(item_roteador.eh_kit)

    def test_segundo_kit_bloqueado_relatorio_eace(self):
        """RN-018: mesmo limite de 1 KIT por INEP do Lado IXC (RN-015),
        agora também no Lado Relatório EACE."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        segundo = KitPadrao.objects.create(descricao="Kit Wi-Fi Outdoor", unidade="Escola")
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace-kit": segundo.pk,
                **self.FORMSET_PRODUTO_EACE_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=ri, eh_kit=True).count(), 1)
        resp_seguinte = self.client.get(resp.url)
        self.assertContains(resp_seguinte, "só é permitido um KIT por INEP")

    def test_lancar_kit_relatorio_eace_gera_entrada_no_historico(self):
        """RN-008: cadastro do KIT do Relatório EACE gera entrada própria
        na linha do tempo, já com o valor resolvido do catálogo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        kit = KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", unidade="Escola", valor_equipamento=Decimal("350.00")
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace-kit": kit.pk,
                **self.FORMSET_PRODUTO_EACE_VAZIO,
            },
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada.campo, "KIT Instalado (Relatório EACE)")
        self.assertEqual(entrada.valor_anterior, "")
        self.assertEqual(entrada.valor_novo, "Kit Wi-Fi Indoor — 1 un. — R$ 350.00")
        self.assertEqual(entrada.autor, self.analista)

    def test_administrador_edita_kit_relatorio_eace(self):
        """RN-018: exceção pontual à imutabilidade da RN-003 — só o item
        marcado como KIT aceita edição (Administrador e Analista, RN-004)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_item_relatorio_eace_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Kit Wi-Fi Indoor (corrigido)", "quantidade": 1, "valor_unitario": "360.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Kit Wi-Fi Indoor (corrigido)")
        self.assertEqual(item.valor_unitario, Decimal("360.00"))

    def test_administrador_exclui_kit_relatorio_eace(self):
        """RN-018/RN-004: exclusão do item KIT do Relatório EACE — só
        Administrador, mesma regra do Lado IXC."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("ri_item_relatorio_eace_delete", kwargs={"item_pk": item.pk}))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(RiItemRelatorioEace.objects.filter(pk=item.pk).exists())

    def test_analista_nao_pode_excluir_kit_relatorio_eace(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.post(reverse("ri_item_relatorio_eace_delete", kwargs={"item_pk": item.pk}))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(RiItemRelatorioEace.objects.filter(pk=item.pk).exists())

    def test_produto_relatorio_eace_pode_ser_editado(self):
        """RN-018 (ampliada, 2026-08-27): a exceção de editar/excluir
        passou a valer para qualquer item deste lado, não só o KIT —
        usuário pediu depois de sincronizar um Produto (ex.: "Nobreak",
        FEAT-024) e não conseguir corrigi-lo. Administrador e Analista
        (RN-004), mesma permissão do Lado IXC."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Cabo de rede", quantidade=10, valor_unitario="5.50",
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_item_relatorio_eace_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Cabo alterado", "quantidade": 5, "valor_unitario": "6.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Cabo alterado")
        self.assertEqual(item.quantidade, 5)

    def test_produto_relatorio_eace_exclusao_restrita_a_administrador(self):
        """RN-018 (ampliada)/RN-004: exclusão de Produto segue restrita a
        Administrador, mesma regra já valia para o KIT."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Cabo de rede", quantidade=10, valor_unitario="5.50",
        )
        self.client.force_login(self.analista)
        resp_analista = self.client.post(
            reverse("ri_item_relatorio_eace_delete", kwargs={"item_pk": item.pk})
        )
        self.assertEqual(resp_analista.status_code, 403)
        self.assertTrue(RiItemRelatorioEace.objects.filter(pk=item.pk).exists())

        self.client.force_login(self.admin)
        resp_admin = self.client.post(
            reverse("ri_item_relatorio_eace_delete", kwargs={"item_pk": item.pk})
        )
        self.assertEqual(resp_admin.status_code, 302)
        self.assertFalse(RiItemRelatorioEace.objects.filter(pk=item.pk).exists())

    def test_lancar_kit_instalado_ixc(self):
        """RN-011 (2ª correção, 2026-08-24): "KIT Instalado" — descrição
        vem do catálogo `KitPadrao` (aba LPU), quantidade sempre 1. Valor
        unitário (ajuste do usuário, 2026-08-24): tirado do formulário —
        nasce 0, sem informação nenhuma pedida."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": kit.pk,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        item = RiItemIxc.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Kit Wi-Fi Indoor")
        self.assertEqual(item.quantidade, 1)
        self.assertEqual(item.valor_unitario, Decimal("0"))

    def test_lancar_kit_do_catalogo_grava_descricao_curta_sem_parenteses(self):
        """Correção (2026-08-25): a descrição gravada é a curta (RN-011,
        mesma do select), não a completa da planilha — que traz um
        qualificador entre parênteses (ex.: "(serviços, materiais e
        equipamentos)") sem utilidade na tela."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            unidade="Escola",
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": kit.pk,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        item = RiItemIxc.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Kit Cobertura Wi-Fi - 4 Access Points")

    def test_segundo_kit_bloqueado_quando_ja_tem_um_lancado(self):
        """RN-011 (ajuste 2026-08-26): 1 KIT por INEP — RI que já tem um
        KIT lançado não pode lançar outro; o já lançado não muda."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        primeiro = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        segundo = KitPadrao.objects.create(descricao="Kit Wi-Fi Outdoor", unidade="Escola")
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="0", eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": segundo.pk,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri, eh_kit=True).count(), 1)
        self.assertEqual(
            RiItemIxc.objects.get(ri=ri, eh_kit=True).descricao_item, "Kit Wi-Fi Indoor"
        )
        resp_seguinte = self.client.get(resp.url)
        self.assertContains(resp_seguinte, "só é permitido um KIT por INEP")

    def test_segundo_kit_via_outro_tambem_bloqueado(self):
        """RN-011: o bloqueio de 1 KIT por INEP vale também para a opção
        "Outro" (kit fora do catálogo), não só para o catálogo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": "outro",
                "kit_outro_numero": "8",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        self.assertEqual(RiItemIxc.objects.filter(ri=ri, eh_kit=True).count(), 1)

    def test_produto_lancado_normalmente_mesmo_com_kit_ja_lancado(self):
        """RN-011: o bloqueio é só para KIT — lançar Produto continua
        funcionando normalmente num RI que já tem KIT."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": cabo.pk,
                "form-0-quantidade": 5,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(RiItemIxc.objects.filter(ri=ri, descricao_item="Cabo de rede").exists())

    def test_painel_esconde_seletor_de_kit_quando_ja_lancado(self):
        """RN-011: com KIT já lançado, a tela não oferece mais o seletor —
        só edita/exclui o item já lançado (RN-004)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "já tem um KIT lançado")
        self.assertNotContains(resp, 'id="id_kit"')

    def test_lancar_kit_outro_gera_descricao_no_padrao_do_catalogo(self):
        """RN-011 (opção "Outro", 2026-08-24): kit instalado diferente de
        tudo no catálogo — pessoa escolhe "Outro" e digita o número de
        Access Points; descrição gravada segue o padrão do catálogo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": "outro",
                "kit_outro_numero": "20",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        self.assertEqual(resp.status_code, 302)
        item = RiItemIxc.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Kit Cobertura Wi-Fi - 20 Access Points")
        self.assertEqual(item.quantidade, 1)
        self.assertEqual(item.valor_unitario, Decimal("0"))

    def test_lancar_kit_outro_sem_numero_nao_cria_item(self):
        """RN-011: "Outro" sem o número de Access Points não lança nada e
        mostra erro, em vez de gravar uma descrição incompleta."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": "outro", **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 0)
        self.assertContains(resp, "Verifique o KIT selecionado")

    def test_select_kit_instalado_inclui_opcao_outro(self):
        """RN-011: a opção "Outro" está sempre disponível, mesmo com o
        catálogo cheio para o Lote desta escola."""
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        valores_kit = [choice[0] for choice in resp.context["kit_form"].fields["kit"].choices]
        self.assertIn("outro", valores_kit)

    def test_kit_instalado_so_lista_catalogo_com_unidade_escola(self):
        """RN-011 (3ª correção, 2026-08-24): o select de "KIT Instalado"
        só mostra o catálogo cuja Unidade é "Escola" — nunca item avulso."""
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_kit = [choice[1] for choice in resp.context["kit_form"].fields["kit"].choices]
        self.assertTrue(any("Kit Wi-Fi Indoor" in opcao for opcao in opcoes_kit))
        self.assertFalse(any("Cabo de rede" in opcao for opcao in opcoes_kit))

    def test_select_do_kit_mostra_descricao_curta_nao_a_completa(self):
        """RN-011 (2026-08-24): a Descrição completa da LPU é grande
        demais para o select — mostra a Descrição curta."""
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 8 Access Points (serviços, materiais e equipamentos)",
            unidade="Escola",
        )
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_kit = [choice[1] for choice in resp.context["kit_form"].fields["kit"].choices]
        self.assertIn("Kit Cobertura Wi-Fi - 8 Access Points", opcoes_kit)
        self.assertNotIn(
            "Kit Cobertura Wi-Fi - 8 Access Points (serviços, materiais e equipamentos)",
            opcoes_kit,
        )

    def test_produto_exclui_catalogo_com_unidade_escola(self):
        """RN-011: o select de "Produto" nunca mostra o KIT (Unidade
        "Escola") do catálogo — só itens avulsos."""
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_produto = [
            choice[1] for choice in resp.context["produto_formset"].empty_form.fields["produto"].choices
        ]
        self.assertTrue(any("Cabo de rede" in opcao for opcao in opcoes_produto))
        self.assertFalse(any("Kit Wi-Fi Indoor" in opcao for opcao in opcoes_produto))

    def test_produto_sem_valor_de_equipamento_sai_da_lista_do_lado_ixc(self):
        """RN-055: produto sem "Equipamentos (R$)" na LPU
        (`valor_equipamento` nulo) — só "Serviços (R$)" preenchido — não
        aparece no select de "Produto" do Lado IXC; ele nunca é faturado
        aqui (RN-013 usa só `valor_faturavel` = valor_equipamento) e ia
        sempre gerar R$ 0,00 sem explicação."""
        KitPadrao.objects.create(
            descricao="Injetor PoE", unidade="Unidade",
            valor_equipamento=None, valor_servico="564.04",
        )
        KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro",
            valor_equipamento="10.00", valor_servico="2.00",
        )
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_produto = [
            choice[1] for choice in resp.context["produto_formset"].empty_form.fields["produto"].choices
        ]
        self.assertTrue(any("Cabo de rede" in opcao for opcao in opcoes_produto))
        self.assertFalse(any("Injetor PoE" in opcao for opcao in opcoes_produto))

    def test_produto_sem_valor_de_equipamento_continua_no_lado_relatorio_eace(self):
        """RN-055: o filtro é só do Lado IXC — o Lado Relatório EACE (3º
        lado, RN-018) usa `valor_total` (Equipamento + Serviço), então um
        produto só-com-Serviço continua faturável ali e continua na lista."""
        KitPadrao.objects.create(
            descricao="Injetor PoE", unidade="Unidade",
            valor_equipamento=None, valor_servico="564.04",
        )
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_produto_eace = [
            choice[1]
            for choice in resp.context["produto_formset_eace"].empty_form.fields["produto"].choices
        ]
        self.assertTrue(any("Injetor PoE" in opcao for opcao in opcoes_produto_eace))

    def test_catalogo_filtra_por_lote_da_escola(self):
        """RN-011: o catálogo mostra só as entradas do Lote desta escola —
        mesmo kit/produto pode ter preço (e existência) diferente por
        Lote (RN-010)."""
        self.escola.lote = 9
        self.escola.save()
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", lote=9, unidade="Escola")
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", lote=11, unidade="Escola")
        KitPadrao.objects.create(descricao="Só do Lote 11", lote=11, unidade="Unidade")
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        opcoes_kit = [choice[1] for choice in resp.context["kit_form"].fields["kit"].choices]
        self.assertTrue(any("Kit Wi-Fi Indoor" in opcao for opcao in opcoes_kit))
        self.assertFalse(any("Só do Lote 11" in opcao for opcao in opcoes_kit))

    def test_lancar_multiplos_produtos_ixc(self):
        """RN-011: "+" lança 0 ou mais produtos numa única submissão, cada
        um escolhido no catálogo `KitPadrao`, com Quantidade digitada
        manualmente. Valor unitário (ajuste do usuário, 2026-08-24):
        tirado do formulário — nasce 0."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        roteador = KitPadrao.objects.create(
            descricao="Roteador extra", unidade="Unidade", valor_equipamento="50.00"
        )
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "2",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": roteador.pk,
                "form-0-quantidade": 1,
                "form-1-produto": cabo.pk,
                "form-1-quantidade": 10,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 2)
        item_roteador = RiItemIxc.objects.get(ri=ri, descricao_item="Roteador extra")
        self.assertEqual(item_roteador.valor_unitario, Decimal("0"))
        self.assertTrue(RiItemIxc.objects.filter(ri=ri, descricao_item="Cabo de rede").exists())

    def test_lancar_kit_e_produto_ixc_gera_entrada_no_historico(self):
        """RN-008 (esclarecida em 2026-08-26): cadastro de KIT Instalado e
        de Produto no Lado IXC gera entrada própria na linha do tempo, cada
        um com o próprio rótulo — não só a mudança de status."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": kit.pk,
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": cabo.pk,
                "form-0-quantidade": 5,
            },
        )
        entrada_kit = RiHistorico.objects.get(ri=ri, campo="KIT Instalado (Lado IXC)")
        self.assertEqual(entrada_kit.tipo, RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada_kit.valor_anterior, "")
        self.assertEqual(entrada_kit.valor_novo, "Kit Wi-Fi Indoor — 1 un.")
        self.assertEqual(entrada_kit.autor, self.analista)
        entrada_produto = RiHistorico.objects.get(ri=ri, campo="Produto (Lado IXC)")
        self.assertEqual(entrada_produto.valor_anterior, "")
        self.assertEqual(entrada_produto.valor_novo, "Cabo de rede — 5 un.")

    def test_entrada_de_cadastro_ixc_aparece_como_cadastrou_na_tela(self):
        """Sem valor anterior (cadastro novo, não alteração), a linha do
        tempo usa "Cadastrou", não "Alterou" — feedback claro pedido pelo
        usuário junto com o critério de aceite (RN-008, 2026-08-26)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Cadastrou")
        self.assertContains(resp, "KIT Instalado (Lado IXC)")
        self.assertContains(resp, "Kit Wi-Fi Indoor — 1 un.")

    def test_lancar_produto_do_catalogo_grava_descricao_curta_sem_parenteses(self):
        """Correção (2026-08-25): mesma regra do KIT Instalado vale para
        "Produtos" — grava a descrição curta, sem o qualificador entre
        parênteses do catálogo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        nobreak = KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", unidade="Unidade",
            valor_equipamento="1551.93",
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": nobreak.pk,
                "form-0-quantidade": 1,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        item = RiItemIxc.objects.get(ri=ri)
        self.assertEqual(item.descricao_item, "Nobreak")

    def test_remover_linha_de_produto_no_meio_nao_quebra_submissao(self):
        """RN-011 (botão de remover linha, 2026-08-24): o JS só apaga a
        div da linha no navegador, sem reindexar as demais — o índice do
        meio some da submissão (como se nunca tivesse sido preenchido) e
        os outros dois continuam lançando normalmente."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        roteador = KitPadrao.objects.create(
            descricao="Roteador extra", unidade="Unidade", valor_equipamento="50.00"
        )
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "3",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": roteador.pk,
                "form-0-quantidade": 1,
                # form-1 removida pelo "x" antes de enviar — sem essas chaves.
                "form-2-produto": cabo.pk,
                "form-2-quantidade": 10,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 2)
        self.assertTrue(RiItemIxc.objects.filter(ri=ri, descricao_item="Roteador extra").exists())
        self.assertTrue(RiItemIxc.objects.filter(ri=ri, descricao_item="Cabo de rede").exists())

    def test_lista_de_produtos_nasce_sem_nenhuma_linha_visivel(self):
        """RN-011: a lista de "Produtos" não vem com nenhuma linha aberta —
        só aparece ao clicar no "+"."""
        KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(len(resp.context["produto_formset"].forms), 0)

    def test_lancar_produto_sem_selecionar_nenhum_nao_cria_item(self):
        """RN-011: submissão do "+" sem nenhuma linha preenchida e sem
        mudar a Data de Ativação (nem Município/Estado, reenviados sem
        mudança) não cria item nem quebra a página.

        Post/Redirect/Get (bug reportado 2026-08-26): a resposta desse
        erro é um redirect, não a página renderizada direto no POST — um
        F5 do usuário na página resultante não reenvia o formulário nem
        repete a mensagem de erro à toa."""
        ri = Ri.objects.create(
            escola=self.escola,
            status=Ri.ANDAMENTO,
            municipio_ixc="Fortaleza",
            estado_ixc="CE",
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": "",
                "form-0-quantidade": "",
                "data_ativacao": "",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 0)

        resp_seguinte = self.client.get(resp.url)
        self.assertContains(resp_seguinte, "Selecione um KIT, um produto ou informe a Data de Ativação.")
        # Formset vazio é válido (nenhuma linha obrigatória, extra=0) — não
        # deve aparecer o aviso de campo inválido.
        self.assertNotContains(resp_seguinte, "Verifique os produtos preenchidos.")

        # A mensagem é consumida ao ser exibida — um F5 (novo GET na mesma
        # URL) não deve trazê-la de volta.
        resp_f5 = self.client.get(resp.url)
        self.assertNotContains(resp_f5, "Selecione um KIT, um produto ou informe a Data de Ativação.")

    def test_reenviar_formulario_ja_preenchido_sem_mudar_nada_mostra_mensagem_neutra(self):
        """RN-011 (correção 2026-09-02, bug reportado pelo usuário — INEP
        35275505): quando o KIT já está lançado e a Data de Ativação já
        está preenchida, reenviar o mesmo formulário sem nenhuma alteração
        nova (ex.: clique duplo em "Salvar") não pode repetir "Selecione
        um KIT, um produto ou informe a Data de Ativação" — a mensagem é
        enganosa, já que os dois já estão lá, visíveis na própria tela."""
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        ri = Ri.objects.create(
            escola=self.escola,
            status=Ri.ANDAMENTO,
            data_ativacao=date(2026, 8, 26),
            municipio_ixc="Fortaleza",
            estado_ixc="CE",
            cnpj="00.000.000/0001-00",
            cnpj_ficticio="11.111.111/0001-11",
        )
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        self.client.force_login(self.analista)
        # Mesma submissão que o navegador manda ao reenviar a tela já
        # preenchida sem mexer em nada: sem campo "kit" (o <select> nem
        # aparece mais — RN-011, kit_ja_lancado), sem linha nova de
        # produto, e os demais campos com os MESMOS valores já salvos.
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "2026-08-26",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                "cnpj": "00.000.000/0001-00",
                "cnpj_ficticio": "11.111.111/0001-11",
            },
            follow=True,
        )
        self.assertContains(resp, "Nenhuma alteração para salvar.")
        self.assertNotContains(resp, "Selecione um KIT, um produto ou informe a Data de Ativação.")

    def test_salvar_data_ativacao_sozinha_sem_produto(self):
        """RN-011 (2026-08-24): "Data Ativação" — um valor só do RI,
        salvo mesmo sem lançar nenhum produto na mesma submissão."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": "",
                "form-0-quantidade": "",
                "data_ativacao": "2026-08-24",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        ri.refresh_from_db()
        self.assertEqual(str(ri.data_ativacao), "2026-08-24")
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 0)

    def test_lancar_produto_tambem_salva_data_ativacao_na_mesma_submissao(self):
        """RN-011: Data Ativação e lançamento de produto podem ser salvos
        juntos, na mesma submissão."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="metro", valor_equipamento="10.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": cabo.pk,
                "form-0-quantidade": 5,
                "data_ativacao": "2026-08-24",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        ri.refresh_from_db()
        self.assertEqual(str(ri.data_ativacao), "2026-08-24")
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 1)

    def test_salvar_municipio_estado_ixc_normaliza_uf_maiuscula(self):
        """RN-014 (2026-08-26): Município/Estado do Lado IXC, salvos na
        mesma submissão do formulário único; UF sempre gravada em
        maiúsculas, mesmo padrão de Escola.estado."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "ce",
            },
        )
        self.assertEqual(resp.status_code, 302)
        ri.refresh_from_db()
        self.assertEqual(ri.municipio_ixc, "Fortaleza")
        self.assertEqual(ri.estado_ixc, "CE")

    def test_salvar_data_ativacao_municipio_estado_ixc_gera_entrada_no_historico(self):
        """RN-008 (esclarecida em 2026-08-26): cadastro de Data de
        Ativação/Município/Estado (Lado IXC) gera uma entrada por campo
        alterado, com o valor anterior vazio (primeiro cadastro)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "2026-08-24",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        entrada_data = RiHistorico.objects.get(ri=ri, campo="Data de Ativação")
        self.assertEqual(entrada_data.valor_anterior, "")
        self.assertEqual(entrada_data.valor_novo, "24/08/2026")
        entrada_municipio = RiHistorico.objects.get(ri=ri, campo="Município (Lado IXC)")
        self.assertEqual(entrada_municipio.valor_novo, "Fortaleza")
        entrada_estado = RiHistorico.objects.get(ri=ri, campo="Estado (Lado IXC)")
        self.assertEqual(entrada_estado.valor_novo, "CE")

    def test_salvar_cnpj_e_cnpj_ficticio_gera_entrada_no_historico(self):
        """RN-048 (2026-09-01): mesmo padrão de Município/Estado — cadastro
        de CNPJ/CNPJ Fictício (Lado IXC) gera uma entrada por campo
        alterado, com o valor anterior vazio (primeiro cadastro)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "",
                "municipio_ixc": "",
                "estado_ixc": "",
                "cnpj": "00.000.000/0001-00",
                "cnpj_ficticio": "11.111.111/0001-11",
            },
        )
        entrada_cnpj = RiHistorico.objects.get(ri=ri, campo="CNPJ (Lado IXC)")
        self.assertEqual(entrada_cnpj.valor_anterior, "")
        self.assertEqual(entrada_cnpj.valor_novo, "00.000.000/0001-00")
        entrada_ficticio = RiHistorico.objects.get(ri=ri, campo="CNPJ Fictício (Lado IXC)")
        self.assertEqual(entrada_ficticio.valor_novo, "11.111.111/0001-11")

    def test_editar_data_ativacao_gera_entrada_so_do_campo_alterado(self):
        """Só o campo de fato alterado nessa submissão gera entrada — não
        os 3 juntos a cada "Salvar" do Lado IXC."""
        ri = Ri.objects.create(
            escola=self.escola, status=Ri.ANDAMENTO,
            data_ativacao=date(2026, 8, 20), municipio_ixc="Fortaleza", estado_ixc="CE",
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "2026-08-25",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
            },
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada.campo, "Data de Ativação")
        self.assertEqual(entrada.valor_anterior, "20/08/2026")
        self.assertEqual(entrada.valor_novo, "25/08/2026")

    def test_municipio_e_estado_ixc_nao_bloqueiam_o_salvar(self):
        """RN-014 (revista em 2026-08-26): Município/Estado NÃO travam o
        "Salvar" do Lado IXC — a exigência é só na hora de enviar o
        e-mail/baixar a planilha (RN-013), não a cada lançamento. Um KIT
        lançado sem os dois salva normalmente."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 1)

    def test_cnpj_e_cnpj_ficticio_nao_bloqueiam_o_salvar(self):
        """RN-048 (2026-09-01): mesmo padrão de Município/Estado (RN-014) —
        CNPJ/CNPJ Fictício NÃO travam o "Salvar" do Lado IXC; a exigência é
        só na hora de enviar o e-mail/baixar a planilha (RN-013). Um KIT
        lançado sem os dois salva normalmente."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 1)

    def test_estado_ixc_precisa_ter_2_letras(self):
        """RN-014: UF fora do padrão (2 letras) não é salva — não bloqueia
        o restante do Lado IXC, só essa submissão específica."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "",
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "C",
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Informe a UF com 2 letras")
        ri.refresh_from_db()
        self.assertEqual(ri.estado_ixc, "")

    def test_divergencia_municipio_estado_ixc_mostra_alerta_sem_bloquear(self):
        """RN-014: Município/Estado do Lado IXC diferentes do cadastro da
        Escola (Fortaleza/CE, ver setUp) mostram alerta visual, mas a
        submissão salva normalmente (não bloqueia)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "",
                "municipio_ixc": "Recife",
                "estado_ixc": "PE",
            },
        )
        self.assertEqual(resp.status_code, 302)
        ri.refresh_from_db()
        self.assertEqual(ri.municipio_ixc, "Recife")

        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "diferem do cadastro da Escola")

    def test_municipio_estado_ixc_iguais_ao_cadastro_nao_mostra_alerta(self):
        """RN-014: mesmo valor dos dois lados (ou campo vazio) não gera
        alerta de divergência."""
        ri = Ri.objects.create(
            escola=self.escola,
            status=Ri.IMPLANTACAO_EACE,
            municipio_ixc="Fortaleza",
            estado_ixc="CE",
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, "diferem do cadastro da Escola")

    def test_municipio_estado_ixc_vem_preenchido_com_dado_da_escola_quando_ainda_vazio(self):
        """RN-014 (revista em 2026-09-02): Município/Estado do Lado IXC
        nascem preenchidos com o cadastro da Escola (INEP) quando o RI ainda
        não tem valor próprio salvo — evita digitar de novo algo que o
        sistema já tem. Continua editável: é só o valor inicial que muda."""
        Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(
            resp.context["data_ativacao_form"].initial.get("municipio_ixc"), "Fortaleza"
        )
        self.assertEqual(resp.context["data_ativacao_form"].initial.get("estado_ixc"), "CE")

    def test_municipio_estado_ixc_ja_salvo_nao_e_sobrescrito_pelo_dado_da_escola(self):
        """RN-014: uma vez que o Lado IXC tem valor próprio salvo (mesmo
        diferente do cadastro da Escola), o pré-preenchimento não entra em
        ação — só vale quando o campo do Lado IXC ainda está vazio."""
        Ri.objects.create(
            escola=self.escola, status=Ri.ANDAMENTO,
            municipio_ixc="Recife", estado_ixc="PE",
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(resp.context["data_ativacao_form"].initial.get("municipio_ixc"), "Recife")
        self.assertEqual(resp.context["data_ativacao_form"].initial.get("estado_ixc"), "PE")

    def test_mes_operacao_ixc_vem_preenchido_com_mes_corrente_quando_ainda_vazio(self):
        """RN-053: select "Mês da Operação" nasce selecionado no mês
        corrente quando o RI ainda não tem valor próprio salvo — mesmo
        padrão de pré-preenchimento de Município/Estado (RN-014)."""
        Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(
            resp.context["data_ativacao_form"].initial.get("mes_operacao_ixc"),
            timezone.now().month,
        )

    def test_mes_operacao_ixc_ja_salvo_nao_e_sobrescrito_pelo_mes_corrente(self):
        """RN-053: uma vez que o Lado IXC tem valor próprio salvo (RI de
        operação retroativa/futura), o pré-preenchimento com o mês
        corrente não entra em ação."""
        mes_diferente = 1 if timezone.now().month != 1 else 2
        Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO, mes_operacao_ixc=mes_diferente)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(
            resp.context["data_ativacao_form"].initial.get("mes_operacao_ixc"), mes_diferente
        )

    def test_salvar_mes_operacao_ixc_gera_entrada_no_historico_com_nome_do_mes(self):
        """RN-008/RN-053: cadastro do Mês da Operação gera entrada no
        histórico com o NOME do mês (ex.: "Agosto"), não o número salvo."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                **self.FORMSET_PRODUTO_VAZIO,
                "data_ativacao": "",
                "municipio_ixc": "",
                "estado_ixc": "",
                "mes_operacao_ixc": "8",
            },
        )
        entrada = RiHistorico.objects.get(ri=ri, campo="Mês da Operação (Lado IXC)")
        self.assertEqual(entrada.valor_anterior, "")
        self.assertEqual(entrada.valor_novo, "Agosto")

    def test_divergencia_kit_declarado_e_instalado_mostra_alerta_sem_bloquear(self):
        """RN-002 (esclarecida em 2026-08-26): KIT instalado diferente do
        Kit declarado (Escola.kit_inicial = "Kit Wi-Fi Indoor", ver setUp)
        mostra alerta visual — mesma mecânica da RN-014 (município) —, mas
        o lançamento salva normalmente (não bloqueia)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 15 Access Points", unidade="Escola",
        )
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 1)

        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "difere do Kit declarado")

    def test_kit_declarado_e_instalado_iguais_nao_mostra_alerta(self):
        """RN-002: mesma descrição nos dois lados (mesmo padrão da RN-014
        para município) não gera alerta de divergência."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        RiItemIxc.objects.create(
            ri=ri, descricao_item=kit.descricao_curta, quantidade=1,
            valor_unitario=Decimal("0"), eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, "difere do Kit declarado")

    def test_catalogo_vazio_mostra_aviso_no_painel_ixc(self):
        """RN-011: sem nenhuma entrada de KIT (Unidade "Escola") no
        catálogo para o Lote desta escola, o painel do IXC avisa em vez de
        oferecer um select vazio."""
        self.escola.lote = 99
        self.escola.save()
        KitPadrao.objects.create(descricao="Kit de outro lote", lote=1, unidade="Escola")
        Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, 'Nenhum KIT (Unidade "Escola") cadastrado no catálogo')

    def test_analista_edita_item_ixc(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Roteador Wi-Fi 6", "quantidade": 2, "valor_unitario": "150.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Roteador Wi-Fi 6")
        self.assertEqual(item.quantidade, 2)

    def test_editar_item_ixc_gera_entrada_no_historico(self):
        """RN-008 (esclarecida em 2026-08-26): edição de item do Lado IXC
        gera entrada com valor anterior e novo — não só o cadastro."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Roteador Wi-Fi 6", "quantidade": 2, "valor_unitario": "150.00"},
        )
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada.campo, "Item do Lado IXC (editado)")
        self.assertEqual(entrada.valor_anterior, "Roteador — 1 un. — R$ 100.00")
        self.assertEqual(entrada.valor_novo, "Roteador Wi-Fi 6 — 2 un. — R$ 150.00")
        self.assertEqual(entrada.autor, self.analista)

    def test_administrador_exclui_item_ixc(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item.pk}))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(RiItemIxc.objects.filter(pk=item.pk).exists())

    def test_excluir_item_ixc_gera_entrada_no_historico(self):
        """RN-008 (esclarecida em 2026-08-26): exclusão de item do Lado
        IXC também gera entrada, com o valor removido."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.admin)
        self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item.pk}))
        entrada = RiHistorico.objects.get(ri=ri, tipo=RiHistorico.LOG_CAMPO)
        self.assertEqual(entrada.campo, "Item do Lado IXC (excluído)")
        self.assertEqual(entrada.valor_anterior, "Roteador — 1 un. — R$ 100.00")
        self.assertEqual(entrada.valor_novo, "Excluído")
        self.assertEqual(entrada.autor, self.admin)

    def test_formulario_de_edicao_ixc_usa_ponto_decimal(self):
        """value de <input type=number> precisa de ponto - LANGUAGE_CODE
        pt-br localiza `{{ valor }}` com vírgula, o que o navegador rejeita
        num campo numérico (o campo fica em branco)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="350.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, 'value="350.00"')
        self.assertNotContains(resp, 'value="350,00"')

    def test_analista_nao_pode_excluir_item_ixc(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.analista)
        resp = self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item.pk}))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(RiItemIxc.objects.filter(pk=item.pk).exists())

    # RN-003 (2026-08-26): confronto formal Lado IXC × Lado Relatório EACE
    # — Descrição (qual KIT/Produto) + Quantidade, sem Valor Unitário
    # (Lado IXC nasce sempre 0,00, RN-011). Divergência bloqueia o envio ao
    # financeiro (RN-001) e destaca em vermelho o item do Lado IXC.

    def test_kit_diferente_entre_ixc_e_relatorio_eace_bloqueia_envio_financeiro(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Outdoor", quantidade=1, valor_unitario="350.00", eh_kit=True,
        )
        # A divergência é recalculada nas ações de lançamento/edição/
        # exclusão — dispara aqui editando o item do Lado IXC (sem mudar
        # nada de fato) só para acionar o recálculo depois do setUp acima
        # ter criado os itens direto no banco (sem passar pela view).
        item_ixc = ri.itens_ixc.get()
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item_ixc.pk}),
            {"descricao_item": item_ixc.descricao_item, "quantidade": 1, "valor_unitario": "0"},
        )
        self.assertTrue(
            RiDivergencia.objects.filter(
                ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, bloqueia=True, resolvida_em__isnull=True
            ).exists()
        )
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ANDAMENTO)

    def test_kit_e_produtos_iguais_nao_gera_divergencia(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, "municipio_ixc": "Fortaleza", "estado_ixc": "CE", **self.FORMSET_PRODUTO_VAZIO},
        )
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_relatorio_eace", "eace-kit": kit.pk, **self.FORMSET_PRODUTO_EACE_VAZIO},
        )
        self.assertFalse(
            RiDivergencia.objects.filter(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, resolvida_em__isnull=True).exists()
        )
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
        )
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_produto_com_quantidade_diferente_gera_divergencia(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        cabo = KitPadrao.objects.create(descricao="Cabo de rede", unidade="metro", valor_equipamento="5.50")
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "municipio_ixc": "Fortaleza", "estado_ixc": "CE",
                "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "0", "form-MIN_NUM_FORMS": "0", "form-MAX_NUM_FORMS": "1000",
                "form-0-produto": cabo.pk, "form-0-quantidade": 10,
            },
        )
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace_produto-TOTAL_FORMS": "1", "eace_produto-INITIAL_FORMS": "0",
                "eace_produto-MIN_NUM_FORMS": "0", "eace_produto-MAX_NUM_FORMS": "1000",
                "eace_produto-0-produto": cabo.pk, "eace_produto-0-quantidade": 5,
            },
        )
        divergencia = RiDivergencia.objects.get(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, resolvida_em__isnull=True)
        self.assertIn("Cabo de rede", divergencia.descricao)
        self.assertIn("10 un.", divergencia.descricao)
        self.assertIn("5 un.", divergencia.descricao)

    def test_corrigir_item_ixc_resolve_divergencia_automaticamente(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item_ixc = RiItemIxc.objects.create(
            ri=ri, descricao_item="Cabo de rede", quantidade=10, valor_unitario="0",
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Cabo de rede", quantidade=5, valor_unitario="5.50",
        )
        self.client.force_login(self.analista)
        # 1ª edição (sem mudar quantidade) só para acionar o 1º recálculo
        # do confronto e abrir a divergência.
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item_ixc.pk}),
            {"descricao_item": "Cabo de rede", "quantidade": 10, "valor_unitario": "0"},
        )
        self.assertTrue(
            RiDivergencia.objects.filter(
                ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, resolvida_em__isnull=True
            ).exists()
        )
        # Corrige a quantidade para igualar ao Lado Relatório EACE.
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item_ixc.pk}),
            {"descricao_item": "Cabo de rede", "quantidade": 5, "valor_unitario": "0"},
        )
        divergencia = RiDivergencia.objects.get(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO)
        self.assertIsNotNone(divergencia.resolvida_em)

    def test_lado_ixc_vazio_nao_gera_divergencia_mesmo_com_relatorio_eace_preenchido(self):
        """RN-003 (ajustada em 2026-09-02, pedido do usuário): Lado IXC
        vazio (nada lançado ainda) não é divergência, mesmo com o Lado
        Relatório EACE já tendo KIT/Produtos — a comparação só faz
        sentido quando os dois lados têm algum valor lançado."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="350.00", eh_kit=True,
        )
        sincronizar_divergencia_kit_relatorio(ri)
        self.assertFalse(
            RiDivergencia.objects.filter(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO).exists()
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertFalse(resp.context["divergencia_kit_relatorio"]["diverge"])

    def test_lado_relatorio_eace_vazio_nao_gera_divergencia_mesmo_com_ixc_preenchido(self):
        """Mesma regra, lado invertido — Lado Relatório EACE vazio (ainda
        não sincronizado) não é divergência com o Lado IXC já lançado."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        sincronizar_divergencia_kit_relatorio(ri)
        self.assertFalse(
            RiDivergencia.objects.filter(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO).exists()
        )

    def test_divergencia_e_resolvida_quando_lado_ixc_fica_vazio_de_novo(self):
        """Divergência real (2 lados preenchidos, KITs diferentes) some
        quando o item do Lado IXC é excluído e o lado fica vazio — mesma
        mecânica de auto-resolução já usada quando os valores passam a
        bater."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        item_ixc = RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Outdoor", quantidade=1, valor_unitario="350.00", eh_kit=True,
        )
        sincronizar_divergencia_kit_relatorio(ri)
        self.assertTrue(
            RiDivergencia.objects.filter(
                ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO, resolvida_em__isnull=True
            ).exists()
        )
        self.client.force_login(self.admin)
        self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item_ixc.pk}))
        divergencia = RiDivergencia.objects.get(ri=ri, tipo=RiDivergencia.TIPO_KIT_RELATORIO)
        self.assertIsNotNone(divergencia.resolvida_em)

    def test_item_ixc_divergente_aparece_destacado_na_tela(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Outdoor", quantidade=1, valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Diverge do Relatório EACE")
        self.assertContains(resp, "bloqueia o envio ao financeiro")

    # RN-020 (2026-08-27): com o RI em "Faturamento Concluído", os campos
    # do Lado IXC e do Lado Relatório EACE ficam bloqueados para os dois
    # perfis — só o Administrador troca o status para liberar de novo
    # (guard de status coberto em RiStatusUpdateViewTests).

    def test_bloqueia_salvar_ixc_com_ri_em_faturamento_concluido(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        self.client.force_login(self.admin)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_ixc",
                "kit": kit.pk,
                "municipio_ixc": "Fortaleza",
                "estado_ixc": "CE",
                **self.FORMSET_PRODUTO_VAZIO,
            },
        )
        self.assertRedirects(resp, reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertFalse(RiItemIxc.objects.filter(ri=ri).exists())

    def test_bloqueia_editar_item_ixc_com_ri_em_faturamento_concluido(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Roteador Wi-Fi 6", "quantidade": 2, "valor_unitario": "150.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Roteador")

    def test_bloqueia_excluir_item_ixc_com_ri_em_faturamento_concluido(self):
        """Bloqueio vale até para o Administrador, que teria permissão de
        excluir (RN-004) fora desse status."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item.pk}))
        self.assertRedirects(resp, reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertTrue(RiItemIxc.objects.filter(pk=item.pk).exists())

    def test_bloqueia_salvar_relatorio_eace_com_ri_em_faturamento_concluido(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        kit = KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", unidade="Escola",
            valor_equipamento=Decimal("300.00"), valor_servico=Decimal("50.00"),
        )
        self.client.force_login(self.admin)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace-kit": kit.pk,
                **self.FORMSET_PRODUTO_EACE_VAZIO,
            },
        )
        self.assertRedirects(resp, reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertFalse(RiItemRelatorioEace.objects.filter(ri=ri).exists())

    def test_bloqueia_editar_kit_relatorio_eace_com_ri_em_faturamento_concluido(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_item_relatorio_eace_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Kit Wi-Fi Indoor (corrigido)", "quantidade": 1, "valor_unitario": "360.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Kit Wi-Fi Indoor")

    def test_bloqueia_excluir_kit_relatorio_eace_com_ri_em_faturamento_concluido(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        item = RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1,
            valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("ri_item_relatorio_eace_delete", kwargs={"item_pk": item.pk}))
        self.assertRedirects(resp, reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertTrue(RiItemRelatorioEace.objects.filter(pk=item.pk).exists())

    def test_tela_esconde_formularios_e_acoes_dos_2_lados_em_faturamento_concluido(self):
        """RN-020 (Lado Relatório EACE) + RN-052 (Lado IXC — "Faturamento
        Concluído" é só um dos casos de "fora de Em Andamento"): os dois
        lados ficam sem ação de editar/excluir item; o Lado Relatório EACE
        esconde o formulário inteiro (RN-020, comportamento inalterado); o
        Lado IXC mantém os campos visíveis (usuário precisa continuar
        vendo Data Ativação/CNPJ/Município/Estado já lançados), só que
        desabilitados (RN-052)."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="0", eh_kit=True,
        )
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="350.00", eh_kit=True,
        )
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))

        # RN-020: Lado Relatório EACE — formulário inteiro some, sem ação
        # de editar/excluir item.
        self.assertContains(resp, "Bloqueado (Faturamento Concluído)")
        self.assertContains(
            resp,
            'RI em "Faturamento Concluído" — lançamento bloqueado até o Administrador trocar o status (RN-020).',
        )
        self.assertNotContains(resp, 'name="acao" value="salvar_relatorio_eace"')
        self.assertNotContains(resp, "ri_item_relatorio_eace_delete")

        # RN-052: Lado IXC — formulário continua renderizado (campos
        # visíveis), mas desabilitado; sem ação de editar/excluir item.
        self.assertFalse(resp.context["lado_ixc_editavel"])
        self.assertTrue(resp.context["kit_form"].fields["kit"].disabled)
        self.assertTrue(resp.context["data_ativacao_form"].fields["data_ativacao"].disabled)
        self.assertContains(resp, 'name="acao" value="salvar_ixc"')
        self.assertContains(resp, 'Somente visualização (fora de "Em Andamento")')
        self.assertContains(
            resp, 'RI fora de "Em Andamento" — campos do Lado IXC em modo visualização (RN-052).'
        )
        self.assertNotContains(resp, "ri_item_ixc_delete")

    def test_lado_ixc_bloqueia_lancamento_fora_de_em_andamento(self):
        """RN-052 (2026-09-02): "Salvar" do Lado IXC só aceita POST com o
        RI em "Em Andamento" — em qualquer outro status (aqui,
        "Implantação EACE", o status inicial) o lançamento é recusado,
        sem criar item."""
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
            follow=True,
        )
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 0)
        # Mensagem exibida via django.contrib.messages, auto-escapada pelo
        # template (`"` vira `&quot;`) — checa um trecho sem aspas.
        self.assertContains(resp, "só são editáveis com o RI em")
        self.assertContains(resp, "RN-052")

    def test_administrador_tambem_nao_lanca_ixc_fora_de_em_andamento(self):
        """RN-052: o bloqueio vale para os dois perfis, sem exceção de
        Administrador — mesmo espírito da RN-020."""
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 0)

    def test_lado_ixc_bloqueia_edicao_e_exclusao_de_item_fora_de_em_andamento(self):
        """RN-052: com o item já lançado, editar/excluir também exige o RI
        em "Em Andamento" — mesmo padrão de bloqueio de campo da RN-020,
        agora aplicado a este lado."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        item = RiItemIxc.objects.create(
            ri=ri, descricao_item="Roteador", quantidade=1, valor_unitario="100.00"
        )
        self.client.force_login(self.admin)
        self.client.post(
            reverse("ri_item_ixc_update", kwargs={"item_pk": item.pk}),
            {"descricao_item": "Roteador Wi-Fi 6", "quantidade": 2, "valor_unitario": "150.00"},
        )
        item.refresh_from_db()
        self.assertEqual(item.descricao_item, "Roteador")

        self.client.post(reverse("ri_item_ixc_delete", kwargs={"item_pk": item.pk}))
        self.assertTrue(RiItemIxc.objects.filter(pk=item.pk).exists())

    def test_lado_ixc_libera_lancamento_quando_ri_volta_para_em_andamento(self):
        """RN-052: não é um travamento permanente — com o RI em "Em
        Andamento" o Lado IXC aceita lançamento normalmente."""
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_ixc", "kit": kit.pk, **self.FORMSET_PRODUTO_VAZIO},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemIxc.objects.filter(ri=ri).count(), 1)

    def test_lado_relatorio_eace_continua_editavel_fora_de_em_andamento(self):
        """RN-052 não altera o Lado Relatório EACE — usuário pediu o
        bloqueio só para o Lado IXC (2º lado); RN-020 segue sozinha
        valendo ali. Com o RI em "Implantação EACE" (fora de "Em
        Andamento", já bloqueado para o Lado IXC), o Relatório EACE
        continua aceitando lançamento normalmente."""
        kit = KitPadrao.objects.create(descricao="Kit Wi-Fi Indoor", unidade="Escola")
        ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "salvar_relatorio_eace", "eace-kit": kit.pk, **self.FORMSET_PRODUTO_EACE_VAZIO},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=ri).count(), 1)


class RiDetailStatusEnvioEmailTests(TestCase):
    """RN-051 (2026-09-02): troca de status direto na tela de detalhe do
    RI (mesmo padrão do drill-down do grid) — "Envio de Email para
    Faturamento" só aparece no <select> quando as regras de negócio do
    envio (RN-013) estão satisfeitas hoje; o botão/modal "Enviar e-mail"
    aparece na própria tela assim que o status muda para esse, sem F5
    (HTMX, `origem=detail`)."""

    def setUp(self):
        self.analista = User.objects.create_user(
            username="analista-envio-detail", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )
        self.escola = Escola.objects.create(
            inep="80000001", nome="Escola Envio Email Detail", municipio="Fortaleza", estado="CE",
        )

    def _preencher_requisitos_envio(self, ri):
        """Deixa o RI pronto para "Envio de Email para Faturamento"
        (mesma checagem de `gerar_planilha_faturamento`, RN-013)."""
        RiItemIxc.objects.create(
            ri=ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario="0", eh_kit=True,
        )
        ri.data_ativacao = date(2026, 8, 1)
        ri.municipio_ixc = "Fortaleza"
        ri.estado_ixc = "CE"
        ri.cnpj = "00.000.000/0001-00"
        ri.cnpj_ficticio = "11.111.111/0001-11"
        ri.save()

    def test_opcao_envio_email_nao_aparece_quando_nao_pronto(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, '<option value="envio_email_faturamento"')

    def test_opcao_envio_email_aparece_quando_pronto(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self._preencher_requisitos_envio(ri)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, '<option value="envio_email_faturamento"')

    def test_opcao_envio_email_continua_visivel_quando_ja_e_o_status_atual(self):
        """Mesmo que algo torne o RI "não pronto" depois (ex.: item do KIT
        excluído), o status atual continua representado no <select> — sem
        isso, o navegador selecionaria outra opção qualquer, mentindo
        sobre o status real."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, '<option value="envio_email_faturamento" selected')

    def test_botao_enviar_email_nao_aparece_fora_do_status_de_envio(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self._preencher_requisitos_envio(ri)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, f'data-abrir-modal-email="modal-email-{ri.pk}"')

    def test_botao_enviar_email_aparece_quando_ja_esta_no_status_de_envio(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ENVIO_EMAIL_FATURAMENTO)
        self._preencher_requisitos_envio(ri)
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, f'data-abrir-modal-email="modal-email-{ri.pk}"')
        self.assertContains(resp, f'INEP {self.escola.inep} — {self.escola.nome}')

    def test_trocar_status_via_htmx_da_tela_de_detalhe_mostra_botao_enviar_email_sem_reload(self):
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self._preencher_requisitos_envio(ri)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {
                "status": Ri.ENVIO_EMAIL_FATURAMENTO,
                "next": reverse("ri_detail", kwargs={"inep": self.escola.inep}),
                "origem": "detail",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Status do RI atualizado.")
        self.assertContains(resp, f'id="status-pill-{ri.pk}"')
        self.assertContains(resp, f'id="acao-envio-email-{ri.pk}"')
        self.assertContains(resp, f'data-abrir-modal-email="modal-email-{ri.pk}"')
        ri.refresh_from_db()
        self.assertEqual(ri.status, Ri.ENVIO_EMAIL_FATURAMENTO)

    def test_trocar_status_via_htmx_da_tela_de_detalhe_para_status_sem_envio_nao_mostra_botao(self):
        """Origem "detail" também repõe o bloco vazio quando o novo status
        não é "Envio de Email para Faturamento" — o botão não fica
        "grudado" na tela depois de uma correção manual do status."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.CORRECAO_MEGA)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {
                "status": Ri.ANDAMENTO,
                "next": reverse("ri_detail", kwargs={"inep": self.escola.inep}),
                "origem": "detail",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertContains(resp, f'id="acao-envio-email-{ri.pk}"')
        self.assertNotContains(resp, f'data-abrir-modal-email="modal-email-{ri.pk}"')

    def test_origem_grid_continua_devolvendo_o_fragmento_do_grid(self):
        """Sem `origem` no POST (comportamento de antes desta feature) ou
        com `origem=grid`, a resposta continua sendo o fragmento do
        drill-down do grid — não o da tela de detalhe."""
        ri = Ri.objects.create(escola=self.escola, status=Ri.ANDAMENTO)
        self._preencher_requisitos_envio(ri)
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_status_update", kwargs={"pk": ri.pk}),
            {"status": Ri.ENVIO_EMAIL_FATURAMENTO, "next": reverse("grid_inep")},
            HTTP_HX_REQUEST="true",
        )
        self.assertContains(resp, f'id="form-status-{ri.pk}"')
        self.assertNotContains(resp, f'id="status-pill-{ri.pk}"')


_MEDIA_ROOT_TESTE_HISTORICO = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_HISTORICO)
class RiHistoricoTests(TestCase):
    """FEAT-014/RN-008: linha do tempo de comunicação do RI — mensagem
    (com anexo opcional) e anexo isolado; log automático fica coberto em
    RiStatusUpdateViewTests (status), RiResponsavelUpdateViewTests
    (responsável) e RiDetailViewTests (cadastro/edição/exclusão dos itens
    do Lado IXC e do Relatório EACE). Envio/recebimento de e-mail (RN-008)
    fica de fora por ora — depende da FEAT-008/009, ainda não implementadas
    (decisão do usuário em 2026-08-22, ver checklist). MEDIA_ROOT isolado
    num diretório temporário para os anexos de teste não irem para o
    `media/` real."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_HISTORICO, ignore_errors=True)

    def setUp(self):
        self.analista = User.objects.create_user(
            username="analista-hist", password="senha-teste-123", perfil=User.PERFIL_ANALISTA
        )
        self.escola = Escola.objects.create(
            inep="40000001", nome="Escola Histórico RI", municipio="Fortaleza", estado="CE"
        )
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)

    def test_mensagem_aparece_na_linha_do_tempo(self):
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Cliente confirmou o endereço."},
        )
        self.assertEqual(resp.status_code, 302)
        entrada = RiHistorico.objects.get(ri=self.ri)
        self.assertEqual(entrada.tipo, RiHistorico.MENSAGEM)
        self.assertEqual(entrada.mensagem, "Cliente confirmou o endereço.")
        self.assertEqual(entrada.autor, self.analista)

    def test_anexo_isolado_sem_mensagem(self):
        arquivo = SimpleUploadedFile("comprovante.txt", b"conteudo do anexo")
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "", "anexo": arquivo},
        )
        entrada = RiHistorico.objects.get(ri=self.ri)
        self.assertEqual(entrada.tipo, RiHistorico.ANEXO)
        self.assertTrue(entrada.anexo.name)

    def test_mensagem_com_anexo_junto(self):
        arquivo = SimpleUploadedFile("nota.txt", b"conteudo")
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Segue anexo.", "anexo": arquivo},
        )
        entrada = RiHistorico.objects.get(ri=self.ri)
        self.assertEqual(entrada.tipo, RiHistorico.MENSAGEM)
        self.assertTrue(entrada.anexo.name)

    def test_exige_mensagem_ou_anexo(self):
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": ""},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Escreva uma mensagem ou anexe um arquivo.")
        self.assertFalse(RiHistorico.objects.filter(ri=self.ri).exists())

    def test_mensagem_acima_do_limite_e_rejeitada(self):
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "x" * 251},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "no máximo 250 caracteres")
        self.assertFalse(RiHistorico.objects.filter(ri=self.ri).exists())

    def test_mais_recente_primeiro(self):
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Primeira mensagem."},
        )
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Segunda mensagem."},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        conteudo = resp.content.decode()
        posicao_segunda = conteudo.index("Segunda mensagem.")
        posicao_primeira = conteudo.index("Primeira mensagem.")
        self.assertLess(posicao_segunda, posicao_primeira)

    def _criar_15_entradas_historico(self):
        """15 entradas com `criado_em` forçado (índice maior = mais
        recente) — `auto_now_add` ignora valor passado no `create()`, por
        isso o ajuste vem de um `update()` logo depois, sem depender da
        precisão de timestamp do banco entre chamadas."""
        agora = timezone.now()
        for indice in range(15):
            entrada = RiHistorico.objects.create(
                ri=self.ri, tipo=RiHistorico.MENSAGEM, autor=self.analista,
                mensagem=f"Mensagem numero {indice}",
            )
            RiHistorico.objects.filter(pk=entrada.pk).update(
                criado_em=agora - timedelta(minutes=15 - indice)
            )

    def test_historico_pagina_1_mostra_so_as_10_mais_recentes(self):
        """RN-008 (2026-08-26): a linha do tempo pagina de 10 em 10 — não
        traz tudo de uma vez, só a página pedida (pedido do usuário para
        não consumir memória com o histórico inteiro)."""
        self._criar_15_entradas_historico()
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Página 1 de 2")
        self.assertContains(resp, "Mensagem numero 14")
        self.assertNotContains(resp, "Mensagem numero 4")

    def test_historico_pagina_2_mostra_as_entradas_mais_antigas(self):
        """Página seguinte só é buscada quando o usuário clica — aqui,
        simulada pelo parâmetro `historico_page` na própria URL."""
        self._criar_15_entradas_historico()
        self.client.force_login(self.analista)
        resp = self.client.get(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}) + "?historico_page=2"
        )
        self.assertContains(resp, "Página 2 de 2")
        self.assertContains(resp, "Mensagem numero 0")
        self.assertNotContains(resp, "Mensagem numero 14")

    def test_historico_sem_paginacao_quando_cabe_em_uma_pagina(self):
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Única mensagem."},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, "Página 1 de")

    def test_entrada_de_email_automatico_sem_autor_renderiza(self):
        """FEAT-009: a leitura automática da resposta do financeiro grava
        `autor=None` (sem usuário logado, `apps/ri/services.py`) — a
        página não pode quebrar com esse tipo de entrada na linha do
        tempo."""
        RiHistorico.objects.create(
            ri=self.ri, tipo=RiHistorico.EMAIL, autor=None, mensagem="E-mail de resposta do financeiro recebido."
        )
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Sistema")

    def test_htmx_adicionar_historico_retorna_fragmento_sem_redirecionar(self):
        """FEAT-019: com o header do HTMX, a linha do tempo volta como
        fragmento (200) em vez de redirecionar — a tela do RI não recarrega
        por completo ao registrar uma nova entrada."""
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "adicionar_historico", "mensagem": "Cliente confirmou o endereço."},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'id="historico"')
        self.assertContains(resp, "Cliente confirmou o endereço.")
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, mensagem="Cliente confirmou o endereço."
            ).exists()
        )


class KitPadraoModelTests(TestCase):
    """RN-010: catálogo de valores fixos por kit, com valor de equipamento
    e de serviço separados por Lote."""

    def test_valor_total_soma_equipamento_e_servico(self):
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 6 Access Points", lote=9, unidade="Escola",
            valor_equipamento="14457.89", valor_servico="20673.51",
        )
        self.assertEqual(float(item.valor_total), 35131.40)

    def test_valor_total_trata_equipamento_ausente_como_zero(self):
        item = KitPadrao.objects.create(
            descricao="Manutenção de Rede Interna", lote=9, unidade="Escola/Mês",
            valor_servico="380.38",
        )
        self.assertEqual(float(item.valor_total), 380.38)

    def test_valor_faturavel_e_so_o_equipamento_nao_soma_servico(self):
        """Correção pedida pelo usuário em 2026-08-31: o Valor Unitário
        faturado (KIT ou Produto) vem só do equipamento — o valor de
        serviço não entra, ao contrário de `valor_total`. Exemplo real
        reportado: Kit 15 tinha que ficar em R$ 26.353,13 (só equipamento),
        não ~R$ 51 mil (equipamento + serviço)."""
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 15 Access Points", lote=9, unidade="Escola",
            valor_equipamento="26353.13", valor_servico="25000.00",
        )
        self.assertEqual(item.valor_faturavel, Decimal("26353.13"))
        self.assertNotEqual(item.valor_faturavel, item.valor_total)

    def test_valor_faturavel_trata_equipamento_ausente_como_zero(self):
        item = KitPadrao.objects.create(
            descricao="Manutenção de Rede Interna", lote=9, unidade="Escola/Mês",
            valor_servico="380.38",
        )
        self.assertEqual(item.valor_faturavel, Decimal("0"))

    def test_kit_fechado_por_escola_quando_unidade_e_escola(self):
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 1 Access Point", lote=9, unidade="Escola",
        )
        self.assertTrue(item.kit_fechado_por_escola)

    def test_kit_fechado_por_escola_vale_para_escola_mes(self):
        item = KitPadrao.objects.create(
            descricao="Manutenção de Rede Interna", lote=9, unidade="Escola/Mês",
        )
        self.assertTrue(item.kit_fechado_por_escola)

    def test_nao_e_kit_fechado_quando_unidade_e_item_avulso(self):
        item = KitPadrao.objects.create(
            descricao="Access Point adicional Indoor", lote=9, unidade="Unidade",
        )
        self.assertFalse(item.kit_fechado_por_escola)

    def test_mesma_descricao_permite_lote_diferente(self):
        KitPadrao.objects.create(descricao="Kit X", lote=9, unidade="Escola")
        KitPadrao.objects.create(descricao="Kit X", lote=11, unidade="Escola")
        self.assertEqual(KitPadrao.objects.filter(descricao="Kit X").count(), 2)

    def test_mesma_descricao_e_lote_e_bloqueada(self):
        KitPadrao.objects.create(descricao="Kit X", lote=9, unidade="Escola")
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                KitPadrao.objects.create(descricao="Kit X", lote=9, unidade="Escola")

    def test_descricao_curta_tira_qualificador_entre_parenteses(self):
        """RN-011 (2026-08-24): a lista do Lado IXC usa a Descrição curta,
        não a Descrição completa (grande demais para um select)."""
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 8 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
        )
        self.assertEqual(item.descricao_curta, "Kit Cobertura Wi-Fi - 8 Access Points")

    def test_descricao_curta_sem_parenteses_usa_a_descricao_inteira(self):
        item = KitPadrao.objects.create(descricao="Kit X", lote=9, unidade="Escola")
        self.assertEqual(item.descricao_curta, "Kit X")

    def test_descricao_curta_digitada_a_mao_nao_e_sobrescrita(self):
        """Campo pode ser digitado à mão (admin) para um caso que a regra
        automática não resolva bem — não é sobrescrita ao salvar de novo."""
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 8 Access Points (serviços, materiais e equipamentos)",
            descricao_curta="Kit 8 APs",
            lote=9, unidade="Escola",
        )
        item.valor_servico = "100.00"
        item.save()
        item.refresh_from_db()
        self.assertEqual(item.descricao_curta, "Kit 8 APs")

    def test_numero_access_points_derivado_da_descricao(self):
        """RN-010 ampliada (FEAT-016): número extraído automaticamente,
        mesmo padrão de derivação já usado por descricao_curta."""
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
        )
        self.assertEqual(item.numero_access_points, 4)

    def test_numero_access_points_vazio_quando_descricao_nao_segue_o_padrao(self):
        item = KitPadrao.objects.create(
            descricao="Cabo de rede", lote=9, unidade="metro",
        )
        self.assertIsNone(item.numero_access_points)

    def test_numero_access_points_digitado_a_mao_nao_e_sobrescrito(self):
        item = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points",
            numero_access_points=99,
            lote=9, unidade="Escola",
        )
        item.valor_servico = "100.00"
        item.save()
        item.refresh_from_db()
        self.assertEqual(item.numero_access_points, 99)

    def test_resolver_kit_declarado_por_numero_cruza_com_access_points(self):
        """RN-010 ampliada: Escola.kit_inicial só com o número da EACE (ex.
        "4") cruza pela quantidade de Access Points, não pelo texto."""
        esperado = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points (serviços, materiais e equipamentos)",
            lote=9, unidade="Escola",
        )
        self.assertEqual(KitPadrao.resolver_kit_declarado("4", lote=9), esperado)

    def test_resolver_kit_declarado_por_texto_completo_continua_igual(self):
        """Comportamento original da RN-010 não muda quando kit_inicial já
        é o texto completo."""
        esperado = KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=9, unidade="Escola",
        )
        self.assertEqual(KitPadrao.resolver_kit_declarado("Kit Wi-Fi Indoor", lote=9), esperado)

    def test_resolver_kit_declarado_sem_correspondencia_retorna_none(self):
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", lote=9, unidade="Escola",
        )
        self.assertIsNone(KitPadrao.resolver_kit_declarado("8", lote=9))
        self.assertIsNone(KitPadrao.resolver_kit_declarado("", lote=9))

    def test_resolver_kit_declarado_respeita_o_lote(self):
        """O mesmo número de Access Points pode ter preço/registro
        diferente por Lote (RN-010) — não cruza fora do Lote informado."""
        do_lote_9 = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", lote=9, unidade="Escola",
        )
        do_lote_11 = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", lote=11, unidade="Escola",
        )
        self.assertEqual(KitPadrao.resolver_kit_declarado("4", lote=9), do_lote_9)
        self.assertEqual(KitPadrao.resolver_kit_declarado("4", lote=11), do_lote_11)

    def test_resolver_nobreak_declarado_cruza_por_descricao_curta(self):
        """RN-017 (correção 2026-08-27): "Nobreak" (`Escola.nobreak_inicial`)
        cruza com a Descrição completa da LPU pela `descricao_curta`
        derivada automaticamente (remove o qualificador entre
        parênteses)."""
        esperado = KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=9, unidade="Unidade",
        )
        self.assertEqual(KitPadrao.resolver_nobreak_declarado("Nobreak", lote=9), esperado)

    def test_resolver_nobreak_declarado_sem_correspondencia_retorna_none(self):
        self.assertIsNone(KitPadrao.resolver_nobreak_declarado("Nobreak", lote=9))
        self.assertIsNone(KitPadrao.resolver_nobreak_declarado("", lote=9))

    def test_resolver_nobreak_declarado_respeita_o_lote(self):
        do_lote_9 = KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=9, unidade="Unidade",
        )
        do_lote_11 = KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=11, unidade="Unidade",
        )
        self.assertEqual(KitPadrao.resolver_nobreak_declarado("Nobreak", lote=9), do_lote_9)
        self.assertEqual(KitPadrao.resolver_nobreak_declarado("Nobreak", lote=11), do_lote_11)

    def test_resolver_nobreak_declarado_usa_catalogo_em_memoria_sem_consulta_extra(self):
        item = KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=9, unidade="Unidade",
        )
        catalogo = list(KitPadrao.objects.all())
        with self.assertNumQueries(0):
            self.assertEqual(
                KitPadrao.resolver_nobreak_declarado("Nobreak", lote=9, catalogo=catalogo), item
            )


def _criar_planilha_lpu(tmp_path, linhas, lotes=(9, 11), linha_lotes=3, linha_cabecalho=4):
    """Monta um .xlsx no mesmo formato da aba LPU do CONSOLIDADO EACE.xlsx
    real: bloco de 3 colunas (Equipamentos/Serviços/Valor Total) por Lote,
    com o rótulo do Lote mesclado só na primeira coluna do bloco."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LPU"
    ws.cell(row=1, column=1, value="TABELA 1 - LISTA DE PREÇOS UNITÁRIOS (LPU)")

    for indice, titulo in enumerate(["Item", "Descrição", "Unidade"], start=1):
        ws.cell(row=linha_cabecalho, column=indice, value=titulo)

    coluna = 4
    for lote in lotes:
        ws.cell(row=linha_lotes, column=coluna, value=f"LOTE {lote}")
        for titulo in ("Equipamentos (R$)", "Serviços (R$)", "Valor Total (R$)"):
            ws.cell(row=linha_cabecalho, column=coluna, value=titulo)
            coluna += 1

    for offset, linha in enumerate(linhas, start=1):
        for indice, valor in enumerate(linha, start=1):
            ws.cell(row=linha_cabecalho + offset, column=indice, value=valor)

    caminho = tmp_path / "planilha_lpu_teste.xlsx"
    wb.save(caminho)
    return caminho


class ImportarCatalogoLpuTests(TestCase):
    """FEAT-015/RN-010: importação do catálogo KitPadrao a partir da aba
    LPU do CONSOLIDADO EACE.xlsx."""

    def setUp(self):
        self._tmp_dir = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp_dir.name)
        self.addCleanup(self._tmp_dir.cleanup)

    def test_importa_valores_de_equipamento_e_servico_por_lote(self):
        caminho = _criar_planilha_lpu(self.tmp_path, [
            ["REDE INTERNA"],
            [1, "Kit Cobertura Wi-Fi - 1 Access Point", "Escola",
             8276.72, 15728.61, 24005.33, 7955.58, 15118.34, 23073.92],
        ])

        call_command("importar_catalogo_lpu", str(caminho))

        self.assertEqual(KitPadrao.objects.count(), 2)
        lote9 = KitPadrao.objects.get(descricao="Kit Cobertura Wi-Fi - 1 Access Point", lote=9)
        self.assertEqual(lote9.unidade, "Escola")
        self.assertEqual(float(lote9.valor_equipamento), 8276.72)
        self.assertEqual(float(lote9.valor_servico), 15728.61)
        lote11 = KitPadrao.objects.get(descricao="Kit Cobertura Wi-Fi - 1 Access Point", lote=11)
        self.assertEqual(float(lote11.valor_equipamento), 7955.58)

    def test_item_avulso_sem_valor_de_equipamento(self):
        caminho = _criar_planilha_lpu(self.tmp_path, [
            ["REDE INTERNA"],
            [19, "Manutenção de Rede Interna", "Escola/Mês",
             None, 380.38, None, None, 365.62, None],
        ])

        call_command("importar_catalogo_lpu", str(caminho))

        item = KitPadrao.objects.get(descricao="Manutenção de Rede Interna", lote=9)
        self.assertIsNone(item.valor_equipamento)
        self.assertEqual(float(item.valor_servico), 380.38)

    def test_linha_de_secao_nao_gera_registro(self):
        caminho = _criar_planilha_lpu(self.tmp_path, [
            ["REDE EXTERNA"],
            [1, "Construção de Rede de acesso em fibra óptica", "km",
             None, 8500, None, None, 8500, None],
            ["REDE INTERNA"],
        ])

        call_command("importar_catalogo_lpu", str(caminho))

        self.assertEqual(KitPadrao.objects.count(), 2)  # só o item 1, 1 por lote
        self.assertFalse(KitPadrao.objects.filter(descricao__in=["REDE EXTERNA", "REDE INTERNA"]).exists())

    def test_rodape_de_notas_e_ignorado(self):
        caminho = _criar_planilha_lpu(self.tmp_path, [
            ["REDE INTERNA"],
            [1, "Kit Cobertura Wi-Fi - 1 Access Point", "Escola",
             8276.72, 15728.61, 24005.33, 7955.58, 15118.34, 23073.92],
            [None],
            ["Notas:"],
            [None, "Texto de nota que não é item, não deve virar registro."],
        ])

        call_command("importar_catalogo_lpu", str(caminho))

        self.assertEqual(KitPadrao.objects.count(), 2)

    def test_reimportar_atualiza_sem_duplicar(self):
        caminho = _criar_planilha_lpu(self.tmp_path, [
            ["REDE INTERNA"],
            [1, "Kit Cobertura Wi-Fi - 1 Access Point", "Escola",
             8276.72, 15728.61, 24005.33, 7955.58, 15118.34, 23073.92],
        ])
        call_command("importar_catalogo_lpu", str(caminho))

        caminho_atualizado = _criar_planilha_lpu(self.tmp_path, [
            ["REDE INTERNA"],
            [1, "Kit Cobertura Wi-Fi - 1 Access Point", "Escola",
             9000.00, 15728.61, 24728.61, 7955.58, 15118.34, 23073.92],
        ])
        call_command("importar_catalogo_lpu", str(caminho_atualizado))

        self.assertEqual(KitPadrao.objects.count(), 2)
        lote9 = KitPadrao.objects.get(descricao="Kit Cobertura Wi-Fi - 1 Access Point", lote=9)
        self.assertEqual(float(lote9.valor_equipamento), 9000.00)

    def test_aba_inexistente_gera_erro_claro(self):
        wb = openpyxl.Workbook()
        wb.active.title = "OUTRA ABA"
        caminho = self.tmp_path / "planilha_sem_aba.xlsx"
        wb.save(caminho)

        with self.assertRaises(Exception):
            call_command("importar_catalogo_lpu", str(caminho))

    def test_arquivo_inexistente_gera_erro_claro(self):
        with self.assertRaises(Exception):
            call_command("importar_catalogo_lpu", str(self.tmp_path / "nao_existe.xlsx"))


_CABECALHO_PLANILHA_EACE = (
    "Projeto;Cod Fornecedor;Fornecedor;CNPJ;Num Obra;Cod Produto;Descrição do Item;"
    "Qtde Produto;Valor Unit UR;Valor Produto;Prod Serv;Previsão de execução;"
    "Num Provisório;Num OSP;Validação OSP;Status OSP;Nota Fiscal;Enviado SAP;"
    # RN-046: "Fase escola" (coluna S) e "Status escola" (coluna T, real).
    "Fase escola;Status escola;Unique ID\n"
)


def _csv_planilha_eace(nome="EACE.csv", linhas=None):
    conteudo = _CABECALHO_PLANILHA_EACE + "".join(linhas or [])
    return SimpleUploadedFile(nome, conteudo.encode("utf-8"), content_type="text/csv")


# RN-021 (ajuste 2026-09-02, pedido do usuário): mesmas colunas de
# `_CABECALHO_PLANILHA_EACE`, mas no formato BRUTO exportado direto do
# sistema da EACE — vírgula como delimitador, campos entre aspas (a
# vírgula também aparece dentro dos valores numéricos, formato BR:
# "21.765,83" — só funciona citado). O sistema detecta sozinho qual dos
# dois formatos foi enviado (`detectar_delimitador_planilha_eace`).
_CABECALHO_PLANILHA_EACE_BRUTA = (
    'Projeto,Cod Fornecedor,Fornecedor,CNPJ,Num Obra,Cod Produto,Descrição do Item,'
    'Qtde Produto,Valor Unit UR,Valor Produto,Prod Serv,Previsão de execução,'
    'Num Provisório,Num OSP,Validação OSP,Status OSP,Nota Fiscal,Enviado SAP,'
    'Fase escola,Status escola,Unique ID\n'
)


def _csv_planilha_eace_bruta(nome="Documento correto.csv", linhas=None):
    conteudo = _CABECALHO_PLANILHA_EACE_BRUTA + "".join(linhas or [])
    return SimpleUploadedFile(nome, conteudo.encode("utf-8"), content_type="text/csv")


def _linha_planilha_eace_bruta(
    inep, descricao, qtd="1", num_osp="2919", validacao_osp="Aprovado", nota_fiscal="289",
    status_escola="", valor="21.765,83",
):
    # Fornecedor com vírgula dentro do valor + "Valor Unit UR"/"Valor
    # Produto" no formato BR (vírgula decimal) — os dois só ficam intactos
    # (não quebram em coluna a mais) porque estão entre aspas; é
    # justamente o caso que o parser tratado (`;`) não tinha que resolver.
    return (
        f'"{inep}","19001","Fornecedor Teste, Filial SP","00.000.000/0001-00","OBRA","",'
        f'"{descricao}","{qtd}","{valor}","{valor}","Material","06/07/2026","3188",'
        f'"{num_osp}","{validacao_osp}","","{nota_fiscal}","Não","5","{status_escola}","UID-TESTE"\n'
    )


class DetectarDelimitadorPlanilhaEaceTests(TestCase):
    """RN-021 (ajuste 2026-09-02): detecção automática do delimitador da
    Planilha EACE — `;` (formato já tratado) ou `,` (formato bruto, campos
    entre aspas, exportado direto da EACE)."""

    def test_reconhece_formato_tratado_com_ponto_e_virgula(self):
        self.assertEqual(detectar_delimitador_planilha_eace(_CABECALHO_PLANILHA_EACE), ";")

    def test_reconhece_formato_bruto_com_virgula_e_aspas(self):
        self.assertEqual(detectar_delimitador_planilha_eace(_CABECALHO_PLANILHA_EACE_BRUTA), ",")

    def test_arquivo_fora_dos_2_formatos_nao_e_reconhecido(self):
        self.assertIsNone(detectar_delimitador_planilha_eace("Projeto,Descricao\n"))


_MEDIA_ROOT_TESTE_PLANILHA_EACE = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_PLANILHA_EACE)
class PlanilhaEaceViewTests(TestCase):
    """FEAT-023/RN-021: upload da Planilha EACE (Administrador). MEDIA_ROOT
    isolado num diretório temporário para o arquivo de teste não ir para o
    `media/` real."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_PLANILHA_EACE, ignore_errors=True)

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin-planilha-eace", password="senha-teste-123",
            perfil=User.PERFIL_ADMINISTRADOR,
        )
        self.analista = User.objects.create_user(
            username="analista-planilha-eace", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )

    def test_upload_valido_cria_planilha_ativa(self):
        self.client.force_login(self.admin)
        arquivo = _csv_planilha_eace(
            linhas=["53004230;;;;;;Kit Cobertura Wi-Fi - 4 Access Points;1;100,00;100,00\n"]
        )
        resp = self.client.post(reverse("planilha_eace"), {"arquivo": arquivo})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(PlanilhaEace.objects.count(), 1)
        planilha = PlanilhaEace.objects.first()
        self.assertEqual(planilha.nome_original, "EACE.csv")
        self.assertEqual(planilha.enviado_por, self.admin)

    def test_upload_com_coluna_faltando_e_rejeitado(self):
        self.client.force_login(self.admin)
        arquivo = SimpleUploadedFile(
            "EACE.csv", b"Projeto;Descricao\n53004230;Kit Wi-Fi\n", content_type="text/csv"
        )
        resp = self.client.post(reverse("planilha_eace"), {"arquivo": arquivo})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(PlanilhaEace.objects.count(), 0)
        self.assertContains(resp, "Colunas obrigatórias ausentes")

    def test_upload_aceita_arquivo_bruto_exportado_direto_da_eace(self):
        """RN-021 (ajuste 2026-09-02, pedido do usuário): a planilha
        exportada direto do sistema da EACE vem com vírgula como
        delimitador e campos entre aspas (a vírgula também aparece dentro
        dos valores numéricos) — o sistema detecta e aceita sozinho, sem
        o usuário precisar tratar o arquivo à mão antes de subir."""
        self.client.force_login(self.admin)
        arquivo = _csv_planilha_eace_bruta(
            linhas=[_linha_planilha_eace_bruta("53004230", "Kit Cobertura Wi-Fi - 4 Access Points")]
        )
        resp = self.client.post(reverse("planilha_eace"), {"arquivo": arquivo})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(PlanilhaEace.objects.count(), 1)
        self.assertEqual(PlanilhaEace.objects.first().nome_original, "Documento correto.csv")

    def test_upload_bruto_com_coluna_faltando_e_rejeitado(self):
        self.client.force_login(self.admin)
        arquivo = SimpleUploadedFile(
            "Documento correto.csv",
            'Projeto,Descricao\n"53004230","Kit Wi-Fi"\n'.encode("utf-8"),
            content_type="text/csv",
        )
        resp = self.client.post(reverse("planilha_eace"), {"arquivo": arquivo})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(PlanilhaEace.objects.count(), 0)
        self.assertContains(resp, "Colunas obrigatórias ausentes")

    def test_novo_upload_substitui_anterior(self):
        self.client.force_login(self.admin)
        self.client.post(reverse("planilha_eace"), {
            "arquivo": _csv_planilha_eace(nome="EACE_v1.csv"),
        })
        arquivo_antigo = PlanilhaEace.objects.first().arquivo
        self.client.post(reverse("planilha_eace"), {
            "arquivo": _csv_planilha_eace(nome="EACE_v2.csv"),
        })
        self.assertEqual(PlanilhaEace.objects.count(), 1)
        self.assertEqual(PlanilhaEace.objects.first().nome_original, "EACE_v2.csv")
        self.assertFalse(arquivo_antigo.storage.exists(arquivo_antigo.name))

    def test_analista_nao_acessa(self):
        self.client.force_login(self.analista)
        resp = self.client.get(reverse("planilha_eace"))
        self.assertEqual(resp.status_code, 403)

    def test_pagina_exibe_planilha_ativa(self):
        PlanilhaEace.substituir(_csv_planilha_eace(), self.admin)
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("planilha_eace"))
        self.assertContains(resp, "EACE.csv")

    def test_input_de_arquivo_mostra_rotulo_em_portugues(self):
        """Usuário reportou (2026-08-27) que o input de arquivo mostrava
        o texto nativo do navegador em inglês ("Choose File"/"No file
        chosen") — substituído por rótulo + texto próprios da tela."""
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("planilha_eace"))
        self.assertContains(resp, "Escolher arquivo")
        self.assertContains(resp, "Nenhum arquivo selecionado")
        self.assertNotContains(resp, "Choose File")


def _linha_planilha_eace(
    inep, descricao, qtd="1", num_osp="2919", validacao_osp="Aprovado", nota_fiscal="289",
    status_escola="",
):
    # RN-046: `status_escola` vazio por padrão nos testes que não tratam
    # dela (campo gravado por item, sem efeito no status do RI).
    return (
        f"{inep};19001;Fornecedor Teste;00.000.000/0001-00;OBRA;COD1;{descricao};{qtd};"
        f"100,00;100,00;Material;06/07/2026;3188;{num_osp};{validacao_osp};;{nota_fiscal};Não;"
        f"5;{status_escola};UID-TESTE\n"
    )


_MEDIA_ROOT_TESTE_SINCRONIZADOR = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
class SincronizarRelatorioEaceDaPlanilhaTests(TestCase):
    """FEAT-024/RN-022: Sincronizador do Lado Relatório EACE a partir da
    Planilha EACE (RN-021), casada com o catálogo `KitPadrao` pelo INEP.
    MEDIA_ROOT isolado num diretório temporário para o arquivo de teste
    não ir para o `media/` real."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_SINCRONIZADOR, ignore_errors=True)

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin-sync-eace", password="senha-teste-123",
            perfil=User.PERFIL_ADMINISTRADOR,
        )
        self.analista = User.objects.create_user(
            username="analista-sync-eace", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )
        self.escola = Escola.objects.create(
            inep="60000001", nome="Escola Sincronizador", municipio="Fortaleza", estado="CE"
        )
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)
        self.kit = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", unidade="Escola",
            valor_equipamento=Decimal("300.00"), valor_servico=Decimal("50.00"),
        )
        self.nobreak = KitPadrao.objects.create(
            descricao="Nobreak", unidade="Unidade", valor_equipamento=Decimal("150.00"),
        )

    def _upload_planilha(self, linhas):
        PlanilhaEace.substituir(_csv_planilha_eace(linhas=linhas), self.admin)

    def test_sincroniza_a_partir_do_arquivo_bruto_exportado_direto_da_eace(self):
        """RN-021 (ajuste 2026-09-02, pedido do usuário): mesmo
        Sincronizador, agora alimentado por uma Planilha EACE no formato
        BRUTO (vírgula + aspas, "Fornecedor" com vírgula dentro do valor)
        — sem precisar tratar o arquivo à mão antes de subir; o resultado
        é idêntico ao do formato já tratado."""
        PlanilhaEace.substituir(
            _csv_planilha_eace_bruta(linhas=[
                _linha_planilha_eace_bruta(
                    self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO"
                ),
            ]),
            self.admin,
        )
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["criados"]), 1)
        item = resultado["criados"][0]
        self.assertEqual(item.descricao_item, "Kit Cobertura Wi-Fi - 4 Access Points")
        self.assertTrue(item.eh_kit)
        self.assertEqual(item.quantidade, 1)
        # Valor vem sempre do catálogo (KitPadrao), nunca do valor bruto da
        # planilha — confere que o "Fornecedor" com vírgula dentro do
        # campo (só protegido pelas aspas) não descolou as colunas
        # seguintes da linha.
        self.assertEqual(item.valor_unitario, Decimal("300.00"))
        self.assertEqual(item.num_osp, "2919")
        self.assertEqual(item.validacao_osp, "Aprovado")
        self.assertEqual(item.nota_fiscal, "289")

    def test_sincroniza_kit_apesar_do_sufixo_extra_no_texto(self):
        """RN-022: a Descrição real da planilha traz sufixo de fornecedor/
        UF que o catálogo não tem — o casamento usa o número de Access
        Points, não o texto inteiro."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO"
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["criados"]), 1)
        item = resultado["criados"][0]
        self.assertEqual(item.descricao_item, "Kit Cobertura Wi-Fi - 4 Access Points")
        self.assertTrue(item.eh_kit)
        self.assertEqual(item.quantidade, 1)
        self.assertEqual(item.valor_unitario, Decimal("300.00"))
        self.assertEqual(item.num_osp, "2919")
        self.assertEqual(item.validacao_osp, "Aprovado")
        self.assertEqual(item.nota_fiscal, "289")

    def test_sincroniza_produto_avulso_por_prefixo(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["criados"]), 1)
        item = resultado["criados"][0]
        self.assertEqual(item.descricao_item, "Nobreak")
        self.assertFalse(item.eh_kit)
        self.assertEqual(item.quantidade, 2)
        self.assertEqual(item.valor_unitario, Decimal("150.00"))

    def test_item_sem_correspondencia_nao_e_lancado(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Produto Nunca Cadastrado - Equip - MEGA - CO"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(resultado["criados"], [])
        self.assertIn(
            "Produto Nunca Cadastrado - Equip - MEGA - CO", resultado["sem_correspondencia"]
        )
        self.assertEqual(RiItemRelatorioEace.objects.count(), 0)

    def test_sem_planilha_ativa_gera_erro(self):
        with self.assertRaises(PlanilhaEaceSincronizacaoError):
            sincronizar_relatorio_eace_da_planilha(self.ri)

    def test_sem_linha_para_o_inep_gera_erro(self):
        self._upload_planilha([_linha_planilha_eace("99999999", "Nobreak")])
        with self.assertRaises(PlanilhaEaceSincronizacaoError):
            sincronizar_relatorio_eace_da_planilha(self.ri)

    def test_kit_ja_lancado_e_ignorado_sem_bloquear_produtos(self):
        """RN-015 (estendida pela RN-018): 1 KIT por INEP também vale para
        o Sincronizador — a linha do KIT é ignorada, o resto continua."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO"
            ),
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["criados"]), 1)
        self.assertFalse(resultado["criados"][0].eh_kit)
        self.assertEqual(resultado["kit_ignorado"], ["Kit Cobertura Wi-Fi - 4 Access Points"])

    def test_sincronizar_duas_vezes_nao_duplica(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        primeira = sincronizar_relatorio_eace_da_planilha(self.ri)
        segunda = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(primeira["criados"]), 1)
        self.assertEqual(segunda["criados"], [])
        self.assertEqual(segunda["duplicados"], ["Nobreak"])
        self.assertEqual(RiItemRelatorioEace.objects.count(), 1)

    def test_sincronizar_nao_altera_status_do_ri(self):
        """RN-024 (retirada, 2026-09-02, pedido do usuário): sincronizar o
        Relatório EACE não altera mais o status do RI, nem quando alguma
        linha traz "Status escola" = "Conectada" — só lança os itens do
        Lado 3, sem nenhum efeito sobre o status."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Conectada"
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.IMPLANTACAO_EACE)
        self.assertIsNone(self.ri.concluido_em)
        self.assertEqual(len(resultado["criados"]), 1)
        self.assertFalse(
            RiHistorico.objects.filter(ri=self.ri, tipo=RiHistorico.LOG_STATUS).exists()
        )

    def test_sincronizar_nao_encerra_correcao_mega(self):
        """RN-024 (retirada): RI em "Correção MEGA" também não é mais
        concluído automaticamente por "Status escola" = "Conectada"."""
        self.ri.status = Ri.CORRECAO_MEGA
        self.ri.save(update_fields=["status"])
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Conectada"
            ),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.CORRECAO_MEGA)

    def test_view_sincronizador_nao_altera_status_e_mostra_mensagem_de_sucesso(self):
        """Mesmo comportamento pela view (botão Sincronizador): item
        lançado, status intacto, mensagem de sucesso."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Conectada"
            ),
        ])
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
            follow=True,
        )
        mensagens = list(resp.context["messages"])
        self.assertEqual(mensagens[-1].tags, "success")
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.IMPLANTACAO_EACE)

    def test_view_lanca_via_botao_sincronizador(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(RiItemRelatorioEace.objects.get(ri=self.ri).descricao_item, "Nobreak")

    def test_salvar_vazio_apos_sincronizar_confirma_que_ja_esta_salvo(self):
        """Clicar em "Salvar" (formulário manual) por engano depois de
        sincronizar não deve soar como erro — os itens já estão salvos e o
        campo "Kit" some da tela (RN-015), então o formulário vazio não
        tem nada para lançar."""
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {
                "acao": "salvar_relatorio_eace",
                "eace_produto-TOTAL_FORMS": "0",
                "eace_produto-INITIAL_FORMS": "0",
                "eace_produto-MIN_NUM_FORMS": "0",
                "eace_produto-MAX_NUM_FORMS": "1000",
            },
            follow=True,
        )
        self.assertContains(resp, "já estão salvos")
        mensagens = list(resp.context["messages"])
        self.assertEqual(mensagens[-1].tags, "success")

    def test_view_sem_planilha_ativa_mostra_mensagem_de_erro(self):
        self.client.force_login(self.analista)
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
            follow=True,
        )
        self.assertContains(resp, "Nenhuma Planilha EACE ativa")
        self.assertEqual(RiItemRelatorioEace.objects.count(), 0)

    def test_nota_fiscal_e_por_item_nao_por_ri(self):
        """RN-022 (ampliada): a Nota Fiscal vem da linha da planilha que
        originou cada item — pode variar entre o KIT e um Produto do
        mesmo INEP (caso real observado pelo usuário: 289 no KIT, 290 no
        Nobreak), então fica gravada por item, não um valor só por RI."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO",
                nota_fiscal="289",
            ),
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", nota_fiscal="290"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        notas_fiscais = {item.descricao_item: item.nota_fiscal for item in resultado["criados"]}
        self.assertEqual(notas_fiscais["Kit Cobertura Wi-Fi - 4 Access Points"], "289")
        self.assertEqual(notas_fiscais["Nobreak"], "290")

    def test_item_lancado_manualmente_nao_tem_campos_fechados(self):
        """RN-022 (ampliada): Num OSP/Validação OSP/Nota Fiscal são
        campos fechados — só o Sincronizador preenche; um item lançado
        manualmente (fora dele) nasce com os três em branco."""
        item = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Cabo de rede", quantidade=1, valor_unitario="5.50",
        )
        self.assertEqual(item.num_osp, "")
        self.assertEqual(item.validacao_osp, "")
        self.assertEqual(item.nota_fiscal, "")

    def test_tela_mostra_campos_fechados_em_verde_apos_sincronizar(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Num OSP")
        self.assertContains(resp, "Validação OSP")
        self.assertContains(resp, "Nota Fiscal")

    def test_status_escola_gravado_por_item(self):
        """RN-046: o valor da coluna "Status escola" (coluna T) da linha
        que originou o item fica gravado nele, igual a Num OSP/Validação
        OSP/Nota Fiscal (RN-022 ampliada)."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Em implantação"
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(resultado["criados"][0].status_escola, "Em implantação")

    def test_item_lancado_manualmente_nao_tem_status_escola(self):
        item = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Cabo de rede", quantidade=1, valor_unitario="5.50",
        )
        self.assertEqual(item.status_escola, "")

    def test_comparar_status_escola_sem_divergencia_quando_todos_iguais(self):
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
            status_escola="Em implantação",
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=1, valor_unitario=Decimal("150.00"),
            status_escola="Em implantação",
        )
        resultado = comparar_status_escola_relatorio(self.ri)
        self.assertFalse(resultado["diverge"])

    def test_comparar_status_escola_diverge_entre_produtos_do_mesmo_inep(self):
        """RN-046: "Status escola" diferente entre 2 produtos do mesmo RI
        é divergência, mesmo sem nenhum lado de referência."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
            status_escola="Conectada",
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=1, valor_unitario=Decimal("150.00"),
            status_escola="Em implantação",
        )
        resultado = comparar_status_escola_relatorio(self.ri)
        self.assertTrue(resultado["diverge"])
        self.assertEqual(resultado["valores"], ["Conectada", "Em implantação"])

    def test_comparar_status_escola_ignora_item_sem_valor(self):
        """RN-046: item lançado manualmente (sem "Status escola") não
        entra na comparação — não é tratado como um 3º valor divergente."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
            status_escola="Conectada",
        )
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Cabo de rede", quantidade=1, valor_unitario="5.50",
        )
        resultado = comparar_status_escola_relatorio(self.ri)
        self.assertFalse(resultado["diverge"])

    def test_tela_mostra_alerta_e_itens_vermelhos_quando_status_escola_diverge(self):
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO",
                status_escola="Conectada",
            ),
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Em implantação"
            ),
        ])
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertContains(resp, "Divergência Status EACE")
        self.assertContains(resp, "ring-red-400")

    def test_sincronizar_de_novo_atualiza_status_escola_de_item_ja_lancado(self):
        """RN-046 (correção, 2026-08-28): usuário reportou item sincronizado
        antes de este campo existir — "Status escola" ficava sempre em
        branco, mesmo sincronizando de novo, porque o item já lançado
        (mesma Descrição + Quantidade) só entrava em "duplicados", sem
        atualizar nada. Sincronizar de novo agora atualiza só esse campo,
        sem duplicar o item nem tocar nos demais campos fechados."""
        item_antigo = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=2, valor_unitario=Decimal("150.00"),
            num_osp="4230", validacao_osp="Aprovado",
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2",
                num_osp="4230", validacao_osp="Aprovado", status_escola="Em implantação",
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        item_antigo.refresh_from_db()
        self.assertEqual(resultado["criados"], [])
        self.assertEqual(resultado["duplicados"], ["Nobreak"])
        self.assertEqual(item_antigo.status_escola, "Em implantação")
        self.assertEqual(RiItemRelatorioEace.objects.count(), 1)

    def test_sincronizar_de_novo_atualiza_status_escola_do_kit_ja_lancado(self):
        """RN-046 (correção, 2026-08-28): mesmo bug do teste acima, mas
        para o KIT — cai no ramo "kit_ignorado" (RN-015) antes de chegar
        no de "duplicados", e por isso nunca recebia o backfill."""
        kit_antigo = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO",
                status_escola="Em implantação",
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        kit_antigo.refresh_from_db()
        self.assertEqual(resultado["kit_ignorado"], ["Kit Cobertura Wi-Fi - 4 Access Points"])
        self.assertEqual(kit_antigo.status_escola, "Em implantação")

    def test_sincronizar_de_novo_atualiza_nota_fiscal_de_item_ja_lancado(self):
        """Usuário pediu (2026-08-28) que os 4 campos fechados (não só
        "Status Equip") sejam sempre atualizados a cada nova planilha —
        cobre o caso real de a EACE emitir a Nota Fiscal só depois de o
        item já ter sido sincronizado sem ela."""
        item_antigo = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=2, valor_unitario=Decimal("150.00"),
            num_osp="4230", validacao_osp="Aprovado",
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2",
                num_osp="4230", validacao_osp="Aprovado", nota_fiscal="381",
            ),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        item_antigo.refresh_from_db()
        self.assertEqual(item_antigo.nota_fiscal, "381")

    def test_sincronizar_de_novo_nao_apaga_campo_fechado_quando_planilha_vem_vazia(self):
        """Planilha nova com a coluna vazia não deve apagar um valor já
        gravado — falta de dado não é tratada como "backfill negativo"
        (mesmo critério conservador já usado no "Status Equip")."""
        item_antigo = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Nobreak", quantidade=2, valor_unitario=Decimal("150.00"),
            nota_fiscal="381",
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2", nota_fiscal="",
            ),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        item_antigo.refresh_from_db()
        self.assertEqual(item_antigo.nota_fiscal, "381")

    def test_tela_nao_mostra_alerta_quando_status_escola_igual(self):
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - Equip - MEGA - CO",
                status_escola="Conectada",
            ),
            _linha_planilha_eace(
                self.escola.inep, "Nobreak - Equip - MEGA - CO", status_escola="Conectada"
            ),
        ])
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        resp = self.client.get(reverse("ri_detail", kwargs={"inep": self.escola.inep}))
        self.assertNotContains(resp, "Divergência Status EACE")


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
class SincronizadorSubstituiPelaUltimaPlanilhaTests(TestCase):
    """RN-062 (2026-09-04): fora de "Implantação EACE"/"Em Andamento", a
    Planilha EACE ativa vira a fonte de verdade do Lado 3 — usuário
    reportou que uma planilha nova com menos equipamentos que a anterior
    não removia os itens que sumiram. Nesses 2 status iniciais, o
    comportamento de sempre (só soma, nunca remove/atualiza) continua —
    ver os testes desta classe que usam `Ri.ANDAMENTO`."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_SINCRONIZADOR, ignore_errors=True)

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin-sync-rn062", password="senha-teste-123",
            perfil=User.PERFIL_ADMINISTRADOR,
        )
        self.analista = User.objects.create_user(
            username="analista-sync-rn062", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )
        self.escola = Escola.objects.create(
            inep="61000001", nome="Escola RN-062", municipio="Fortaleza", estado="CE"
        )
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.AGUARDANDO_FINANCEIRO)
        self.nobreak = KitPadrao.objects.create(
            descricao="Nobreak", unidade="Unidade", valor_equipamento=Decimal("150.00"),
        )
        self.cabo = KitPadrao.objects.create(
            descricao="Cabo de rede", unidade="Metro", valor_equipamento=Decimal("2.00"),
        )

    def _upload_planilha(self, linhas):
        PlanilhaEace.substituir(_csv_planilha_eace(linhas=linhas), self.admin)

    def test_quantidade_diferente_atualiza_o_mesmo_item_em_vez_de_duplicar(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="3"),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(resultado["criados"], [])
        self.assertEqual(len(resultado["atualizados"]), 1)
        item, resumo_anterior = resultado["atualizados"][0]
        self.assertEqual(resumo_anterior, "Nobreak — 3 un. — R$ 150.00")
        self.assertEqual(item.quantidade, 2)
        # Continua 1 único item — não vira 3 + 2 = 5 (bug relatado).
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=self.ri).count(), 1)

    def test_item_ausente_na_planilha_nova_e_removido(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
            _linha_planilha_eace(self.escola.inep, "Cabo de rede - Equip - MEGA - CO", qtd="10"),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=self.ri).count(), 2)

        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["removidos"]), 1)
        descricao, quantidade, valor_unitario, eh_kit = resultado["removidos"][0]
        self.assertEqual(descricao, "Cabo de rede")
        self.assertEqual(quantidade, 10)
        self.assertFalse(eh_kit)
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=self.ri).count(), 1)
        self.assertEqual(
            RiItemRelatorioEace.objects.get(ri=self.ri).descricao_item, "Nobreak"
        )

    def test_item_lancado_manualmente_nunca_e_removido_pela_sincronizacao(self):
        """Decisão do usuário (2026-09-04): item lançado manualmente
        (`origem_sincronizador=False`) nunca é apagado por uma
        sincronização, mesmo ausente da planilha ativa."""
        manual = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Cabo de rede", quantidade=10,
            valor_unitario=Decimal("2.00"),
        )
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(resultado["removidos"], [])
        manual.refresh_from_db()  # não levanta DoesNotExist
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=self.ri).count(), 2)

    def test_item_manual_confirmado_por_planilha_passa_a_poder_ser_removido_depois(self):
        """Depois que uma planilha real confirma a mesma Descrição de um
        item lançado manualmente, ele passa a ser tratado como vindo do
        Sincronizador — uma planilha seguinte sem essa Descrição pode
        removê-lo (RN-062)."""
        manual = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Cabo de rede", quantidade=10,
            valor_unitario=Decimal("2.00"),
        )
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Cabo de rede - Equip - MEGA - CO", qtd="10"),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        manual.refresh_from_db()
        self.assertTrue(manual.origem_sincronizador)

        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["removidos"]), 1)
        self.assertEqual(RiItemRelatorioEace.objects.filter(pk=manual.pk).count(), 0)

    def test_status_implantacao_eace_preserva_comportamento_antigo(self):
        """Em "Implantação EACE"/"Em Andamento" (RN-062), a Planilha EACE
        continua só somando — quantidade diferente cria outro item em vez
        de atualizar, e nada é removido (comportamento intencionalmente
        mantido, fora do escopo desta melhoria)."""
        self.ri.status = Ri.ANDAMENTO
        self.ri.save(update_fields=["status"])
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="3"),
        ])
        sincronizar_relatorio_eace_da_planilha(self.ri)
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2"),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(len(resultado["criados"]), 1)
        self.assertEqual(resultado["atualizados"], [])
        self.assertEqual(resultado["removidos"], [])
        self.assertEqual(RiItemRelatorioEace.objects.filter(ri=self.ri).count(), 2)

    def test_kit_diferente_ignorado_nao_remove_o_kit_ja_lancado(self):
        """RN-015 continua valendo dentro do modo RN-062: uma planilha com
        um KIT diferente do já lançado cai em "kit_ignorado" (não
        substitui o KIT automaticamente) e o KIT já lançado não é
        removido por "não veio na planilha"."""
        kit_4ap = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", unidade="Escola",
            valor_equipamento=Decimal("300.00"),
        )
        KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 8 Access Points", unidade="Escola",
            valor_equipamento=Decimal("500.00"),
        )
        kit_antigo = RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("300.00"), eh_kit=True,
            origem_sincronizador=True,
        )
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola.inep, "Kit Cobertura Wi-Fi - 8 Access Points - Equip - MEGA - CO"
            ),
        ])
        resultado = sincronizar_relatorio_eace_da_planilha(self.ri)
        self.assertEqual(
            resultado["kit_ignorado"], ["Kit Cobertura Wi-Fi - 8 Access Points"]
        )
        self.assertEqual(resultado["removidos"], [])
        self.assertEqual(RiItemRelatorioEace.objects.filter(pk=kit_antigo.pk).count(), 1)

    def test_view_grava_historico_de_atualizacao_e_remocao(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="3"),
            _linha_planilha_eace(self.escola.inep, "Cabo de rede - Equip - MEGA - CO", qtd="10"),
        ])
        self.client.force_login(self.analista)
        self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
        )
        self._upload_planilha([
            _linha_planilha_eace(self.escola.inep, "Nobreak - Equip - MEGA - CO", qtd="2"),
        ])
        resp = self.client.post(
            reverse("ri_detail", kwargs={"inep": self.escola.inep}),
            {"acao": "sincronizar_planilha_eace"},
            follow=True,
        )
        mensagens = [str(m) for m in resp.context["messages"]]
        self.assertIn("1 atualizado(s)", mensagens[-1])
        self.assertIn("1 removido(s)", mensagens[-1])
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_CAMPO,
                valor_anterior="Nobreak — 3 un. — R$ 150.00",
                valor_novo__icontains="Nobreak — 2 un.",
            ).exists()
        )
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_CAMPO,
                campo__icontains="removido",
                valor_novo="Removido — não veio mais na Planilha EACE ativa",
            ).exists()
        )


@override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
class SincronizarRelatorioEaceDeTodasAsRiTests(TestCase):
    """FEAT-025/RN-023: botão "Sincronizar todas as RI" do card "Arquivo
    ativo" — aplica o Sincronizador (RN-022/FEAT-024) ao RI atual de cada
    Escola de uma vez, sem abrir RI por RI."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin-sync-lote", password="senha-teste-123",
            perfil=User.PERFIL_ADMINISTRADOR,
        )
        self.analista = User.objects.create_user(
            username="analista-sync-lote", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )
        self.kit = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", unidade="Escola",
            valor_equipamento=Decimal("300.00"), valor_servico=Decimal("50.00"),
        )
        self.escola_1 = Escola.objects.create(inep="70000001", nome="Escola Lote 1")
        self.escola_2 = Escola.objects.create(inep="70000002", nome="Escola Lote 2")
        self.ri_1 = Ri.objects.create(escola=self.escola_1, status=Ri.IMPLANTACAO_EACE)
        self.ri_2 = Ri.objects.create(escola=self.escola_2, status=Ri.IMPLANTACAO_EACE)

    def _upload_planilha(self, linhas):
        PlanilhaEace.substituir(_csv_planilha_eace(linhas=linhas), self.admin)

    def test_sincroniza_ri_atual_de_cada_escola(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
            _linha_planilha_eace(self.escola_2.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        resultados_por_ri = {ri.pk: resultado for ri, resultado in processados}
        self.assertEqual(len(resultados_por_ri[self.ri_1.pk]["criados"]), 1)
        self.assertEqual(len(resultados_por_ri[self.ri_2.pk]["criados"]), 1)

    def test_status_escola_conectada_nao_conclui_ri_no_lote(self):
        """RN-024 (retirada, 2026-09-02): "Status escola" = "Conectada"
        não conclui mais o RI, nem no Sincronizador em lote."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO",
                status_escola="Conectada",
            ),
            _linha_planilha_eace(self.escola_2.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        sincronizar_relatorio_eace_de_todas_as_ri()
        self.ri_1.refresh_from_db()
        self.assertEqual(self.ri_1.status, Ri.IMPLANTACAO_EACE)
        self.assertIsNone(self.ri_1.concluido_em)

    def test_status_escola_gravado_por_item_tambem_no_lote(self):
        """RN-046: mesma gravação por item vale para o lote — não é lógica
        separada, é a mesma `sincronizar_relatorio_eace_da_planilha`."""
        self._upload_planilha([
            _linha_planilha_eace(
                self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO",
                status_escola="Em implantação",
            ),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        resultados_por_ri = {ri.pk: resultado for ri, resultado in processados}
        item = resultados_por_ri[self.ri_1.pk]["criados"][0]
        self.assertEqual(item.status_escola, "Em implantação")

    def test_usa_ri_mais_recente_quando_escola_tem_mais_de_um(self):
        ri_antigo = self.ri_1
        ri_recente = Ri.objects.create(escola=self.escola_1, status=Ri.IMPLANTACAO_EACE)
        self._upload_planilha([
            _linha_planilha_eace(self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        ris_processados = [ri.pk for ri, _ in processados if ri.escola_id == self.escola_1.pk]
        self.assertEqual(ris_processados, [ri_recente.pk])
        self.assertNotIn(ri_antigo.pk, ris_processados)

    def test_ri_bloqueado_por_faturamento_concluido_e_pulado(self):
        self.ri_1.status = Ri.FATURAMENTO_CONCLUIDO
        self.ri_1.save(update_fields=["status"])
        self._upload_planilha([
            _linha_planilha_eace(self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        resultados_por_ri = {ri.pk: resultado for ri, resultado in processados}
        self.assertEqual(resultados_por_ri[self.ri_1.pk], RI_BLOQUEADO_FATURAMENTO_CONCLUIDO)
        self.assertFalse(self.ri_1.itens_relatorio_eace.exists())

    def test_ri_sem_linha_na_planilha_marcado_no_resumo_sem_travar_os_demais(self):
        self._upload_planilha([
            _linha_planilha_eace(self.escola_2.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        resultados_por_ri = {ri.pk: resultado for ri, resultado in processados}
        self.assertEqual(resultados_por_ri[self.ri_1.pk], RI_SEM_LINHA_NA_PLANILHA)
        self.assertEqual(len(resultados_por_ri[self.ri_2.pk]["criados"]), 1)

    def test_sem_planilha_ativa_levanta_erro(self):
        with self.assertRaises(PlanilhaEaceSincronizacaoError):
            sincronizar_relatorio_eace_de_todas_as_ri()

    def test_escola_sem_ri_ganha_ri_novo_quando_tem_linha_na_planilha(self):
        """RN-049 (2026-09-02): Escola sem nenhum RI ainda, mas com linha
        na Planilha EACE ativa para o INEP dela, ganha um RI novo aqui —
        nasce em "Implantação EACE" (mesmo status do "Iniciar RI" manual)
        e já processa o item do Lado Relatório EACE nele, na mesma
        passada."""
        escola_sem_ri = Escola.objects.create(inep="70000009", nome="Escola Sem RI Ainda")
        self._upload_planilha([
            _linha_planilha_eace(escola_sem_ri.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        processados = sincronizar_relatorio_eace_de_todas_as_ri()
        ri_novo = Ri.objects.get(escola=escola_sem_ri)
        self.assertEqual(ri_novo.status, Ri.IMPLANTACAO_EACE)
        resultados_por_ri = {ri.pk: resultado for ri, resultado in processados}
        self.assertEqual(len(resultados_por_ri[ri_novo.pk]["criados"]), 1)

    def test_escola_sem_ri_e_sem_linha_na_planilha_continua_sem_ri(self):
        """RN-049: sem nenhuma linha pra sincronizar, não cria RI vazio à
        toa (CLAUDE.md §9) — mesmo comportamento de antes desta regra."""
        escola_sem_ri = Escola.objects.create(inep="70000009", nome="Escola Sem RI Ainda")
        self._upload_planilha([
            _linha_planilha_eace(self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        sincronizar_relatorio_eace_de_todas_as_ri()
        self.assertFalse(Ri.objects.filter(escola=escola_sem_ri).exists())

    def test_nao_gera_consulta_proporcional_ao_numero_de_escolas(self):
        """Mesmo prefetch do grid (FEAT-007): número de consultas não
        cresce junto com o número de Escolas/RIs processados."""
        for numero in range(3, 8):
            escola = Escola.objects.create(inep=f"7000000{numero}", nome=f"Escola Lote {numero}")
            Ri.objects.create(escola=escola, status=Ri.IMPLANTACAO_EACE)
        self._upload_planilha([
            _linha_planilha_eace(self.escola_1.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
        ])
        # 9, não 11 (RN-003 ajustada em 2026-09-02): o KIT lançado aqui só
        # entra no Lado Relatório EACE — Lado IXC vazio não gera mais
        # `RiDivergencia` (1 INSERT a menos que antes da correção). RN-062
        # (2026-09-04) tira mais 1 consulta: "já existe KIT lançado?"
        # (RN-015) passa a ser calculado a partir da lista de itens já
        # carregada, em vez de um `.filter(eh_kit=True).exists()` à parte.
        with self.assertNumQueries(9):
            sincronizar_relatorio_eace_de_todas_as_ri()


class PlanilhaEaceSincronizarTodasViewTests(TestCase):
    """FEAT-025/RN-023: view do botão "Sincronizar todas as RI"."""

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(_MEDIA_ROOT_TESTE_SINCRONIZADOR, ignore_errors=True)

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin-view-lote", password="senha-teste-123",
            perfil=User.PERFIL_ADMINISTRADOR,
        )
        self.analista = User.objects.create_user(
            username="analista-view-lote", password="senha-teste-123",
            perfil=User.PERFIL_ANALISTA,
        )
        self.kit = KitPadrao.objects.create(
            descricao="Kit Cobertura Wi-Fi - 4 Access Points", unidade="Escola",
            valor_equipamento=Decimal("300.00"), valor_servico=Decimal("50.00"),
        )
        self.escola = Escola.objects.create(inep="71000001", nome="Escola View Lote")
        self.ri = Ri.objects.create(escola=self.escola, status=Ri.IMPLANTACAO_EACE)

    @override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
    def test_administrador_dispara_sincronizacao_e_gera_log(self):
        PlanilhaEace.substituir(
            _csv_planilha_eace(linhas=[
                _linha_planilha_eace(self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
            ]),
            self.admin,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"))
        self.assertRedirects(resp, reverse("planilha_eace"))
        self.assertEqual(self.ri.itens_relatorio_eace.count(), 1)
        self.assertTrue(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_CAMPO, campo__icontains="Sincronizador em lote"
            ).exists()
        )

    @override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
    def test_sem_item_novo_nao_e_tratado_como_erro(self):
        """Usuário reportou (2026-08-27) que a mensagem aparecia em
        vermelho (estilo de erro) mesmo com a sincronização tendo rodado
        certinho — "0 item novo" é normal no lote (RI sem linha na
        planilha, ou já sincronizado antes), não uma falha."""
        PlanilhaEace.substituir(
            _csv_planilha_eace(linhas=[
                _linha_planilha_eace("99999999", "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
            ]),
            self.admin,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"), follow=True)
        mensagens = list(resp.context["messages"])
        self.assertEqual(len(mensagens), 1)
        self.assertEqual(mensagens[0].tags, "success")
        self.assertIn("0 INEP(s) atualizado(s)", str(mensagens[0]))

    @override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
    def test_mensagem_mostra_so_a_contagem_de_ineps_atualizados(self):
        """RN-023 (ajustada em 2026-08-27): usuário pediu só a contagem —
        sem o detalhamento por INEP (já sincronizado, bloqueado, sem
        linha na planilha) que a mensagem trazia antes."""
        outra_escola = Escola.objects.create(inep="71000002", nome="Escola View Lote 2")
        Ri.objects.create(escola=outra_escola, status=Ri.FATURAMENTO_CONCLUIDO)
        PlanilhaEace.substituir(
            _csv_planilha_eace(linhas=[
                _linha_planilha_eace(self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
                _linha_planilha_eace(outra_escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO"),
            ]),
            self.admin,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"), follow=True)
        mensagens = [str(m) for m in resp.context["messages"]]
        self.assertEqual(mensagens, ["Sincronização em lote: 1 INEP(s) atualizado(s)."])

    def test_analista_nao_acessa(self):
        self.client.force_login(self.analista)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"))
        self.assertEqual(resp.status_code, 403)

    def test_get_nao_processa(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("planilha_eace_sincronizar_todas"))
        self.assertRedirects(resp, reverse("planilha_eace"))
        self.assertEqual(self.ri.itens_relatorio_eace.count(), 0)

    def test_sem_planilha_ativa_mostra_mensagem_de_erro(self):
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"), follow=True)
        mensagens = [str(m) for m in resp.context["messages"]]
        self.assertTrue(any("Nenhuma Planilha EACE ativa" in m for m in mensagens))

    @override_settings(MEDIA_ROOT=_MEDIA_ROOT_TESTE_SINCRONIZADOR)
    def test_status_escola_conectada_nao_entra_na_contagem_sem_item_novo(self):
        """RN-024 (retirada, 2026-09-02): sem item novo (já lançado antes)
        e sem mais troca de status, o INEP não entra na contagem de
        "atualizados" só por causa de "Status escola" = "Conectada"."""
        RiItemRelatorioEace.objects.create(
            ri=self.ri, descricao_item="Kit Cobertura Wi-Fi - 4 Access Points",
            quantidade=1, valor_unitario=Decimal("350.00"), eh_kit=True,
        )
        PlanilhaEace.substituir(
            _csv_planilha_eace(linhas=[
                _linha_planilha_eace(
                    self.escola.inep, "Kit Cobertura Wi-Fi - 4 Access Points - MEGA - CO",
                    status_escola="Conectada",
                ),
            ]),
            self.admin,
        )
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("planilha_eace_sincronizar_todas"), follow=True)
        mensagens = [str(m) for m in resp.context["messages"]]
        self.assertEqual(mensagens, ["Sincronização em lote: 0 INEP(s) atualizado(s)."])
        self.ri.refresh_from_db()
        self.assertEqual(self.ri.status, Ri.IMPLANTACAO_EACE)
        self.assertIsNone(self.ri.concluido_em)
        self.assertFalse(
            RiHistorico.objects.filter(
                ri=self.ri, tipo=RiHistorico.LOG_STATUS, campo="Status do RI (Sincronizador)"
            ).exists()
        )


class DashboardFinanceiroTests(TestCase):
    """FEAT-026 (RN-025/RN-026): cards do dashboard financeiro."""

    def setUp(self):
        # Catálogo (LPU), Lote 9: Kit = 1.500,00 (1.000 + 500); Nobreak = 400,00 (300 + 100).
        KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=9, unidade="Escola",
            valor_equipamento="1000.00", valor_servico="500.00",
        )
        KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=9, unidade="Unidade",
            valor_equipamento="300.00", valor_servico="100.00",
        )

    def test_soma_kit_e_nobreak_de_todas_as_escolas(self):
        """RN-025: cada escola contribui com Kit (RN-010) + Nobreak
        (RN-017 corrigida); 2 escolas iguais somam o dobro."""
        Escola.objects.create(inep="10000001", nome="Escola A", kit_inicial="Kit Wi-Fi Indoor", lote=9)
        Escola.objects.create(inep="10000002", nome="Escola B", kit_inicial="Kit Wi-Fi Indoor", lote=9)
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_total_projeto"], Decimal("3800.00"))

    def test_escola_sem_correspondencia_no_catalogo_conta_zero(self):
        """RN-025: escola sem Kit/Nobreak no catálogo não trava o
        dashboard — contribui com R$ 0,00 (opção conservadora)."""
        Escola.objects.create(
            inep="10000003", nome="Escola Sem Catálogo", kit_inicial="Kit Inexistente", lote=9,
            nobreak_inicial="",
        )
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_total_projeto"], Decimal("0"))

    def test_valor_faturado_soma_so_ri_em_faturamento_concluido(self):
        """RN-026: item do Lado 3 de RI que não está em 'Faturamento
        Concluído' não entra na soma, mesmo já lançado."""
        escola_concluida = Escola.objects.create(inep="10000004", nome="Escola Concluída", lote=9)
        escola_andamento = Escola.objects.create(inep="10000005", nome="Escola Andamento", lote=9)
        ri_concluido = Ri.objects.create(escola=escola_concluida, status=Ri.FATURAMENTO_CONCLUIDO)
        ri_andamento = Ri.objects.create(escola=escola_andamento, status=Ri.ANDAMENTO)
        RiItemRelatorioEace.objects.create(
            ri=ri_concluido, descricao_item="Kit Wi-Fi Indoor", quantidade=2, valor_unitario="250.00",
        )
        RiItemRelatorioEace.objects.create(
            ri=ri_andamento, descricao_item="Kit Wi-Fi Indoor", quantidade=5, valor_unitario="999.00",
        )
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_faturado"], Decimal("500.00"))

    def test_diferenca_e_percentual_quando_falta_faturar(self):
        Escola.objects.create(inep="10000006", nome="Escola A", kit_inicial="Kit Wi-Fi Indoor", lote=9)
        escola_ri = Escola.objects.create(inep="10000007", nome="Escola B", nobreak_inicial="", lote=9)
        ri = Ri.objects.create(escola=escola_ri, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Produto X", quantidade=1, valor_unitario="500.00",
        )
        resultado = montar_dashboard_financeiro()
        # total_projeto = 1900,00 (escola A: 1500+400) + 0 (escola B sem kit/nobreak) = 1900,00
        self.assertEqual(resultado["valor_total_projeto"], Decimal("1900.00"))
        self.assertEqual(resultado["valor_faturado"], Decimal("500.00"))
        self.assertFalse(resultado["meta_atingida"])
        self.assertEqual(resultado["valor_faltante"], Decimal("1400.00"))
        self.assertEqual(resultado["valor_excedente"], Decimal("0"))

    def test_meta_atingida_quando_faturado_ultrapassa_o_total(self):
        escola_ri = Escola.objects.create(inep="10000008", nome="Escola C", nobreak_inicial="", lote=9)
        ri = Ri.objects.create(escola=escola_ri, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="1500.00",
        )
        # Sem nenhuma Escola com Kit/Nobreak resolvido: valor_total_projeto = 0,00.
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_total_projeto"], Decimal("0"))
        self.assertEqual(resultado["valor_faturado"], Decimal("1500.00"))
        self.assertTrue(resultado["meta_atingida"])
        self.assertEqual(resultado["valor_faltante"], Decimal("0"))
        self.assertEqual(resultado["valor_excedente"], Decimal("1500.00"))
        self.assertEqual(resultado["percentual_faturado_pct"], Decimal("100"))
        self.assertEqual(resultado["percentual_faltante_pct"], Decimal("0"))

    def test_percentual_faturado_passa_de_100_quando_ha_meta_real(self):
        """Correção (2026-08-27, pedido do usuário): com uma meta real (>0)
        e faturado acima dela, o texto/badge mostra mais de 100% — só a
        barra de 2 segmentos do card (CSS) fica capada em 100%."""
        escola = Escola.objects.create(
            inep="10000009", nome="Escola D", kit_inicial="Kit Wi-Fi Indoor", nobreak_inicial="", lote=9,
        )
        ri = Ri.objects.create(escola=escola, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="3000.00",
        )
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_total_projeto"], Decimal("1500.00"))  # meta real
        self.assertEqual(resultado["valor_faturado"], Decimal("3000.00"))
        self.assertEqual(resultado["percentual_faturado_pct"], Decimal("200.00"))  # 3.000/1.500
        self.assertEqual(resultado["percentual_faturado_css"], "100.00")  # barra capada
        self.assertEqual(resultado["percentual_faltante_css"], "0.00")


class DashboardFinanceiroPorEstadoTests(TestCase):
    """FEAT-026 ampliada (RN-027): filtro por estado nos 2 cards e gráfico
    "Faturado por Estado"."""

    def setUp(self):
        # Catálogo (LPU), Lote 9: Kit = 1.500,00 (1.000 + 500); Nobreak = 400,00 (300 + 100).
        KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=9, unidade="Escola",
            valor_equipamento="1000.00", valor_servico="500.00",
        )
        KitPadrao.objects.create(
            descricao="Nobreak (serviço, material, equipamento)", lote=9, unidade="Unidade",
            valor_equipamento="300.00", valor_servico="100.00",
        )
        self.escola_sp = Escola.objects.create(
            inep="20000001", nome="Escola SP", kit_inicial="Kit Wi-Fi Indoor", lote=9, estado="SP",
        )  # Kit 1.500 + Nobreak 400 (padrão) = 1.900,00
        self.escola_rj = Escola.objects.create(
            inep="20000002", nome="Escola RJ", kit_inicial="Kit Wi-Fi Indoor", lote=9, estado="RJ",
            nobreak_inicial="",
        )  # Kit 1.500 + 0 (sem Nobreak) = 1.500,00
        ri_sp = Ri.objects.create(escola=self.escola_sp, status=Ri.FATURAMENTO_CONCLUIDO)
        ri_rj = Ri.objects.create(escola=self.escola_rj, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri_sp, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="700.00",
        )
        RiItemRelatorioEace.objects.create(
            ri=ri_rj, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="300.00",
        )

    def test_sem_estado_soma_todas_as_escolas_e_ris(self):
        resultado = montar_dashboard_financeiro()
        self.assertEqual(resultado["valor_total_projeto"], Decimal("3400.00"))
        self.assertEqual(resultado["valor_faturado"], Decimal("1000.00"))
        self.assertIsNone(resultado["estado_filtrado"])

    def test_estado_filtra_os_2_cards_por_uf(self):
        resultado_sp = montar_dashboard_financeiro(estado="SP")
        self.assertEqual(resultado_sp["valor_total_projeto"], Decimal("1900.00"))
        self.assertEqual(resultado_sp["valor_faturado"], Decimal("700.00"))
        self.assertEqual(resultado_sp["estado_filtrado"], "SP")

        resultado_rj = montar_dashboard_financeiro(estado="RJ")
        self.assertEqual(resultado_rj["valor_total_projeto"], Decimal("1500.00"))
        self.assertEqual(resultado_rj["valor_faturado"], Decimal("300.00"))

    def test_estado_sem_nenhuma_escola_ou_ri_conta_zero(self):
        resultado = montar_dashboard_financeiro(estado="CE")
        self.assertEqual(resultado["valor_total_projeto"], Decimal("0"))
        self.assertEqual(resultado["valor_faturado"], Decimal("0"))

    def test_faturamento_por_estado_ordena_do_maior_para_o_menor(self):
        """Ordena pela % da meta já faturada (não pelo valor bruto); cada
        linha também traz a meta (Kit + Nobreak) daquele estado."""
        linhas = montar_faturamento_por_estado()
        self.assertEqual([linha["estado"] for linha in linhas], ["SP", "RJ"])
        self.assertEqual(linhas[0]["valor"], Decimal("700.00"))
        self.assertEqual(linhas[0]["meta"], Decimal("1900.00"))  # SP: Kit 1.500 + Nobreak 400
        self.assertEqual(linhas[1]["valor"], Decimal("300.00"))
        self.assertEqual(linhas[1]["meta"], Decimal("1500.00"))  # RJ: só Kit (sem Nobreak)

    def test_faturamento_por_estado_percentual_css_usa_ponto_decimal(self):
        """A barra é proporcional ao quanto da META do próprio estado já
        foi faturado (não ao valor bruto comparado entre estados)."""
        linhas = montar_faturamento_por_estado()
        self.assertEqual(linhas[0]["percentual_css"], "36.84")  # SP: 700 / 1.900 * 100
        self.assertEqual(linhas[1]["percentual_css"], "20.00")  # RJ: 300 / 1.500 * 100

    def test_estado_com_escola_mas_sem_faturamento_aparece_com_zero(self):
        Escola.objects.create(inep="20000003", nome="Escola MG", estado="MG", nobreak_inicial="")
        linhas = montar_faturamento_por_estado()
        linha_mg = next(linha for linha in linhas if linha["estado"] == "MG")
        self.assertEqual(linha_mg["valor"], Decimal("0"))
        self.assertEqual(linha_mg["meta"], Decimal("0"))  # sem Kit/Nobreak resolvido
        self.assertEqual(linha_mg["percentual_css"], "0.00")

    def test_escola_sem_estado_nao_entra_no_grafico(self):
        Escola.objects.create(inep="20000004", nome="Escola Sem UF", nobreak_inicial="")
        linhas = montar_faturamento_por_estado()
        self.assertNotIn("", [linha["estado"] for linha in linhas])
        self.assertEqual(len(linhas), 2)  # só SP e RJ, do setUp

    def test_ordena_pela_porcentagem_nao_pelo_valor_bruto(self):
        """Pedido do usuário (2026-08-27): do que está mais perto de 100%
        para o que tem 0% — não do maior valor em R$ para o menor. Um
        estado com valor bruto MAIOR mas % MENOR deve vir por último."""
        KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=11, unidade="Escola",
            valor_equipamento="18000.00", valor_servico="2000.00",  # meta = 20.000,00
        )
        escola_go = Escola.objects.create(
            inep="20000005", nome="Escola GO", kit_inicial="Kit Wi-Fi Indoor", lote=11,
            estado="GO", nobreak_inicial="",
        )  # meta = 20.000,00
        ri_go = Ri.objects.create(escola=escola_go, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri_go, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="1000.00",
        )  # GO: valor bruto 1.000 — maior que SP (700) e RJ (300) —, mas só 5% da meta

        linhas = montar_faturamento_por_estado()
        # Por valor bruto seria GO (1.000) > SP (700) > RJ (300); pela % (o
        # critério real) é SP (36,84%) > RJ (20,00%) > GO (5,00%, por último
        # apesar do maior valor bruto).
        self.assertEqual(
            [linha["estado"] for linha in linhas], ["SP", "RJ", "GO"]
        )


class DashboardFaturamentoPorMunicipioTests(TestCase):
    """FEAT-026 ampliada (RN-027 ampliada): drill-down por Município dentro
    de um estado — pedido do usuário para o gráfico "expandir" ao clicar
    num estado, mostrando os municípios com a mesma informação."""

    def setUp(self):
        KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=9, unidade="Escola",
            valor_equipamento="1000.00", valor_servico="500.00",
        )
        self.escola_campinas = Escola.objects.create(
            inep="40000001", nome="Escola Campinas", kit_inicial="Kit Wi-Fi Indoor", lote=9,
            estado="SP", municipio="Campinas", nobreak_inicial="",
        )
        self.escola_santos = Escola.objects.create(
            inep="40000002", nome="Escola Santos", kit_inicial="Kit Wi-Fi Indoor", lote=9,
            estado="SP", municipio="Santos", nobreak_inicial="",
        )
        self.escola_rj = Escola.objects.create(
            inep="40000003", nome="Escola RJ", kit_inicial="Kit Wi-Fi Indoor", lote=9,
            estado="RJ", municipio="Campinas", nobreak_inicial="",  # município homônimo, outro estado
        )
        ri_campinas = Ri.objects.create(escola=self.escola_campinas, status=Ri.FATURAMENTO_CONCLUIDO)
        ri_santos = Ri.objects.create(escola=self.escola_santos, status=Ri.FATURAMENTO_CONCLUIDO)
        ri_rj = Ri.objects.create(escola=self.escola_rj, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri_campinas, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="900.00",
        )
        RiItemRelatorioEace.objects.create(
            ri=ri_santos, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="150.00",
        )
        RiItemRelatorioEace.objects.create(
            ri=ri_rj, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="1500.00",
        )

    def test_sem_estado_retorna_lista_vazia(self):
        self.assertEqual(montar_faturamento_por_municipio(None), [])
        self.assertEqual(montar_faturamento_por_municipio(""), [])

    def test_agrupa_so_dentro_do_estado_informado(self):
        """Campinas existe em SP e em RJ (homônimo) — só entra o de SP."""
        linhas = montar_faturamento_por_municipio("SP")
        self.assertEqual(
            {linha["municipio"] for linha in linhas}, {"Campinas", "Santos"}
        )
        campinas = next(linha for linha in linhas if linha["municipio"] == "Campinas")
        self.assertEqual(campinas["valor"], Decimal("900.00"))  # não soma o Campinas/RJ (1.500)
        self.assertEqual(campinas["meta"], Decimal("1500.00"))  # Kit, sem Nobreak

    def test_ordena_pela_maior_porcentagem_da_meta(self):
        linhas = montar_faturamento_por_municipio("SP")
        self.assertEqual([linha["municipio"] for linha in linhas], ["Campinas", "Santos"])

    def test_municipio_ordena_pela_porcentagem_nao_pelo_valor_bruto(self):
        """Mesmo critério do gráfico de Estado (2026-08-27): um município
        com valor bruto MAIOR mas % MENOR da própria meta vem por último."""
        KitPadrao.objects.create(
            descricao="Kit Wi-Fi Indoor", lote=11, unidade="Escola",
            valor_equipamento="18000.00", valor_servico="2000.00",  # meta = 20.000,00
        )
        escola_sorocaba = Escola.objects.create(
            inep="40000004", nome="Escola Sorocaba", kit_inicial="Kit Wi-Fi Indoor", lote=11,
            estado="SP", municipio="Sorocaba", nobreak_inicial="",
        )  # meta = 20.000,00
        ri_sorocaba = Ri.objects.create(escola=escola_sorocaba, status=Ri.FATURAMENTO_CONCLUIDO)
        RiItemRelatorioEace.objects.create(
            ri=ri_sorocaba, descricao_item="Kit Wi-Fi Indoor", quantidade=1, valor_unitario="1000.00",
        )  # valor bruto 1.000 — maior que Campinas (900) —, mas só 5% da meta de 20.000

        linhas = montar_faturamento_por_municipio("SP")
        # Por valor bruto seria Sorocaba (1.000) > Campinas (900) > Santos
        # (150); pela % (o critério real) é Campinas (60,00%) > Santos
        # (10,00%) > Sorocaba (5,00%, por último apesar do maior valor bruto).
        self.assertEqual(
            [linha["municipio"] for linha in linhas], ["Campinas", "Santos", "Sorocaba"]
        )

    def test_municipio_filtra_os_2_cards_junto_com_o_estado(self):
        resultado = montar_dashboard_financeiro(estado="SP", municipio="Santos")
        self.assertEqual(resultado["valor_total_projeto"], Decimal("1500.00"))
        self.assertEqual(resultado["valor_faturado"], Decimal("150.00"))
        self.assertEqual(resultado["municipio_filtrado"], "Santos")

    def test_municipio_sem_estado_e_ignorado_pelo_service(self):
        """RN-027 ampliada: município só é aplicado junto com estado (nome
        se repete entre UFs) — a view já impede isso, mas o service também
        não aplica o filtro sozinho, por segurança."""
        resultado = montar_dashboard_financeiro(municipio="Santos")
        self.assertEqual(resultado["valor_total_projeto"], Decimal("4500.00"))  # todas as escolas
        self.assertIsNone(resultado["municipio_filtrado"])


class AcessoLiberadoBloqueiaTelasDoRiTests(TestCase):
    """FEAT-029 (RN-045): o aviso de "aguardando liberação" também cobre
    as telas do app `ri` (Grid de INEPs e detalhe do RI) — o bloqueio é
    aplicado por middleware, mas o critério de aceite pede pelo menos 1
    tela verificada em cada área do sistema."""

    def setUp(self):
        self.desligado = User.objects.create_user(
            username="desligado-teste", password="senha-teste-123", acesso_liberado=False,
        )
        self.escola = Escola.objects.create(inep="90000001", nome="Escola Teste", estado="SP")
        Ri.objects.create(escola=self.escola)
        self.client.force_login(self.desligado)

    def test_grid_de_ineps_mostra_o_aviso_em_vez_da_lista(self):
        resp = self.client.get(reverse("grid_inep"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Aguardando liberação do Administrador")
        self.assertNotContains(resp, "Escola Teste")

    def test_detalhe_do_ri_mostra_o_aviso_em_vez_dos_dados(self):
        resp = self.client.get(reverse("ri_detail", args=["90000001"]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Aguardando liberação do Administrador")
        self.assertNotContains(resp, "Escola Teste")
